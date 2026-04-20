"""routers/benchmark.py — Phase 4 Step 4D refactoring."""

from fastapi import APIRouter
from fastapi import Depends, HTTPException, UploadFile, File, Form
from core.deps import get_current_user, require_admin
from core.config import FRED_API_KEY, ANTHROPIC_KEY, FB_URL
from core.firebase import fb_read, fb_put, fb_auth_param
import requests
import json
import datetime
import os
import tempfile
from pyxlsb import open_workbook



# ── FRED 시계열 설정 ──
FRED_SERIES = {
    # 금리 (할인율 기준점)
    "us_10y":       {"id": "DGS10",        "label": "US 10Y Treasury",    "unit": "%",       "group": "rates"},
    "us_2y":        {"id": "DGS2",         "label": "US 2Y Treasury",     "unit": "%",       "group": "rates"},
    "fed_funds":    {"id": "DFF",          "label": "Fed Funds Rate",     "unit": "%",       "group": "rates"},
    "bbb_spread":   {"id": "BAMLC0A4CBBB", "label": "BBB Corp Spread",    "unit": "%",       "group": "rates"},
    # 에너지/인플레
    "henry_hub":    {"id": "DHHNGSP",      "label": "Henry Hub NatGas",   "unit": "$/MMBtu", "group": "energy"},
    "cpi":          {"id": "CPIAUCSL",     "label": "US CPI (Index)",     "unit": "Index",   "group": "macro"},
    # 환율
    "krw_usd":      {"id": "DEXKOUS",      "label": "KRW/USD",            "unit": "KRW",     "group": "fx"},
}

# ── Stooq 시계열 설정 ──
STOOQ_SYMBOLS = {
    "tan":          {"symbol": "tan.us",   "label": "TAN (Solar ETF)",    "unit": "$",       "group": "equity"},
    "icln":         {"symbol": "icln.us",  "label": "ICLN (Clean Energy)", "unit": "$",      "group": "equity"},
}

router = APIRouter()


def _fred_fetch(series_id: str, days: int = 180):
    """FRED API에서 시리즈 데이터 조회. 최근 N일치."""
    if not FRED_API_KEY:
        return None
    try:
        end = datetime.date.today()
        start = end - datetime.timedelta(days=days)
        res = requests.get(
            "https://api.stlouisfed.org/fred/series/observations",
            params={
                "series_id": series_id,
                "api_key": FRED_API_KEY,
                "file_type": "json",
                "observation_start": start.isoformat(),
                "observation_end": end.isoformat(),
                "sort_order": "asc",
            },
            timeout=10,
        )
        if res.status_code != 200:
            return None
        obs = res.json().get("observations", [])
        # "." = 결측, 제외
        points = [
            {"date": o["date"], "value": float(o["value"])}
            for o in obs if o.get("value") not in (".", "", None)
        ]
        return points
    except Exception:
        return None


def _stooq_fetch(symbol: str, days: int = 180):
    """Stooq에서 일별 종가 조회 (CSV, 키 불필요)."""
    try:
        url = f"https://stooq.com/q/d/l/?s={symbol}&i=d"
        res = requests.get(url, timeout=10)
        if res.status_code != 200 or "Date,Open" not in res.text:
            return None
        lines = res.text.strip().split("\n")[1:]
        end = datetime.date.today()
        start = end - datetime.timedelta(days=days)
        points = []
        for ln in lines:
            parts = ln.split(",")
            if len(parts) < 5:
                continue
            try:
                d = datetime.date.fromisoformat(parts[0])
                if d < start:
                    continue
                points.append({"date": parts[0], "value": float(parts[4])})
            except Exception:
                continue
        return points
    except Exception:
        return None


def _summarize_series(points):
    """시계열 → 최신값/변동/스파크라인 요약 (1년치)."""
    if not points:
        return None
    latest = points[-1]
    prev = points[-2] if len(points) >= 2 else latest
    # 약 일주일 전 (영업일 5개 전)
    week_ago = points[-6] if len(points) >= 6 else points[0]
    # 약 한달 전 (영업일 21개 전)
    month_ago = points[-22] if len(points) >= 22 else points[0]
    # 약 1년 전 (영업일 252개 전)
    year_ago = points[-253] if len(points) >= 253 else points[0]
    # 1년치 주간 샘플링 (52포인트 내외) — 주 1회 데이터만 추출
    # 영업일 기준 5일마다 1개 선택
    sampled = points[::5] if len(points) > 60 else points
    return {
        "latest": latest["value"],
        "latest_date": latest["date"],
        "d_1d": latest["value"] - prev["value"],
        "d_1w": latest["value"] - week_ago["value"],
        "d_1m": latest["value"] - month_ago["value"],
        "d_1y": latest["value"] - year_ago["value"],
        "spark": [p["value"] for p in sampled],
        "spark_dates": [p["date"] for p in sampled],  # 실제 날짜 병행 전달
        "n_points": len(points),
    }


@router.get("/benchmark/market")
def get_market_benchmark(force: int = 0, user=Depends(get_current_user)):
    """FRED + Stooq 시장 벤치마크. 6시간 캐시."""
    today = datetime.date.today().isoformat()
    cache_key = f"benchmark_cache/market/{today}"

    # 캐시 확인 (force=1이면 무시)
    if not force:
        cached = fb_read(cache_key)
        if cached and cached.get("fetched_at"):
            try:
                fetched = datetime.datetime.fromisoformat(cached["fetched_at"])
                age_hrs = (datetime.datetime.utcnow() - fetched).total_seconds() / 3600
                if age_hrs < 6:
                    return cached
            except Exception:
                pass

    if not FRED_API_KEY:
        raise HTTPException(500, "FRED_API_KEY 환경변수 미설정")

    result = {
        "fetched_at": datetime.datetime.utcnow().isoformat()[:19],
        "source": "FRED + Stooq",
        "series": {},
    }

    # FRED 시리즈 (1년치)
    for key, meta in FRED_SERIES.items():
        pts = _fred_fetch(meta["id"], days=400 if key == "cpi" else 365)
        summary = _summarize_series(pts) if pts else None
        result["series"][key] = {
            **meta,
            "data": summary,
            "ok": summary is not None,
        }

    # Stooq 시리즈 (1년치)
    for key, meta in STOOQ_SYMBOLS.items():
        pts = _stooq_fetch(meta["symbol"], days=365)
        summary = _summarize_series(pts) if pts else None
        result["series"][key] = {
            **meta,
            "data": summary,
            "ok": summary is not None,
        }

    # CPI는 YoY % 변화율로 계산 (Index 자체는 의미가 없음)
    cpi = result["series"].get("cpi", {}).get("data")
    if cpi and cpi.get("spark") and len(cpi["spark"]) >= 13:
        try:
            yoy = (cpi["spark"][-1] / cpi["spark"][-13] - 1) * 100
            result["series"]["cpi"]["yoy_pct"] = round(yoy, 2)
        except Exception:
            pass

    # 캐시 저장
    try:
        fb_put(cache_key, result)
    except Exception:
        pass  # 캐시 실패해도 응답은 반환
    return result


# ── LevelTen PPA Index 업로드/조회 ─────────────────
@router.post("/benchmark/levelten/upload")
async def upload_levelten(
    file: UploadFile = File(...),
    quarter: str = Form(...),  # e.g. "2026-Q1"
    user=Depends(get_current_user),
):
    """LevelTen PPA Index 리포트 업로드 → Claude API로 파싱 → Firebase 저장."""
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 환경변수 미설정")

    raw = await file.read()
    filename = file.filename or "levelten.pdf"
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    # 1) 파일 타입별 텍스트 추출
    source_text = ""
    parse_mode = ""
    parsed = None

    if ext in ("csv", "tsv"):
        try:
            source_text = raw.decode("utf-8", errors="ignore")
        except Exception:
            source_text = raw.decode("latin-1", errors="ignore")
        parse_mode = "csv"

    elif ext in ("xlsx", "xls", "xlsb"):
        parse_mode = "excel"
        tmp_path = None
        try:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}")
            tmp_path = tmp.name
            tmp.write(raw)
            tmp.close()

            if ext == "xlsb":
                lines = []
                with open_workbook(tmp_path) as wb:
                    for sheet_name in wb.sheets[:5]:
                        lines.append(f"\n=== Sheet: {sheet_name} ===")
                        with wb.get_sheet(sheet_name) as sh:
                            for row in sh.rows():
                                vals = [str(c.v) if c.v is not None else "" for c in row]
                                lines.append("\t".join(vals))
                source_text = "\n".join(lines)[:40000]
            else:
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    raise HTTPException(400, "openpyxl이 설치되지 않았습니다. requirements.txt에 추가하세요.")
                wb = load_workbook(tmp_path, data_only=True, read_only=True)
                lines = []
                for sheet_name in wb.sheetnames[:5]:
                    lines.append(f"\n=== Sheet: {sheet_name} ===")
                    for row in wb[sheet_name].iter_rows(values_only=True):
                        vals = [str(v) if v is not None else "" for v in row]
                        lines.append("\t".join(vals))
                wb.close()
                source_text = "\n".join(lines)[:40000]
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Excel 파싱 실패: {str(e)}")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try: os.unlink(tmp_path)
                except Exception: pass

    elif ext == "pdf":
        parse_mode = "pdf"
        import base64
        pdf_b64 = base64.standard_b64encode(raw).decode("utf-8")

        headers = {
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        prompt = (
            "You are parsing a LevelTen Energy PPA Price Index report for a Solar+BESS developer. "
            "Extract structured data into strict JSON (no markdown, no prose).\n\n"
            "IMPORTANT PRINCIPLES:\n"
            "1. We focus on Solar PPA and BESS Storage pricing. Skip Wind data entirely.\n"
            "2. ONLY extract values that are explicitly present in the report — tables, charts, or text.\n"
            "3. DO NOT estimate, guess, or use general market knowledge to fill missing values.\n"
            "4. If a value is not in the report, use null. It is BETTER to return null than to invent data.\n"
            "5. For chart-read values (Storage Price Spreads typically shown as charts only), estimate to nearest $0.5 and mark source as 'chart_read'.\n\n"

            "Required schema:\n"
            "{\n"
            '  "quarter": "YYYY-QN",\n'
            '  "report_date": "YYYY-MM-DD or null",\n'
            '  "solar_iso": [\n'
            '    {"region":"ERCOT|PJM|MISO|CAISO|SPP|NYISO|ISO-NE|AESO", "p25":<number $/MWh>,\n'
            '     "qoq_pct":<number or null>, "yoy_pct":<number or null>}\n'
            "  ],\n"
            '  "solar_continental": {\n'
            '    "p25":<number>, "p50":<number>, "p75":<number>,\n'
            '    "p10":<number or null>, "p90":<number or null>,\n'
            '    "qoq_pct":<number or null>, "yoy_pct":<number or null>\n'
            "  },\n"
            '  "solar_hub": [\n'
            '    {"region":"ERCOT", "hub":"HB_NORTH|HB_WEST|HB_SOUTH|HB_HOUSTON|SP15|Alberta|WESTERN HUB|DOM|N ILLINOIS HUB|AEP-DAYTON HUB|SPPNORTH_HUB|SPPSOUTH_HUB|MINN.HUB|ILLINOIS.HUB|INDIANA.HUB|LOUISIANA.HUB|ARKANSAS.HUB|...", "p25":<number>}\n'
            "  ],\n"
            '  "storage_iso": [\n'
            '    {"region":"AESO|CAISO|ERCOT|MISO|PJM|SPP|...", \n'
            '     "min":<number or null>, "p25":<number or null>, "median":<number or null>, "p75":<number or null>, "max":<number or null>,\n'
            '     "source":"levelten_index|chart_read"}\n'
            "  ],\n"
            '  "storage_duration_mix": [\n'
            '    {"region":"ERCOT", "2h":<pct or null>, "3h":<pct or null>, "4h":<pct or null>, "6h":<pct or null>, "8h":<pct or null>, "10h":<pct or null>}\n'
            "  ],\n"
            '  "solar_psv": [\n'
            '    {"region":"ERCOT", "psv_median":<number $/MWh>, "psv_min":<number>, "psv_max":<number>}\n'
            "  ],\n"
            '  "pipeline_breakdown": [\n'
            '    {"cod_year":"2025|2026|2027|2028|2029|2030+", "solar_mw":<number>, "standalone_storage_mw":<number>, "hybrid_mw":<number>}\n'
            "  ],\n"
            '  "storage_available": <true if BESS data found in report, false otherwise>,\n'
            '  "storage_note": "description of BESS data source",\n'
            '  "key_insights": ["1-line insight 1", "1-line insight 2", ...],\n'
            '  "notes": "2-3 sentence summary of quarter trends (Solar + Storage focus)"\n'
            "}\n\n"

            "CRITICAL — Storage Extraction:\n"
            "- LevelTen's 'Storage Price Spreads by ISO' is a BOX PLOT chart showing MIN, P25, MEDIAN, P75, MAX for each ISO.\n"
            "- Read ALL 5 statistics from the box plot. Round to nearest $0.5. Mark source='levelten_index'.\n"
            "- These are LEVELIZED TOLLING AGREEMENT prices in $/kW-month (confirmed by methodology).\n"
            "- Typical ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP (6 ISOs). ISO-NE and NYISO NOT covered by LevelTen Storage Index.\n"
            "- Also extract 'Storage Duration Distribution by ISO' chart → percent for each duration (2h, 3h, 4h, 6h, 8h, 10h).\n\n"

            "CRITICAL — Hub-level Solar P25 Extraction:\n"
            "- Every ISO has a 'PPA Prices by Hub' section showing maps with Solar P25 values labeled on each hub.\n"
            "- Extract every hub + price combination. Examples:\n"
            "  - ERCOT: HB_NORTH, HB_WEST, HB_SOUTH, HB_HOUSTON (4 hubs)\n"
            "  - CAISO: SP15 (1 hub)\n"
            "  - MISO: MINN.HUB, ILLINOIS.HUB, INDIANA.HUB, LOUISIANA.HUB, ARKANSAS.HUB (5 hubs)\n"
            "  - PJM: WESTERN HUB, DOM, AEP-DAYTON HUB, N ILLINOIS HUB (4 hubs)\n"
            "  - SPP: SPPNORTH_HUB, SPPSOUTH_HUB (2 hubs)\n"
            "  - AESO: Alberta (1 hub)\n\n"

            "CRITICAL — Solar PSV (Projected Settlement Value):\n"
            "- Report has 'Projected Settlement Values by Market: Solar' box plot chart.\n"
            "- Read median, min, max for each ISO shown. Values are in $/MWh (can be NEGATIVE).\n"
            "- Typical ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP.\n\n"

            "CRITICAL — Pipeline Breakdown:\n"
            "- Report has 'Technology Breakdown of Pipelines by COD Year' bar chart (in 'Going Hybrid' section).\n"
            "- Extract MW values for each year × technology. Include Solar, Standalone Storage, Hybrid.\n"
            "- DO NOT include Wind.\n\n"

            "General Rules:\n"
            "- Solar prices: USD/MWh. Storage prices: USD/kW-month.\n"
            "- If data not available for a field, use null. NEVER invent numbers.\n"
            "- DO NOT include any Wind data anywhere in the output.\n"
            "- Return ONLY the JSON object. No explanation. No code fences."
        )
        body = {
            "model": "claude-sonnet-4-5",
            "max_tokens": 4000,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "document", "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_b64,
                    }},
                    {"type": "text", "text": prompt},
                ],
            }],
        }
        try:
            res = requests.post("https://api.anthropic.com/v1/messages",
                                headers=headers, json=body, timeout=150)
            if res.status_code != 200:
                raise HTTPException(502, f"Claude API 오류: {res.text[:300]}")
            ai_text = res.json()["content"][0]["text"].strip()
            if ai_text.startswith("```"):
                ai_text = ai_text.split("```", 2)[1]
                if ai_text.startswith("json"): ai_text = ai_text[4:]
                ai_text = ai_text.rsplit("```", 1)[0]
            parsed = json.loads(ai_text.strip())
        except HTTPException:
            raise
        except requests.Timeout:
            raise HTTPException(504, "Claude API 응답 타임아웃 — 리포트가 너무 크거나 서버 혼잡 (재시도 권장)")
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"AI 응답 JSON 파싱 실패: {str(e)}")
    else:
        raise HTTPException(400, f"지원하지 않는 파일 형식: .{ext} (PDF, CSV, XLSX, XLSB 지원)")

    # CSV/Excel인 경우 Claude에게 텍스트 파싱 요청
    if parse_mode in ("csv", "excel"):
        headers = {
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        prompt = (
            "You are parsing LevelTen Energy PPA Price Index tabular data for a Solar+BESS developer. "
            f"The data below is from a {parse_mode.upper()} export.\n\n"
            f"DATA:\n{source_text[:30000]}\n\n"
            "IMPORTANT PRINCIPLES:\n"
            "1. We focus on Solar PPA and BESS Storage pricing. Skip Wind data.\n"
            "2. ONLY extract values that are EXPLICITLY present in the data. NEVER estimate or invent.\n"
            "3. If a value is missing, use null. Do NOT fill with guesses.\n\n"
            "Extract into strict JSON (no markdown, no prose):\n"
            "{\n"
            '  "quarter": "YYYY-QN",\n'
            '  "report_date": "YYYY-MM-DD or null",\n'
            '  "solar_iso": [{"region":"ERCOT|PJM|MISO|CAISO|SPP|ISO-NE|AESO", "p25":<$/MWh>, "qoq_pct":<null>, "yoy_pct":<null>}],\n'
            '  "solar_continental": {"p25":<number>, "p50":<number>, "p75":<number>, "p10":<number|null>, "p90":<number|null>, "qoq_pct":<null>, "yoy_pct":<null>},\n'
            '  "solar_hub": [{"region":"ERCOT", "hub":"North", "p25":<number>}],\n'
            '  "storage_iso": [{"region":"ERCOT|...", "p25":<$/kW-month|null>, "p50":<number|null>, "p75":<number|null>, "source":"table"}],\n'
            '  "storage_available": <true|false>,\n'
            '  "storage_note": "description or \\"Not included in data\\"",\n'
            '  "key_insights": ["actionable insight 1", "insight 2"],\n'
            '  "notes": "2-3 sentence summary (Solar+Storage focus, no Wind)"\n'
            "}\n"
            "If NO storage data in file: storage_iso=[], storage_available=false.\n"
            "Return ONLY the JSON object. No code fences."
        )
        body = {
            "model": "claude-sonnet-4-5",
            "max_tokens": 4000,
            "messages": [{"role": "user", "content": prompt}],
        }
        try:
            res = requests.post("https://api.anthropic.com/v1/messages",
                                headers=headers, json=body, timeout=150)
            if res.status_code != 200:
                raise HTTPException(502, f"Claude API 오류: {res.text[:300]}")
            ai_text = res.json()["content"][0]["text"].strip()
            if ai_text.startswith("```"):
                ai_text = ai_text.split("```", 2)[1]
                if ai_text.startswith("json"): ai_text = ai_text[4:]
                ai_text = ai_text.rsplit("```", 1)[0]
            parsed = json.loads(ai_text.strip())
        except HTTPException:
            raise
        except requests.Timeout:
            raise HTTPException(504, "Claude API 응답 타임아웃 — 리포트가 너무 크거나 서버 혼잡 (재시도 권장)")
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"AI 응답 JSON 파싱 실패: {str(e)}")

    if not parsed:
        raise HTTPException(500, "파싱 결과가 비어있습니다.")

    # 쿼터 형식 검증 (YYYY-QN)
    import re
    if not re.match(r'^\d{4}-Q[1-4]$', quarter.upper()):
        raise HTTPException(400, "쿼터 형식은 YYYY-Q1 ~ YYYY-Q4 여야 합니다.")
    quarter = quarter.upper()

    # 쿼터 덮어쓰기 (사용자가 명시한 값이 우선)
    parsed["quarter"] = quarter
    parsed["uploaded_at"] = datetime.datetime.utcnow().isoformat()[:19]
    parsed["uploaded_by"] = user["email"]
    parsed["filename"] = filename
    parsed["parse_mode"] = parse_mode

    # Backward compat: 새 파서 스키마를 legacy entries 배열로도 변환
    if "entries" not in parsed:
        legacy = []
        for s in parsed.get("solar_iso", []) or []:
            legacy.append({"tech":"solar", "region": s.get("region",""), "term_yr":10,
                           "p25": s.get("p25"), "p50": None, "p75": None})
        for s in parsed.get("storage_iso", []) or []:
            # 새 스키마: min/p25/median/p75/max → legacy: p25/p50/p75
            legacy.append({"tech":"storage", "region": s.get("region",""), "term_yr":10,
                           "p25": s.get("p25"),
                           "p50": s.get("median") or s.get("p50"),
                           "p75": s.get("p75")})
        parsed["entries"] = legacy

    # Firebase 저장: benchmark/levelten/{quarter}
    fb_put(f"benchmark/levelten/{quarter}", parsed)

    return {
        "ok": True,
        "quarter": quarter,
        "solar_iso_count": len(parsed.get("solar_iso", []) or []),
        "storage_iso_count": len(parsed.get("storage_iso", []) or []),
        "entries_count": len(parsed.get("entries", []) or []),
        "data": parsed,
    }


@router.get("/benchmark/levelten")
def get_levelten_all(user=Depends(get_current_user)):
    """모든 분기별 LevelTen 데이터."""
    return fb_read("benchmark/levelten") or {}


@router.get("/benchmark/levelten/latest")
def get_levelten_latest(user=Depends(get_current_user)):
    """가장 최신 분기의 LevelTen 데이터."""
    all_data = fb_read("benchmark/levelten") or {}
    if not all_data:
        return {}
    # 쿼터 문자열 정렬 (YYYY-QN 포맷이라 사전순 정렬로 충분)
    latest_key = sorted(all_data.keys())[-1]
    return {"quarter": latest_key, **all_data[latest_key]}


@router.delete("/benchmark/levelten/{quarter}")
def delete_levelten(quarter: str, user=Depends(require_admin)):
    """특정 분기 LevelTen 데이터 삭제."""
    import re
    if not re.match(r'^\d{4}-Q[1-4]$', quarter.upper()):
        raise HTTPException(400, "쿼터 형식은 YYYY-Q1 ~ YYYY-Q4 여야 합니다.")
    quarter = quarter.upper()
    try:
        requests.delete(f"{FB_URL}/benchmark/levelten/{quarter}.json",
                        params=fb_auth_param(), timeout=5)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, f"삭제 오류: {str(e)}")


# ── 피어 IRR 벤치마크 (내부 수동 입력값) ────────────
@router.get("/benchmark/peer-irr")
def get_peer_irr(user=Depends(get_current_user)):
    """저장된 피어 IRR 벤치마크 조회."""
    return fb_read("benchmark/peer_irr") or {}


@router.post("/benchmark/peer-irr")
def save_peer_irr(payload: dict, user=Depends(get_current_user)):
    """피어 IRR 벤치마크 저장 (Levered Pre-Tax IRR 레인지)."""
    # 필드 검증
    required_numeric = ["solar_min", "solar_max", "hybrid_min", "hybrid_max", "wind_min", "wind_max"]
    data = {}
    for k in required_numeric:
        v = payload.get(k)
        if v is None:
            raise HTTPException(400, f"누락된 필드: {k}")
        try:
            fv = float(v)
            if fv < 0 or fv > 50:
                raise HTTPException(400, f"{k}: 0~50% 범위여야 합니다.")
            data[k] = round(fv, 2)
        except (ValueError, TypeError):
            raise HTTPException(400, f"{k}: 숫자여야 합니다.")
    # min < max 검증
    for tech in ("solar", "hybrid", "wind"):
        if data[f"{tech}_min"] >= data[f"{tech}_max"]:
            raise HTTPException(400, f"{tech}: min < max 이어야 합니다.")
    # 비고
    note = payload.get("note", "")
    if isinstance(note, str):
        data["note"] = note[:200]
    data["updated_at"] = datetime.datetime.utcnow().isoformat()[:19]
    data["updated_by"] = user["email"]
    fb_put("benchmark/peer_irr", data)
    return {"ok": True, "data": data}


# ══════════════════════════════════════════════════
#  BESS Tolling Market Research (AI Web Search)
# ══════════════════════════════════════════════════
@router.post("/benchmark/bess-tolling/research")
def research_bess_tolling(user=Depends(get_current_user)):
    """
    Claude API + web_search 도구로 ISO별 BESS tolling 가격을 실시간 리서치.
    결과: ISO × Duration 별 P25/P75 + 출처 URL + confidence score.
    캐시: benchmark/bess_tolling/latest (수동 새로고침)
    """
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 미설정")

    today_str = datetime.date.today().isoformat()

    prompt = (
        f"You are an energy market research analyst specializing in US battery energy storage "
        f"system (BESS) tolling agreements AND PPA markets. Today: {today_str}.\n\n"
        "ROLE: This research is COMPLEMENTARY to LevelTen's official PPA Price Index.\n"
        "LevelTen publishes official data for 6 ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP.\n"
        "For these 6 ISOs the dashboard uses LevelTen first — your role is DURATION-level BESS detail only.\n\n"
        "YOUR FOCUS (three objectives):\n"
        "  (A) NON-LEVELTEN ISOs — provide BOTH BESS tolling AND PPA market commentary:\n"
        "      - ISO-NE (New England)\n"
        "      - NYISO (New York)\n"
        "      - WECC_DSW (Desert Southwest: AZ, NM, NV — Arizona/New Mexico/Nevada utilities)\n"
        "      - WECC_RM  (Rocky Mountain: UT, CO, WY, ID — PacifiCorp East/RMP, Xcel Colorado)\n"
        "      - WECC_NW  (Northwest: OR, WA, MT — PacifiCorp West, PGE, Puget Sound Energy)\n"
        "      - SERC (Southeast: TVA, Duke, Southern Company territory — NC, SC, GA, AL, TN, KY)\n"
        "  (B) DURATION BREAKDOWN for LevelTen-covered ISOs (ERCOT, CAISO, PJM, MISO, SPP, AESO):\n"
        "      → duration-level prices (2h / 4h / 6h) — LevelTen only gives ISO-level\n"
        "  (C) For WECC sub-regions: include PPA market commentary since LevelTen has ZERO coverage.\n"
        "      Key utility RFPs to reference: PacifiCorp IRP RFP, URC (Utah Renewable Communities),\n"
        "      APS (Arizona Public Service), NV Energy, Xcel Energy Colorado, Portland General Electric,\n"
        "      Idaho Power, Puget Sound Energy.\n\n"
        "Research methodology — TRIANGULATION:\n"
        "  1. Capacity market clearing prices (PJM, NYISO, ISO-NE) — adjusted for storage\n"
        "  2. Merchant BESS revenue data (ERCOT ~$30-50/kW-yr, CAISO duck curve premium)\n"
        "  3. Utility RFP announcements when prices are disclosed (PacifiCorp, APS, Xcel, etc.)\n"
        "  4. Public company earnings calls (NextEra, Vistra, AES)\n"
        "  5. Duration-adjustment heuristic:\n"
        "     - 2h: 60-75% of 4h price (arbitrage-dominated)\n"
        "     - 4h: reference (capacity-dominated, NERC/ISO standard)\n"
        "     - 6h+: 110-130% of 4h (long-duration premium)\n"
        "  6. Industry rule-of-thumb benchmarks (2025):\n"
        "     - ERCOT 2h: $3-8/kW-mo  | 4h: $5-12/kW-mo\n"
        "     - CAISO 4h: $10-16/kW-mo (duck curve) | 8h: $13-20/kW-mo\n"
        "     - PJM 4h: $8-13/kW-mo (capacity market) | 2h: $5-9/kW-mo\n"
        "     - SPP/MISO 4h: $6-11/kW-mo\n"
        "     - ISO-NE 4h: $12-18/kW-mo (tight capacity, winter peak)\n"
        "     - NYISO 4h: $11-17/kW-mo (DEC mandate, expensive zones J/K)\n"
        "     - WECC_DSW 4h: $7-12/kW-mo (APS/NV Energy RFPs, solar-shifting demand)\n"
        "     - WECC_RM 4h: $6-11/kW-mo (PacifiCorp/Xcel CO — emerging market, thin liquidity)\n"
        "     - WECC_NW 4h: $6-10/kW-mo (hydro-dominant, moderate storage need)\n"
        "     - SERC 4h: $7-12/kW-mo (vertically integrated utilities, bilateral)\n"
        "Use these as STARTING POINTS, then VERIFY/ADJUST via web_search.\n\n"
        "Use web_search to find CURRENT data from:\n"
        "- Wood Mackenzie, BloombergNEF, S&P Global, LCG Consulting\n"
        "- ISO capacity auction results: PJM BRA, ISO-NE FCA, NYISO ICAP\n"
        "- State PUC filings for RFP results (Utah PSC, Colorado PUC, Oregon PUC, Arizona ACC)\n"
        "- Utility IRP documents (PacifiCorp IRP, APS IRP, Xcel Colorado ERP)\n"
        "- Press releases: NextEra, Invenergy, AES, EDP, Engie, Brookfield\n"
        "- News: Utility Dive, Energy Storage News, Reuters, Canary Media\n\n"
        "TARGET REGIONS (10 total):\n"
        "- PRIMARY (no LevelTen coverage, full research required):\n"
        "    ISO-NE, NYISO, WECC_DSW, WECC_RM, WECC_NW, SERC\n"
        "- SECONDARY (LevelTen-covered, provide duration breakdown only):\n"
        "    ERCOT, CAISO, PJM, MISO, SPP, AESO\n\n"
        "TARGET DURATIONS for each region: 2h, 4h, 6h\n\n"
        "Output: ALL text fields (market_note, methodology_note, caveats) MUST BE IN KOREAN.\n"
        "Use formal nominal/concise style ('~확인됨', '~추정됨', '~범위').\n"
        "For WECC_* regions, market_note MUST include PPA market commentary (utility RFP landscape, "
        "recent clearing prices, Neptune-like Utah projects context).\n"
        "Numbers stay numeric ($X/kW-mo). Region names stay English.\n\n"
        "Return ONLY this JSON structure (no markdown, no code fences):\n"
        "{\n"
        '  "research_date": "YYYY-MM-DD",\n'
        '  "iso_data": [\n'
        '    {\n'
        '      "region": "ERCOT|CAISO|PJM|MISO|SPP|AESO|ISO-NE|NYISO|WECC_DSW|WECC_RM|WECC_NW|SERC",\n'
        '      "levelten_covered": true,  // 6 LevelTen ISOs=true, 나머지 4 (ISO-NE/NYISO/WECC_*/SERC)=false\n'
        '      "durations": [\n'
        '        {"hours": 2, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"},\n'
        '        {"hours": 4, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"},\n'
        '        {"hours": 6, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"}\n'
        '      ],\n'
        '      "market_note": "(한국어) 시장 특성 1-2문장. WECC_*는 PPA 시장 commentary 포함 (주요 utility RFP, recent clearing prices, 인접 주 벤치마크)",\n'
        '      "sources": [\n'
        '        {"url": "https://...", "title": "source title", "date": "YYYY-MM", "key_data": "핵심 수치/인용 (한국어 번역 OK)"}\n'
        '      ]\n'
        '    }\n'
        '  ],\n'
        '  "methodology_note": "(한국어) 추정 방법 요약. LevelTen 공식 index와의 관계 명시: LevelTen은 6개 ISO만 커버, 본 리서치는 (1) 미커버 4개 지역(ISO-NE/NYISO/WECC_DSW/WECC_RM/WECC_NW/SERC) 보완, (2) 전체 duration별(2h/4h/6h) 세분화 목표. capacity market + merchant 수익 + utility RFP 삼각 검증",\n'
        '  "confidence_overall": "high|medium|low",\n'
        '  "caveats": "(한국어) 1-2문장. 예: 본 수치는 AI 리서치 기반 추정치. LevelTen 6개 ISO는 공식 데이터 우선. WECC sub-region 및 SERC는 공식 index 부재 — RFP/IRP 참고치"\n'
        "}\n\n"
        "Rules:\n"
        "- All prices in USD/kW-month, levelized over contract term.\n"
        "- ALWAYS include all 12 regions (10 + WECC split into 3 sub-regions):\n"
        "  ERCOT, CAISO, PJM, MISO, SPP, AESO, ISO-NE, NYISO, WECC_DSW, WECC_RM, WECC_NW, SERC.\n"
        "- Each region must have 3 durations (2h, 4h, 6h) — use benchmark if no evidence, mark confidence='low'.\n"
        "- Confidence guide: 'high' if 3+ sources corroborate; 'medium' if 1-2 sources; 'low' if benchmark/inference only.\n"
        "- For WECC_* regions, market_note MUST include PPA context (not just BESS) — target utilities and recent RFP clearing prices.\n"
        "- Dates must be 2024-2026 (recent only).\n"
        "- All text fields in Korean formal nominal style.\n"
        "- Return valid JSON only."
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-5",
                "max_tokens": 8000,
                "tools": [{
                    "type": "web_search_20250305",
                    "name": "web_search",
                    "max_uses": 10,
                }],
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=180,  # 웹서치 여러 번 → 최대 3분
        )
        if resp.status_code != 200:
            raise HTTPException(502, f"Claude API 오류: {resp.text[:400]}")

        data = resp.json()
        # content blocks 중 text 타입만 합쳐서 JSON 파싱
        text_parts = []
        for block in data.get("content", []):
            if block.get("type") == "text":
                text_parts.append(block.get("text", ""))
        full_text = "".join(text_parts).strip()

        # JSON 추출 (code fence 제거 + { } 범위)
        import re as _re
        clean = _re.sub(r"```(?:json)?\s*", "", full_text).strip().strip("`")
        start = clean.find("{")
        end = clean.rfind("}") + 1
        if start < 0 or end <= start:
            raise HTTPException(500, f"AI 응답에서 JSON을 찾을 수 없음: {full_text[:300]}")
        clean = clean[start:end]

        try:
            parsed = json.loads(clean)
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"JSON 파싱 실패: {str(e)}. 응답: {clean[:400]}")

        # 메타데이터 추가
        parsed["generated_at"] = datetime.datetime.utcnow().isoformat()[:19]
        parsed["generated_by"] = user["email"]
        parsed["source"] = "ai_research"
        # 토큰 사용량 (cost 추적용)
        usage = data.get("usage", {})
        parsed["tokens"] = {
            "input": usage.get("input_tokens", 0),
            "output": usage.get("output_tokens", 0),
        }

        # Firebase 저장 (latest로)
        fb_put("benchmark/bess_tolling/latest", parsed)
        # 히스토리도 (월별 캐시)
        month_key = datetime.date.today().strftime("%Y-%m")
        fb_put(f"benchmark/bess_tolling/history/{month_key}", parsed)

        return {"ok": True, "data": parsed}

    except HTTPException:
        raise
    except requests.Timeout:
        raise HTTPException(504, "AI 리서치 타임아웃 (3분 초과)")
    except Exception as e:
        raise HTTPException(500, f"BESS 리서치 실패: {str(e)}")


@router.get("/benchmark/bess-tolling")
def get_bess_tolling(user=Depends(get_current_user)):
    """저장된 BESS tolling 리서치 결과 조회 (latest)."""
    return fb_read("benchmark/bess_tolling/latest") or {}


@router.get("/benchmark/bess-tolling/history")
def get_bess_tolling_history(user=Depends(get_current_user)):
    """월별 히스토리 (stale 확인용)."""
    return fb_read("benchmark/bess_tolling/history") or {}
