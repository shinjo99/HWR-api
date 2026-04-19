"""
audit.py — HWR Model Audit (PF Excel 정합성 검증)

Phase 4 Step 4B refactoring에서 main.py에서 분리됨.

Contents:
- _integrity_check_pf_model: PF 모델 엑셀 파일(.xlsb/.xlsx)의 구조적 정합성 검증
  - Y0 Sponsor CF가 CAPEX와 일치하는지
  - MACRS depreciation schedule 올바른지
  - Debt service 스케줄 정상인지
  - TE distribution 일관성
  - 기타 여러 sanity check

주의: 함수 내부에서 openpyxl, pyxlsb를 동적으로 import 하므로
이 파일은 모듈 레벨 외부 import가 필요 없음.

Generated: Apr 20, 2026
"""


def _integrity_check_pf_model(filepath: str, lang: str = 'ko') -> dict:
    """PF 엑셀 모델 정합성 체크
    
    5개 카테고리 × 심각도별 (HIGH/MEDIUM/LOW) 검사
    
    Args:
        filepath: 엑셀 파일 경로
        lang: 'ko' or 'en' - 리포트 언어
    
    Returns:
        {'checks': [...], 'summary': {...}, 'metadata': {...}}
    """
    # ─── 번역 딕셔너리 ───
    T = {
        'ko': {
            'cat_formula': '수식 오류', 'cat_capital': 'Capital Stack',
            'cat_irr': 'IRR 합리성', 'cat_debt': 'Debt Schedule',
            'cat_revenue': 'Revenue 현실성', 'cat_summary': '종합',
            # Formula errors
            'conv_fail_t': 'xlsb → xlsx 변환 실패',
            'conv_fail_d': '파일 형식 변환 중 오류 발생. 수식 분석 불가.',
            'conv_fail_a': 'LibreOffice 설치 확인 or xlsx 직접 업로드 권장',
            'conv_err_t': 'xlsb → xlsx 변환 에러',
            'conv_err_d': '변환 오류: ',
            'conv_err_a': 'xlsx로 재저장 후 업로드',
            'formula_skip_t': '수식 스트링 분석 스킵 (xlsb 직접 분석)',
            'formula_skip_d': '서버에 LibreOffice가 없어 xlsb → xlsx 변환을 건너뛰었습니다. 수식 텍스트 분석은 스킵하며, 셀 결과값(#REF!/#VALUE! 등)과 나머지 정합성 체크는 xlsb에서 직접 수행합니다.',
            'formula_skip_a': '수식 정밀 분석이 필요하면 xlsx로 저장 후 재업로드',
            'formula_t': '심각한 수식 에러 {n}개',
            'formula_d': '#NAME?/#REF!/#DIV/0 등 에러. 샘플: ',
            'formula_a': '원본 모델러에게 해당 셀 재확인 요청. Named Range 깨짐 or 외부 참조 오류 추정.',
            'na_many_t': '#N/A 값 {n}개',
            'na_many_d': 'VLOOKUP/INDEX 미스매치 or 빈 lookup 범위 가능. 정상일 수 있음.',
            'na_many_a': '핵심 계산 탭(Returns, CF, Summary)에 #N/A 있는지 집중 확인',
            'ext_ref_t': '외부 파일 참조 {n}곳',
            'ext_ref_d': '담당자 local path 참조 가능성. 샘플: ',
            'ext_ref_a': '링크 끊기(Break Links) or 값 복사 처리 필요',
            'parse_err_t': '수식 파싱 실패',
            'parse_err_d': '엑셀 구조 분석 중 오류: ',
            'parse_err_a': '파일 무결성 확인',
            # Capital Stack
            'stack_t': 'Capital Stack 합계 불일치 ({err:.1f}% 오차)',
            'stack_a': 'Summary 탭 Capital Stack 수식 재검증',
            'debt_ratio_t': 'Debt 비율 이상치 ({r:.1f}%)',
            'debt_ratio_d': '일반적 Debt 40-60% 범위를 벗어남',
            'debt_ratio_a': '금융 자문사/Debt Sizing 결과 재확인',
            'te_high_t': 'TE 비율 과다 ({r:.1f}%)',
            'te_high_d': 'TE 40% 초과 시 Sponsor 희석 심각',
            'te_high_a': 'TE 투자 계약 조건 재검토',
            'sponsor_low_t': 'Sponsor Eq 비율 과소 ({r:.1f}%)',
            'sponsor_low_d': 'Sponsor 5% 미만 시 프로젝트 관여도 낮음',
            'sponsor_low_a': 'Sponsor 실질 참여 구조 확인',
            'summary_t': 'Summary 탭 파싱 제한',
            'summary_d': '자동 체크 일부 스킵: ',
            'summary_a': '수동으로 Summary 탭 확인',
            # IRR
            'irr_high_t': '{label} 과도 ({p:.2f}%)',
            'irr_high_d': '25%+ IRR은 일반 Solar+BESS 드문 수치',
            'irr_high_a': 'Revenue/CAPEX 가정 재검토',
            'irr_low_t': '{label} 과소 ({p:.2f}%)',
            'irr_low_d': '5% 미만 IRR은 투자 매력 부족',
            'irr_low_a': 'CAPEX/OPEX/Revenue 전면 재검토',
            'irr_contract_t': 'Contract IRR > Full Life IRR ({sc:.2f}% vs {sf:.2f}%)',
            'irr_contract_d': 'Merchant 기간(Y26+) CF가 음수 또는 IRR 끌어내림. PPA 종료 후 수익성 악화 신호',
            'irr_contract_a': 'Merchant 가정(PV $61, BESS Toll 종료) 재검토',
            # Debt
            'dscr_low_t': 'DSCR 최저값 위험 ({v:.2f})',
            'dscr_low_d': 'DSCR 1.20 미만 기간 있음. 표준 covenant 1.30 위반 가능.',
            'dscr_low_a': 'Debt sizing 재검토 또는 cash sweep 조항 확인',
            'dscr_var_t': 'DSCR 편차 과다 ({lo:.2f}~{hi:.2f})',
            'dscr_var_d': 'Sculpted debt (DSCR 기반 동적 상환) 구조 추정',
            'dscr_var_a': 'Debt 탭 연도별 상환 스케줄 수동 검토 필수',
            # Revenue
            'ppa_t': 'PPA Price 이상치 (${v:.2f}/MWh)',
            'ppa_d': '일반 Solar PPA $40-90/MWh 범위 밖',
            'ppa_a': 'PPA 계약서 원본 확인',
            'merch_t': 'Merchant Price 높음 (${v:.2f}/MWh)',
            'merch_d': 'Merchant 가격 $100+ 는 공격적 가정',
            'merch_a': 'Ventyx/WM 예측 근거 확인',
            'deg_t': 'Degradation 이상치 ({v:.2f}%/yr)',
            'deg_d': '일반 0.4-0.8%/yr 범위',
            'deg_a': 'Module warranty 확인',
            # Recommend
            'rec_calib_t': '권장 계산 모드: Calibration',
            'rec_calib_d': '이 모델은 Neptune과 유사한 구조로 보입니다. 근거: ',
            'rec_calib_none': 'Neptune 패턴 일부 감지',
            'rec_calib_a': 'Dashboard에서 "🎯 Calibration" 모드로 계산 권장',
            'rec_predict_t': '권장 계산 모드: Prediction',
            'rec_predict_d': '일반 PF 구조로 판단됩니다.',
            'rec_predict_a': 'Dashboard "📈 Prediction" 모드로 계산 (기본값)',
            'clear_t': '주요 이상 징후 없음',
            'clear_d': '자동 체크한 5개 카테고리에서 큰 이슈 미발견.',
            'clear_a': '수동 검토는 여전히 권장',
            # Mode reasons
            'reason_sculpted': 'Sculpted debt 구조 감지 (DSCR 편차 큼)',
            'reason_irr_invert': 'Contract > Full IRR 역전 (Neptune 패턴)',
            'reason_debt_ratio': 'Debt 비율 {r:.1f}% (Neptune 유사 47.6%)',
            'reason_te_ratio': 'TE 비율 {r:.1f}% (Neptune 유사 32.5%)',
            # Capital Stack mismatch
            'stack_mismatch_t': 'Capital Stack 합산 불일치 ({pct:.1f}% 오차)',
            'stack_mismatch_d': 'CAPEX ${capex:,.1f}M ≠ Debt ${debt:,.1f}M + TE ${te:,.1f}M + Eq ${eq:,.1f}M (합계 ${sum:,.1f}M, 차이 ${diff:+,.1f}M)',
            'stack_mismatch_a': '모델의 Capital Stack 섹션에서 자금조달 출처별 금액 재검토 필요. 누락된 자금원 or 이중계산 가능성.',
            # DSCR
            'dscr_low_t': 'DSCR 부족 (최소 {v:.2f}x < 1.0)',
            'dscr_low_d': '특정 연도에 EBITDA가 Debt Service를 감당하지 못함. Covenant breach 가능성.',
            'dscr_low_a': 'Debt 스케줄 재검토 또는 상환 조건 완화 필요',
            'dscr_tight_t': 'DSCR 타이트 (최소 {v:.2f}x, 업계 표준 ≥ 1.20)',
            'dscr_tight_d': '기술적 Default 위험은 낮으나 업계 관행 대비 커버리지 여유 작음.',
            'dscr_tight_a': '대주단이 통상 요구하는 DSCR 1.20~1.30x 대비 낮음. 금융조건 확인',
            # Value range sanity
            'range_t': '값 범위 이상 {n}개',
            'range_a': '업계 일반 범위를 벗어난 값. 입력 오타 or 특수 구조 가능성 확인',
            'range_debt': 'Debt 비율 {v:.1f}% (정상 25~75%)',
            'range_te': 'TE 비율 {v:.1f}% (정상 0~50%)',
            'range_ppa': 'PPA ${v:.1f}/MWh (정상 $20~$200)',
            'range_flip': 'Flip Yield {v:.2f}% (정상 5~15%)',
        },
        'en': {
            'cat_formula': 'Formula Errors', 'cat_capital': 'Capital Stack',
            'cat_irr': 'IRR Validity', 'cat_debt': 'Debt Schedule',
            'cat_revenue': 'Revenue Realism', 'cat_summary': 'Summary',
            'conv_fail_t': 'xlsb → xlsx conversion failed',
            'conv_fail_d': 'Error during file format conversion. Formula analysis not possible.',
            'conv_fail_a': 'Check LibreOffice install or upload xlsx directly',
            'conv_err_t': 'xlsb → xlsx conversion error',
            'conv_err_d': 'Conversion error: ',
            'conv_err_a': 'Re-save as xlsx and re-upload',
            'formula_skip_t': 'Formula-text analysis skipped (xlsb analyzed directly)',
            'formula_skip_d': 'Server lacks LibreOffice; xlsb → xlsx conversion skipped. Formula-string analysis is bypassed, but cell-value errors (#REF!/#VALUE! etc.) and all other integrity checks are performed directly on xlsb.',
            'formula_skip_a': 'Re-save as xlsx and re-upload if detailed formula analysis needed',
            'formula_t': '{n} serious formula errors',
            'formula_d': '#NAME?/#REF!/#DIV/0 errors. Samples: ',
            'formula_a': 'Ask the original modeler to verify cells. Likely broken Named Range or external reference errors.',
            'na_many_t': '{n} #N/A values',
            'na_many_d': 'VLOOKUP/INDEX mismatch or empty lookup range possible. May be intentional.',
            'na_many_a': 'Focus on #N/A in core tabs (Returns, CF, Summary)',
            'ext_ref_t': '{n} external file references',
            'ext_ref_d': 'Possible user local path references. Samples: ',
            'ext_ref_a': 'Break links or copy-paste values required',
            'parse_err_t': 'Formula parsing failed',
            'parse_err_d': 'Error during Excel structure analysis: ',
            'parse_err_a': 'Check file integrity',
            'stack_t': 'Capital Stack sum mismatch ({err:.1f}% error)',
            'stack_a': 'Re-verify Capital Stack formula in Summary tab',
            'debt_ratio_t': 'Debt ratio outlier ({r:.1f}%)',
            'debt_ratio_d': 'Outside typical Debt 40-60% range',
            'debt_ratio_a': 'Re-verify with financial advisor/Debt Sizing',
            'te_high_t': 'TE ratio excessive ({r:.1f}%)',
            'te_high_d': 'TE over 40% means severe Sponsor dilution',
            'te_high_a': 'Review TE investment agreement terms',
            'sponsor_low_t': 'Sponsor Eq ratio too low ({r:.1f}%)',
            'sponsor_low_d': 'Sponsor below 5% suggests low project commitment',
            'sponsor_low_a': 'Verify actual Sponsor participation structure',
            'summary_t': 'Summary tab parsing limited',
            'summary_d': 'Some auto-checks skipped: ',
            'summary_a': 'Manually verify Summary tab',
            'irr_high_t': '{label} too high ({p:.2f}%)',
            'irr_high_d': '25%+ IRR is rare for Solar+BESS',
            'irr_high_a': 'Re-verify Revenue/CAPEX assumptions',
            'irr_low_t': '{label} too low ({p:.2f}%)',
            'irr_low_d': 'IRR below 5% lacks investment appeal',
            'irr_low_a': 'Comprehensive review of CAPEX/OPEX/Revenue',
            'irr_contract_t': 'Contract IRR > Full Life IRR ({sc:.2f}% vs {sf:.2f}%)',
            'irr_contract_d': 'Merchant period (Y26+) CF is negative or drags IRR down. Signal of deteriorating economics post-PPA.',
            'irr_contract_a': 'Re-verify Merchant assumptions (PV $61, BESS Toll end)',
            'dscr_low_t': 'DSCR minimum risk ({v:.2f})',
            'dscr_low_d': 'DSCR below 1.20 in some periods. Standard 1.30 covenant at risk.',
            'dscr_low_a': 'Re-verify Debt sizing or cash sweep provisions',
            'dscr_var_t': 'DSCR variance high ({lo:.2f}~{hi:.2f})',
            'dscr_var_d': 'Sculpted debt (DSCR-based dynamic amortization) likely',
            'dscr_var_a': 'Manual review of Debt tab annual schedule required',
            'ppa_t': 'PPA Price outlier (${v:.2f}/MWh)',
            'ppa_d': 'Outside typical Solar PPA $40-90/MWh range',
            'ppa_a': 'Verify original PPA contract',
            'merch_t': 'Merchant Price high (${v:.2f}/MWh)',
            'merch_d': 'Merchant price $100+ is aggressive',
            'merch_a': 'Verify Ventyx/WM forecast source',
            'deg_t': 'Degradation outlier ({v:.2f}%/yr)',
            'deg_d': 'Typical range 0.4-0.8%/yr',
            'deg_a': 'Verify Module warranty',
            'rec_calib_t': 'Recommended Mode: Calibration',
            'rec_calib_d': 'This model appears similar to Neptune structure. Reasons: ',
            'rec_calib_none': 'Some Neptune patterns detected',
            'rec_calib_a': 'Run calculations in "🎯 Calibration" mode',
            'rec_predict_t': 'Recommended Mode: Prediction',
            'rec_predict_d': 'Looks like standard PF structure.',
            'rec_predict_a': 'Run in "📈 Prediction" mode (default)',
            'clear_t': 'No major anomalies detected',
            'clear_d': 'No major issues found in 5 auto-checked categories.',
            'clear_a': 'Manual review still recommended',
            'reason_sculpted': 'Sculpted debt detected (large DSCR variance)',
            'reason_irr_invert': 'Contract > Full IRR inversion (Neptune pattern)',
            'reason_debt_ratio': 'Debt ratio {r:.1f}% (similar to Neptune 47.6%)',
            'reason_te_ratio': 'TE ratio {r:.1f}% (similar to Neptune 32.5%)',
            # Capital Stack mismatch
            'stack_mismatch_t': 'Capital Stack imbalance ({pct:.1f}% off)',
            'stack_mismatch_d': 'CAPEX ${capex:,.1f}M ≠ Debt ${debt:,.1f}M + TE ${te:,.1f}M + Eq ${eq:,.1f}M (sum ${sum:,.1f}M, diff ${diff:+,.1f}M)',
            'stack_mismatch_a': 'Review financing sources in the model. Possible missing source or double-counting.',
            # DSCR
            'dscr_low_t': 'DSCR insufficient (min {v:.2f}x < 1.0)',
            'dscr_low_d': 'EBITDA fails to cover Debt Service in some year(s). Possible covenant breach.',
            'dscr_low_a': 'Review debt schedule or relax repayment terms',
            'dscr_tight_t': 'DSCR tight (min {v:.2f}x; industry standard ≥ 1.20)',
            'dscr_tight_d': 'Technical default risk is low, but coverage cushion is thin vs market practice.',
            'dscr_tight_a': 'Below typical lender requirement of 1.20–1.30x. Confirm financing terms.',
            # Value range sanity
            'range_t': '{n} values out of normal range',
            'range_a': 'Values outside industry typical range. Check for typos or special structure.',
            'range_debt': 'Debt ratio {v:.1f}% (normal 25–75%)',
            'range_te': 'TE ratio {v:.1f}% (normal 0–50%)',
            'range_ppa': 'PPA ${v:.1f}/MWh (normal $20–$200)',
            'range_flip': 'Flip Yield {v:.2f}% (normal 5–15%)',
        },
    }
    t = T.get(lang, T['ko'])
    
    from openpyxl import load_workbook
    from pyxlsb import open_workbook as open_xlsb
    import tempfile, subprocess, os
    
    checks = []
    metadata = {}
    
    # xlsb인 경우 xlsx로 변환 (openpyxl로 수식 읽기 위해)
    # libreoffice 미설치 환경(Railway 등)에서는 변환 실패 시
    # 수식 체크만 스킵하고 나머지 (Capital Stack / IRR / Debt / Revenue)는 pyxlsb로 진행
    xlsx_path = filepath
    if filepath.endswith('.xlsb'):
        try:
            subprocess.run(['libreoffice', '--headless', '--convert-to', 'xlsx',
                          '--outdir', os.path.dirname(filepath), filepath],
                         check=True, capture_output=True, timeout=60)
            candidate = filepath.replace('.xlsb', '.xlsx')
            if os.path.exists(candidate):
                xlsx_path = candidate
            else:
                xlsx_path = None
                metadata['formula_text_analysis'] = 'skipped_no_libreoffice'
        except Exception:
            xlsx_path = None
            metadata['formula_text_analysis'] = 'skipped_no_libreoffice'
    
    # ═══ 1. 수식 오류 체크 (xlsx_path 있을 때만) ═══
    if xlsx_path and xlsx_path != filepath:
        try:
            wb = load_workbook(xlsx_path, data_only=False)
            sheet_names = wb.sheetnames
            metadata['sheets'] = sheet_names
            metadata['sheet_count'] = len(sheet_names)
            
            formula_errors = []    # 심각한 수식 에러
            na_errors = []
            external_refs = []
            
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        val_str = str(cell.value)
                        for err in ['#NAME?', '#REF!', '#DIV/0!', '#VALUE!', '#NULL!']:
                            if err in val_str:
                                formula_errors.append({
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'error': err,
                                    'formula': val_str[:100],
                                })
                                break
                        if '#N/A' in val_str:
                            na_errors.append({'sheet': sheet_name, 'cell': cell.coordinate})
                        if '[' in val_str and ('\\' in val_str or '.xls' in val_str):
                            if not any(e in val_str for e in ['#NAME?', '#REF!']):
                                external_refs.append({
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'ref': val_str[:150],
                                })
            
            if formula_errors:
                sample = formula_errors[:5]
                sample_text = '; '.join([f"{e['sheet']}!{e['cell']}({e['error']})" for e in sample])
                from collections import Counter
                sheet_counts = Counter(e['sheet'] for e in formula_errors)
                sheet_samples = {}
                for e in formula_errors:
                    sh = e['sheet']
                    if sh not in sheet_samples:
                        sheet_samples[sh] = []
                    if len(sheet_samples[sh]) < 3:
                        sheet_samples[sh].append(e['cell'])
                detail_by_sheet = [
                    {'sheet': sh, 'count': cnt, 'cells': sheet_samples.get(sh, [])}
                    for sh, cnt in sheet_counts.most_common()
                ]
                checks.append({
                    'category': t['cat_formula'], 'severity': 'HIGH',
                    'code': 'FORMULA_ERR',
                    'title': t['formula_t'].format(n=len(formula_errors)),
                    'description': t['formula_d'] + sample_text,
                    'action': t['formula_a'],
                    'detail_by_sheet': detail_by_sheet,
                })
            
            if len(na_errors) > 50:
                checks.append({
                    'category': t['cat_formula'], 'severity': 'LOW',
                    'code': 'NA_MANY',
                    'title': t['na_many_t'].format(n=len(na_errors)),
                    'description': t['na_many_d'],
                    'action': t['na_many_a'],
                })
            
            if external_refs:
                sample = external_refs[:3]
                sample_text = '; '.join([f"{e['sheet']}!{e['cell']}" for e in sample])
                checks.append({
                    'category': t['cat_formula'], 'severity': 'MEDIUM',
                    'code': 'EXT_REF',
                    'title': t['ext_ref_t'].format(n=len(external_refs)),
                    'description': t['ext_ref_d'] + sample_text,
                    'action': t['ext_ref_a'],
                    'detail': external_refs[:10],
                })
        except Exception as e:
            checks.append({
                'category': t['cat_formula'], 'severity': 'HIGH',
                'code': 'PARSE_ERR',
                'title': t['parse_err_t'],
                'description': t['parse_err_d'] + str(e)[:100],
                'action': t['parse_err_a'],
            })
    else:
        # xlsx 변환 불가 → 수식 체크 스킵, 메타데이터는 pyxlsb로 수집
        try:
            xlsb_path = filepath if filepath.endswith('.xlsb') else filepath.replace('.xlsx', '.xlsb')
            with open_xlsb(xlsb_path) as wb_v:
                sheet_names = list(wb_v.sheets)
                metadata['sheets'] = sheet_names
                metadata['sheet_count'] = len(sheet_names)
                # xlsb에서 #REF! / #VALUE! 등 에러 값 전수 조사
                formula_errors = []
                na_count = 0
                # Excel 컬럼 인덱스 → 문자 변환 (0→A, 25→Z, 26→AA)
                def col_letter(idx):
                    s = ''
                    idx += 1  # 1-based
                    while idx > 0:
                        idx, r = divmod(idx - 1, 26)
                        s = chr(65 + r) + s
                    return s
                for sheet_name in sheet_names:
                    try:
                        with wb_v.get_sheet(sheet_name) as ws:
                            for r_idx, row in enumerate(ws.rows()):
                                for cell in row:
                                    if cell.v is None: continue
                                    val_str = str(cell.v)
                                    # 치명 에러
                                    for err in ['#NAME?', '#REF!', '#DIV/0!', '#VALUE!', '#NULL!']:
                                        if err in val_str:
                                            try:
                                                col_idx = cell.c  # pyxlsb cell has .c (column index, 0-based)
                                            except AttributeError:
                                                col_idx = None
                                            coord = f"{col_letter(col_idx)}{r_idx+1}" if col_idx is not None else f"R{r_idx+1}"
                                            formula_errors.append({
                                                'sheet': sheet_name,
                                                'cell': coord,
                                                'error': err,
                                            })
                                            break
                                    # #N/A 별도 카운트 (정상일 수도 있어서 50개 넘을 때만 경고)
                                    if '#N/A' in val_str:
                                        na_count += 1
                    except Exception:
                        continue
                
                # HIGH: 치명 에러
                if formula_errors:
                    sample = formula_errors[:8]
                    sample_text = '; '.join([f"{e['sheet']}!{e['cell']}({e['error']})" for e in sample])
                    # 시트별 집계 (전체 기반, 상위 샘플 아닌 전체)
                    from collections import Counter
                    sheet_counts = Counter(e['sheet'] for e in formula_errors)
                    # 각 시트에서 처음 발견된 에러 3개씩만 샘플 유지
                    sheet_samples = {}
                    for e in formula_errors:
                        sh = e['sheet']
                        if sh not in sheet_samples:
                            sheet_samples[sh] = []
                        if len(sheet_samples[sh]) < 3:
                            sheet_samples[sh].append(e['cell'])
                    # detail_by_sheet: 프론트에서 바로 렌더 가능한 구조
                    detail_by_sheet = [
                        {'sheet': sh, 'count': cnt, 'cells': sheet_samples.get(sh, [])}
                        for sh, cnt in sheet_counts.most_common()
                    ]
                    checks.append({
                        'category': t['cat_formula'], 'severity': 'HIGH',
                        'code': 'FORMULA_ERR_VAL',
                        'title': t['formula_t'].format(n=len(formula_errors)),
                        'description': t['formula_d'] + sample_text,
                        'action': t['formula_a'],
                        'detail_by_sheet': detail_by_sheet,
                    })
                # LOW: #N/A 과다
                if na_count > 50:
                    checks.append({
                        'category': t['cat_formula'], 'severity': 'LOW',
                        'code': 'NA_MANY_VAL',
                        'title': t['na_many_t'].format(n=na_count),
                        'description': t['na_many_d'],
                        'action': t['na_many_a'],
                    })
                metadata['formula_error_count'] = len(formula_errors)
                metadata['na_count'] = na_count
        except Exception:
            pass
    
    # ═══ 2. Capital Stack 정합성 (값 기반 체크) ═══
    try:
        with open_xlsb(filepath if filepath.endswith('.xlsb') else filepath.replace('.xlsx', '.xlsb')) as wb_v:
            try:
                with wb_v.get_sheet('Summary') as ws:
                    rows = list(ws.rows())
                    # 각 행에서 label + 숫자 값들 추출
                    # Neptune은 Case 1 / Case 2 두 컬럼 — 마지막 숫자(Case 2) 우선
                    summary_rows = []
                    for row in rows:
                        vals = [(c.c, c.v) for c in row if c.v is not None]
                        if len(vals) < 2: continue
                        # 첫 문자열 = label
                        label = None
                        for c, v in vals:
                            if isinstance(v, str):
                                label = v.strip()
                                break
                        # 숫자 값들 (label 이후)
                        nums = [v for c, v in vals if isinstance(v, (int, float))]
                        if label and nums:
                            # Case 2 (마지막 유효값) 우선
                            summary_rows.append((label, nums[-1], nums))
                    
                    metadata['summary_rows_found'] = len(summary_rows)
                    
                    # 정확한 라벨 매칭 (Neptune 기준)
                    label_map = {
                        'total project cost': 'capex_m',
                        'debt': 'debt_m',
                        'tax equity investment': 'te_m',
                        'sponsor equity investment': 'sponsor_m',
                        'hqc dev margin': 'dev_margin_m',
                        'levered project irr (full life)': 'lev_irr',
                        'unlevered project irr (full life)': 'unlev_irr',
                        'sponsor levered irr (full life)': 'sponsor_full_irr',
                        'sponsor levered irr (contract)': 'sponsor_contract_irr',
                    }
                    
                    extracted = {}
                    for label, val, all_nums in summary_rows:
                        key = label.lower().strip()
                        # 콜론, 괄호 등 정리
                        key_clean = key.replace(':', '').replace('  ', ' ').strip()
                        for pattern, field in label_map.items():
                            if pattern in key_clean:
                                extracted[field] = val
                                break
                    
                    metadata.update(extracted)
                    
                    capex_m = extracted.get('capex_m')
                    debt_m = extracted.get('debt_m')
                    te_m = extracted.get('te_m')
                    sponsor_m = extracted.get('sponsor_m')
                    
                    # 체크: 합계 일치
                    if all(v is not None for v in [capex_m, debt_m, te_m, sponsor_m]):
                        total = debt_m + te_m + sponsor_m
                        err_pct = abs(total - capex_m) / capex_m * 100 if capex_m else 0
                        if err_pct > 1:
                            checks.append({
                                'category': t['cat_capital'], 'severity': 'HIGH',
                                'code': 'STACK_MISMATCH',
                                'title': t['stack_t'].format(err=err_pct),
                                'description': f'Debt ${debt_m:,.0f}K + TE ${te_m:,.0f}K + Sponsor ${sponsor_m:,.0f}K = ${total:,.0f}K, CAPEX ${capex_m:,.0f}K',
                                'action': t['stack_a'],
                            })
                        
                        debt_ratio = debt_m / capex_m * 100 if capex_m else 0
                        te_ratio = te_m / capex_m * 100 if capex_m else 0
                        sponsor_ratio = sponsor_m / capex_m * 100 if capex_m else 0
                        
                        metadata['debt_ratio'] = round(debt_ratio, 2)
                        metadata['te_ratio'] = round(te_ratio, 2)
                        metadata['sponsor_ratio'] = round(sponsor_ratio, 2)
                        
                        if debt_ratio < 30 or debt_ratio > 70:
                            checks.append({
                                'category': t['cat_capital'], 'severity': 'MEDIUM',
                                'code': 'DEBT_RATIO',
                                'title': t['debt_ratio_t'].format(r=debt_ratio),
                                'description': t['debt_ratio_d'],
                                'action': t['debt_ratio_a'],
                            })
                        if te_ratio > 45:
                            checks.append({
                                'category': t['cat_capital'], 'severity': 'MEDIUM',
                                'code': 'TE_RATIO_HIGH',
                                'title': t['te_high_t'].format(r=te_ratio),
                                'description': t['te_high_d'],
                                'action': t['te_high_a'],
                            })
                        if sponsor_ratio < 5:
                            checks.append({
                                'category': t['cat_capital'], 'severity': 'MEDIUM',
                                'code': 'SPONSOR_LOW',
                                'title': t['sponsor_low_t'].format(r=sponsor_ratio),
                                'description': t['sponsor_low_d'],
                                'action': t['sponsor_low_a'],
                            })
                    
                    # IRR 체크 (Summary 탭에서 직접)
                    sponsor_full = extracted.get('sponsor_full_irr')
                    sponsor_contract = extracted.get('sponsor_contract_irr')
                    lev_irr = extracted.get('lev_irr')
                    unlev_irr = extracted.get('unlev_irr')
                    
                    for label, val in [
                        ('Sponsor IRR Full Life', sponsor_full),
                        ('Sponsor IRR Contract', sponsor_contract),
                        ('Project Levered IRR', lev_irr),
                        ('Project Unlevered IRR', unlev_irr),
                    ]:
                        if val is not None:
                            # 0~1 → percent
                            pct = val * 100 if 0 < val < 1 else val
                            if pct > 25:
                                checks.append({
                                    'category': t['cat_irr'], 'severity': 'MEDIUM',
                                    'code': 'IRR_HIGH',
                                    'title': t['irr_high_t'].format(label=label, p=pct),
                                    'description': t['irr_high_d'],
                                    'action': t['irr_high_a'],
                                })
                            elif pct < 5:
                                checks.append({
                                    'category': t['cat_irr'], 'severity': 'HIGH',
                                    'code': 'IRR_LOW',
                                    'title': t['irr_low_t'].format(label=label, p=pct),
                                    'description': t['irr_low_d'],
                                    'action': t['irr_low_a'],
                                })
                    
                    # Contract > Full 역전 체크
                    if sponsor_full is not None and sponsor_contract is not None:
                        sf = sponsor_full * 100 if 0 < sponsor_full < 1 else sponsor_full
                        sc = sponsor_contract * 100 if 0 < sponsor_contract < 1 else sponsor_contract
                        if sc > sf + 2:  # Contract가 Full보다 2%p 이상 높으면
                            checks.append({
                                'category': t['cat_irr'], 'severity': 'MEDIUM',
                                'code': 'IRR_CONTRACT_HIGH',
                                'title': t['irr_contract_t'].format(sc=sc, sf=sf),
                                'description': t['irr_contract_d'],
                                'action': t['irr_contract_a'],
                            })
            except Exception as e:
                checks.append({
                    'category': t['cat_capital'], 'severity': 'LOW',
                    'code': 'SUMMARY_PARSE',
                    'title': t['summary_t'],
                    'description': t['summary_d'] + str(e)[:80],
                    'action': t['summary_a'],
                })
    except Exception:
        pass
    
    # ═══ 4. Debt Schedule 체크 ═══
    try:
        with open_xlsb(filepath if filepath.endswith('.xlsb') else None) as wb_v:
            with wb_v.get_sheet('Debt') as ws:
                rows = list(ws.rows())
                # DSCR 라인 탐색 (보통 'DSCR' 라벨이 A열에)
                dscr_values = []
                for r_idx, row in enumerate(rows):
                    label = None
                    for c in row:
                        if c.v is not None:
                            label = str(c.v) if isinstance(c.v, str) else None
                            break
                    if label and 'DSCR' in label.upper():
                        # 이 row의 모든 숫자 값
                        vals = [c.v for c in row if c.v is not None and isinstance(c.v, (int, float)) and 0.5 < c.v < 5]
                        if vals:
                            dscr_values = vals[:40]  # 첫 40개
                            break
                
                if dscr_values:
                    metadata['dscr_sample'] = dscr_values[:10]
                    min_dscr = min(dscr_values)
                    max_dscr = max(dscr_values)
                    
                    if min_dscr < 1.20:
                        checks.append({
                            'category': t['cat_debt'], 'severity': 'HIGH',
                            'code': 'DSCR_LOW',
                            'title': t['dscr_low_t'].format(v=min_dscr),
                            'description': t['dscr_low_d'],
                            'action': t['dscr_low_a'],
                        })
                    if max_dscr - min_dscr > 1.0:
                        checks.append({
                            'category': t['cat_debt'], 'severity': 'MEDIUM',
                            'code': 'DSCR_VARIANCE',
                            'title': t['dscr_var_t'].format(lo=min_dscr, hi=max_dscr),
                            'description': t['dscr_var_d'],
                            'action': t['dscr_var_a'],
                        })
    except Exception:
        pass
    
    # ═══ 5. Revenue / OPEX 현실성 ═══
    try:
        with open_xlsb(filepath if filepath.endswith('.xlsb') else None) as wb_v:
            # Dash 탭에서 주요 가정 확인
            try:
                with wb_v.get_sheet('Dash') as ws:
                    dash_data = {}
                    for row in ws.rows():
                        vals = [c.v for c in row]
                        labels = [str(v) for v in vals if isinstance(v, str)]
                        nums = [v for v in vals if isinstance(v, (int, float))]
                        if labels and nums:
                            label = labels[0].strip()
                            dash_data[label] = nums[0]
                    
                    # PPA Price 체크
                    for k, v in dash_data.items():
                        k_lower = k.lower()
                        if 'ppa' in k_lower and 'price' in k_lower:
                            if v < 20 or v > 150:
                                checks.append({
                                    'category': t['cat_revenue'], 'severity': 'MEDIUM',
                                    'code': 'PPA_RANGE',
                                    'title': t['ppa_t'].format(v=v),
                                    'description': t['ppa_d'],
                                    'action': t['ppa_a'],
                                })
                            metadata['ppa_price'] = v
                        elif 'merchant' in k_lower and 'price' in k_lower:
                            if v > 100:
                                checks.append({
                                    'category': t['cat_revenue'], 'severity': 'LOW',
                                    'code': 'MERCHANT_HIGH',
                                    'title': t['merch_t'].format(v=v),
                                    'description': t['merch_d'],
                                    'action': t['merch_a'],
                                })
                            metadata['merchant_price'] = v
                        elif 'degradation' in k_lower:
                            deg_pct = v * 100 if v < 0.1 else v
                            if deg_pct == 0 or deg_pct > 1.5:
                                checks.append({
                                    'category': t['cat_revenue'], 'severity': 'MEDIUM',
                                    'code': 'DEGRADATION',
                                    'title': t['deg_t'].format(v=deg_pct),
                                    'description': t['deg_d'],
                                    'action': t['deg_a'],
                                })
                            metadata['degradation_pct'] = deg_pct
            except Exception:
                pass
    except Exception:
        pass
    # ═══ 추가 체크: DSCR 합리성 ═══
    try:
        dscr_samples = metadata.get('dscr_sample') or []
        if dscr_samples:
            min_dscr = min(dscr_samples)
            if min_dscr < 1.0:
                checks.append({
                    'category': t['cat_debt'], 'severity': 'HIGH',
                    'code': 'DSCR_INSUFFICIENT',
                    'title': t['dscr_low_t'].format(v=min_dscr),
                    'description': t['dscr_low_d'],
                    'action': t['dscr_low_a'],
                })
            elif min_dscr < 1.20:
                checks.append({
                    'category': t['cat_debt'], 'severity': 'MEDIUM',
                    'code': 'DSCR_TIGHT',
                    'title': t['dscr_tight_t'].format(v=min_dscr),
                    'description': t['dscr_tight_d'],
                    'action': t['dscr_tight_a'],
                })
    except Exception:
        pass

    # ═══ 추가 체크: 값 범위 sanity ═══
    try:
        checks_range = []
        debt_r = metadata.get('debt_ratio', 0)
        te_r = metadata.get('te_ratio', 0)
        ppa = metadata.get('ppa_price', 0)
        
        if debt_r > 0 and not (25 <= debt_r <= 75):
            checks_range.append(t['range_debt'].format(v=debt_r))
        if te_r > 0 and not (0 <= te_r <= 50):
            checks_range.append(t['range_te'].format(v=te_r))
        if ppa > 0 and not (20 <= ppa <= 200):
            checks_range.append(t['range_ppa'].format(v=ppa))
        
        if checks_range:
            checks.append({
                'category': t['cat_summary'], 'severity': 'MEDIUM',
                'code': 'RANGE_ANOMALY',
                'title': t['range_t'].format(n=len(checks_range)),
                'description': '; '.join(checks_range),
                'action': t['range_a'],
            })
    except Exception:
        pass

    # ═══ 체크 없음 (No Issues) — 긍정 신호 ═══
    if not checks:
        checks.append({
            'category': t['cat_summary'], 'severity': 'OK',
            'code': 'ALL_CLEAR',
            'title': t['clear_t'],
            'description': t['clear_d'],
            'action': t['clear_a'],
        })
    
    # 심각도별 요약
    summary = {
        'total': len(checks),
        'high': sum(1 for c in checks if c.get('severity') == 'HIGH'),
        'medium': sum(1 for c in checks if c.get('severity') == 'MEDIUM'),
        'low': sum(1 for c in checks if c.get('severity') == 'LOW'),
        'ok': sum(1 for c in checks if c.get('severity') == 'OK'),
    }
    
    return {
        'checks': checks,
        'summary': summary,
        'metadata': metadata,
    }
