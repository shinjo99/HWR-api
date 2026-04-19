from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

MACRS_5YR = [0.20, 0.32, 0.192, 0.1152, 0.1152, 0.0576]

def _irr_robust(cfs, guess=0.08):
    """여러 초기값으로 Newton 반복 → 양수 수렴값 반환.
    Overflow 방지: r은 [-0.99, 10.0] 범위로 클램핑."""
    import numpy as np
    def newton(g):
        r = g
        for _ in range(2000):
            # r이 너무 극단값이면 cf/(1+r)**t overflow — 안전 범위 클램핑
            if r <= -0.99: r = -0.99
            elif r >= 10.0: r = 10.0
            try:
                npv  = sum(cf/(1+r)**t for t,cf in enumerate(cfs))
                dnpv = sum(-t*cf/(1+r)**(t+1) for t,cf in enumerate(cfs))
            except (OverflowError, ZeroDivisionError):
                return None  # 발산 — 실패 신호
            if abs(dnpv) < 1e-12: break
            r_new = r - npv/dnpv
            # 한 step 이동 제한 (안정화)
            if r_new <= -0.99: r_new = -0.99
            elif r_new >= 10.0: r_new = 10.0
            if abs(r_new - r) < 1e-8: return r_new
            r = r_new
        return r
    for g in [guess, 0.01, 0.03, 0.05, 0.10, 0.15, 0.20, -0.05]:
        r = newton(g)
        if r is None: continue  # 발산 스킵
        if -0.5 < r < 5.0:  # 합리적 IRR 범위
            try:
                chk = sum(cf/(1+r)**t for t,cf in enumerate(cfs))
                if abs(chk) < 500:  # $500K 오차 허용
                    return r
            except (OverflowError, ZeroDivisionError):
                continue
    # numpy_financial fallback
    try:
        import numpy_financial as npf
        r0 = float(npf.irr(cfs))
        if not np.isnan(r0) and -0.5 < r0 < 5.0:
            return r0
    except Exception:
        pass
    return None  # 진짜 해 없음 (caller가 처리)
import requests
import jwt
import os
import datetime
import json
import tempfile
from pyxlsb import open_workbook

app = FastAPI(title="HWR Dashboard API")

# ── CORS (대시보드에서 호출 허용) ──────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 추후 shinjo99.github.io 로 제한
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── 설정 ──────────────────────────────────────────
JWT_SECRET     = os.environ.get("JWT_SECRET", "hwr-secret-change-this")
ANTHROPIC_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")
FB_URL      = os.environ.get("FB_URL", "https://team-dashboard-c0d7b-default-rtdb.asia-southeast1.firebasedatabase.app")
FB_SECRET   = os.environ.get("FB_SECRET", "")  # Firebase Database Secret
FRED_API_KEY = os.environ.get("FRED_API_KEY", "")  # https://fred.stlouisfed.org/docs/api/api_key.html

# ── 사용자 계정 (환경변수로 관리) ─────────────────
# 형식: "email:password:role,email:password:role"
# 예: "team@hwr.com:hanwha2024:viewer,admin@hwr.com:admin1234:admin"
def get_users():
    raw = os.environ.get("USERS", "team@hwr.com:hanwha2024:viewer")
    users = {}
    for entry in raw.split(","):
        parts = entry.strip().split(":")
        if len(parts) >= 3:
            email, password, role = parts[0], parts[1], parts[2]
            users[email] = {"password": password, "role": role}
    return users

# ── JWT 헬퍼 ──────────────────────────────────────
def create_token(email: str, role: str) -> str:
    payload = {
        "email": email,
        "role": role,
        "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=24)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

def verify_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="토큰이 만료되었습니다. 다시 로그인하세요.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="유효하지 않은 토큰입니다.")

security = HTTPBearer()

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    return verify_token(credentials.credentials)

def require_admin(user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")
    return user

# ── Firebase 헬퍼 ──────────────────────────────────
def fb_auth_param():
    if FB_SECRET:
        return {"auth": FB_SECRET}
    return {}

def fb_read(path: str):
    try:
        res = requests.get(
            f"{FB_URL}/{path}.json",
            params=fb_auth_param(),
            timeout=5
        )
        return res.json() or {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB 읽기 오류: {str(e)}")

def fb_write(path: str, data: dict):
    try:
        res = requests.patch(
            f"{FB_URL}/{path}.json",
            json=data,
            params=fb_auth_param(),
            timeout=5
        )
        if res.status_code != 200:
            raise HTTPException(status_code=500, detail="DB 저장 실패")
        return res.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB 쓰기 오류: {str(e)}")

def fb_put(path: str, data: dict):
    try:
        res = requests.put(
            f"{FB_URL}/{path}.json",
            json=data,
            params=fb_auth_param(),
            timeout=5
        )
        if res.status_code != 200:
            raise HTTPException(status_code=500, detail="DB 저장 실패")
        return res.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB 쓰기 오류: {str(e)}")

def fb_patch(path: str, data: dict):
    """Firebase PATCH — 필드 일부만 업데이트"""
    try:
        requests.patch(
            f"{FB_URL}/{path}.json",
            json=data,
            params=fb_auth_param(),
            timeout=5
        )
    except Exception:
        pass

# ══════════════════════════════════════════════════
#  인증
# ══════════════════════════════════════════════════
class LoginRequest(BaseModel):
    email: str
    password: str

class ValuationCalcRequest(BaseModel):
    project_id: str = ""
    inputs: dict = {}

@app.post("/auth/login")
def login(req: LoginRequest):
    users = get_users()
    user = users.get(req.email)
    if not user or user["password"] != req.password:
        raise HTTPException(status_code=401, detail="이메일 또는 비밀번호가 틀렸습니다.")
    token = create_token(req.email, user["role"])
    return {
        "token": token,
        "email": req.email,
        "role": user["role"],
        "expires_in": 86400
    }

@app.get("/auth/me")
def me(user=Depends(get_current_user)):
    return user

@app.get("/auth/admins")
def get_admins(user=Depends(get_current_user)):
    """승인자 드롭다운용 admin 목록. 로그인한 사용자라면 조회 가능."""
    users = get_users()
    admins = [email for email, u in users.items() if u.get("role") == "admin"]
    return {"admins": sorted(admins)}

# ══════════════════════════════════════════════════
#  PPV
# ══════════════════════════════════════════════════
@app.get("/ppv")
def get_ppv(user=Depends(get_current_user)):
    return fb_read("ppv")

@app.get("/ppv/summary")
def get_ppv_summary(user=Depends(get_current_user)):
    return fb_read("ppv/summary")

class PPVSummary(BaseModel):
    totalRisked: float
    byStage: dict
    projectCount: int

@app.post("/ppv/summary")
def save_ppv_summary(data: PPVSummary, user=Depends(get_current_user)):
    payload = data.dict()
    payload["updatedAt"] = datetime.datetime.now().isoformat()
    payload["updatedBy"] = user["email"]
    fb_write("ppv/summary", payload)
    return {"ok": True, "data": payload}

@app.post("/ppv/snapshot")
def save_snapshot(data: dict, user=Depends(get_current_user)):
    data["ts"] = datetime.datetime.now().isoformat()
    data["by"] = user["email"]
    requests.post(
        f"{FB_URL}/ppv/snapshots.json",
        json=data,
        params=fb_auth_param(),
        timeout=5
    )
    return {"ok": True}

@app.post("/ppv/event")
def save_event(data: dict, user=Depends(get_current_user)):
    data["ts"] = datetime.datetime.now().isoformat()
    data["by"] = user["email"]
    requests.post(
        f"{FB_URL}/ppv/events.json",
        json=data,
        params=fb_auth_param(),
        timeout=5
    )
    return {"ok": True}

@app.post("/ppv/override/{project_name}")
def save_override(project_name: str, data: dict, user=Depends(get_current_user)):
    safe_name = project_name.replace("/", "_").replace(".", "_")
    data["updatedAt"] = datetime.datetime.now().isoformat()
    data["updatedBy"] = user["email"]
    fb_write(f"ppv/overrides/{safe_name}", data)
    return {"ok": True}

# ══════════════════════════════════════════════════
#  재무 (P&L / B/S / C/F)
# ══════════════════════════════════════════════════
@app.get("/financial")
def get_financial(user=Depends(get_current_user)):
    return fb_read("financial")

@app.get("/financial/{stmt}")
def get_stmt(stmt: str, user=Depends(get_current_user)):
    if stmt not in ["pl", "bs", "cf"]:
        raise HTTPException(status_code=400, detail="stmt는 pl, bs, cf 중 하나여야 합니다.")
    return fb_read(f"financial/{stmt}")

@app.get("/financial/{stmt}/{year}")
def get_stmt_year(stmt: str, year: int, user=Depends(get_current_user)):
    if stmt not in ["pl", "bs", "cf"]:
        raise HTTPException(status_code=400, detail="stmt는 pl, bs, cf 중 하나여야 합니다.")
    return fb_read(f"financial/{stmt}/{year}")

class FinancialData(BaseModel):
    year: int
    month: int
    data: dict

@app.post("/financial/{stmt}")
def save_financial(stmt: str, req: FinancialData, user=Depends(get_current_user)):
    if stmt not in ["pl", "bs", "cf"]:
        raise HTTPException(status_code=400, detail="stmt는 pl, bs, cf 중 하나여야 합니다.")
    payload = req.data.copy()
    payload["updatedAt"] = datetime.datetime.now().isoformat()[:16]
    payload["updatedBy"] = user["email"]
    fb_write(f"financial/{stmt}/{req.year}/{req.month}", payload)
    return {"ok": True, "path": f"financial/{stmt}/{req.year}/{req.month}", "data": payload}

# ══════════════════════════════════════════════════
#  매각 현황
# ══════════════════════════════════════════════════
@app.get("/divest")
def get_divest(user=Depends(get_current_user)):
    return fb_read("divest")

@app.post("/divest/{project_name}")
def update_divest(project_name: str, data: dict, user=Depends(get_current_user)):
    safe_name = project_name.replace("/", "_").replace(".", "_")
    data["updatedAt"] = datetime.datetime.now().isoformat()[:16]
    data["updatedBy"] = user["email"]
    fb_write(f"divest/{safe_name}", data)
    return {"ok": True}

# ══════════════════════════════════════════════════
#  Atlas Milestone
# ══════════════════════════════════════════════════
@app.get("/atlas")
def get_atlas(user=Depends(get_current_user)):
    return fb_read("atlas")

@app.post("/atlas/{milestone_id}")
def update_atlas(milestone_id: str, data: dict, user=Depends(get_current_user)):
    data["updatedAt"] = datetime.datetime.now().isoformat()[:16]
    data["updatedBy"] = user["email"]
    fb_write(f"atlas/{milestone_id}", data)
    return {"ok": True}

# ══════════════════════════════════════════════════
#  전체 데이터 (대시보드 초기 로딩용)
# ══════════════════════════════════════════════════
@app.get("/dashboard")
def get_dashboard(user=Depends(get_current_user)):
    return {
        "ppv_summary": fb_read("ppv/summary"),
        "financial": fb_read("financial"),
        "divest": fb_read("divest"),
        "atlas": fb_read("atlas"),
    }

# ══════════════════════════════════════════════════
#  외부 시장 벤치마크 (FRED + LevelTen)
# ══════════════════════════════════════════════════

# FRED 시리즈 매핑: HEUH 투자 의사결정에 의미있는 지표만
FRED_SERIES = {
    # 금리 (할인율 기준점)
    "us_10y":       {"id": "DGS10",        "label": "US 10Y Treasury",    "unit": "%",       "group": "rates"},
    "us_2y":        {"id": "DGS2",         "label": "US 2Y Treasury",     "unit": "%",       "group": "rates"},
    "fed_funds":    {"id": "DFF",          "label": "Fed Funds Rate",     "unit": "%",       "group": "rates"},
    "bbb_spread":   {"id": "BAMLC0A4CBBB", "label": "BBB Corp Spread",    "unit": "%",       "group": "rates"},
    # 에너지/인플레
    "henry_hub":    {"id": "DHHNGSP",      "label": "Henry Hub NatGas",   "unit": "$/MMBtu", "group": "energy"},
    "cpi":          {"id": "CPIAUCSL",     "label": "US CPI (Index)",     "unit": "Index",   "group": "macro"},
    # 환율
    "krw_usd":      {"id": "DEXKOUS",      "label": "KRW/USD",            "unit": "KRW",     "group": "fx"},
}

# TAN ETF는 FRED에 없음 — Stooq로 별도 조회
STOOQ_SYMBOLS = {
    "tan":          {"symbol": "tan.us",   "label": "TAN (Solar ETF)",    "unit": "$",       "group": "equity"},
    "icln":         {"symbol": "icln.us",  "label": "ICLN (Clean Energy)", "unit": "$",      "group": "equity"},
}

def _fred_fetch(series_id: str, days: int = 180):
    """FRED API에서 시리즈 데이터 조회. 최근 N일치."""
    if not FRED_API_KEY:
        return None
    try:
        end = datetime.date.today()
        start = end - datetime.timedelta(days=days)
        res = requests.get(
            "https://api.stlouisfed.org/fred/series/observations",
            params={
                "series_id": series_id,
                "api_key": FRED_API_KEY,
                "file_type": "json",
                "observation_start": start.isoformat(),
                "observation_end": end.isoformat(),
                "sort_order": "asc",
            },
            timeout=10,
        )
        if res.status_code != 200:
            return None
        obs = res.json().get("observations", [])
        # "." = 결측, 제외
        points = [
            {"date": o["date"], "value": float(o["value"])}
            for o in obs if o.get("value") not in (".", "", None)
        ]
        return points
    except Exception:
        return None

def _stooq_fetch(symbol: str, days: int = 180):
    """Stooq에서 일별 종가 조회 (CSV, 키 불필요)."""
    try:
        url = f"https://stooq.com/q/d/l/?s={symbol}&i=d"
        res = requests.get(url, timeout=10)
        if res.status_code != 200 or "Date,Open" not in res.text:
            return None
        lines = res.text.strip().split("\n")[1:]
        end = datetime.date.today()
        start = end - datetime.timedelta(days=days)
        points = []
        for ln in lines:
            parts = ln.split(",")
            if len(parts) < 5:
                continue
            try:
                d = datetime.date.fromisoformat(parts[0])
                if d < start:
                    continue
                points.append({"date": parts[0], "value": float(parts[4])})
            except Exception:
                continue
        return points
    except Exception:
        return None

def _summarize_series(points):
    """시계열 → 최신값/변동/스파크라인 요약 (1년치)."""
    if not points:
        return None
    latest = points[-1]
    prev = points[-2] if len(points) >= 2 else latest
    # 약 일주일 전 (영업일 5개 전)
    week_ago = points[-6] if len(points) >= 6 else points[0]
    # 약 한달 전 (영업일 21개 전)
    month_ago = points[-22] if len(points) >= 22 else points[0]
    # 약 1년 전 (영업일 252개 전)
    year_ago = points[-253] if len(points) >= 253 else points[0]
    # 1년치 주간 샘플링 (52포인트 내외) — 주 1회 데이터만 추출
    # 영업일 기준 5일마다 1개 선택
    sampled = points[::5] if len(points) > 60 else points
    return {
        "latest": latest["value"],
        "latest_date": latest["date"],
        "d_1d": latest["value"] - prev["value"],
        "d_1w": latest["value"] - week_ago["value"],
        "d_1m": latest["value"] - month_ago["value"],
        "d_1y": latest["value"] - year_ago["value"],
        "spark": [p["value"] for p in sampled],
        "spark_dates": [p["date"] for p in sampled],  # 실제 날짜 병행 전달
        "n_points": len(points),
    }

@app.get("/benchmark/market")
def get_market_benchmark(force: int = 0, user=Depends(get_current_user)):
    """FRED + Stooq 시장 벤치마크. 6시간 캐시."""
    today = datetime.date.today().isoformat()
    cache_key = f"benchmark_cache/market/{today}"

    # 캐시 확인 (force=1이면 무시)
    if not force:
        cached = fb_read(cache_key)
        if cached and cached.get("fetched_at"):
            try:
                fetched = datetime.datetime.fromisoformat(cached["fetched_at"])
                age_hrs = (datetime.datetime.utcnow() - fetched).total_seconds() / 3600
                if age_hrs < 6:
                    return cached
            except Exception:
                pass

    if not FRED_API_KEY:
        raise HTTPException(500, "FRED_API_KEY 환경변수 미설정")

    result = {
        "fetched_at": datetime.datetime.utcnow().isoformat()[:19],
        "source": "FRED + Stooq",
        "series": {},
    }

    # FRED 시리즈 (1년치)
    for key, meta in FRED_SERIES.items():
        pts = _fred_fetch(meta["id"], days=400 if key == "cpi" else 365)
        summary = _summarize_series(pts) if pts else None
        result["series"][key] = {
            **meta,
            "data": summary,
            "ok": summary is not None,
        }

    # Stooq 시리즈 (1년치)
    for key, meta in STOOQ_SYMBOLS.items():
        pts = _stooq_fetch(meta["symbol"], days=365)
        summary = _summarize_series(pts) if pts else None
        result["series"][key] = {
            **meta,
            "data": summary,
            "ok": summary is not None,
        }

    # CPI는 YoY % 변화율로 계산 (Index 자체는 의미가 없음)
    cpi = result["series"].get("cpi", {}).get("data")
    if cpi and cpi.get("spark") and len(cpi["spark"]) >= 13:
        try:
            yoy = (cpi["spark"][-1] / cpi["spark"][-13] - 1) * 100
            result["series"]["cpi"]["yoy_pct"] = round(yoy, 2)
        except Exception:
            pass

    # 캐시 저장
    try:
        fb_put(cache_key, result)
    except Exception:
        pass  # 캐시 실패해도 응답은 반환
    return result


# ── LevelTen PPA Index 업로드/조회 ─────────────────
@app.post("/benchmark/levelten/upload")
async def upload_levelten(
    file: UploadFile = File(...),
    quarter: str = Form(...),  # e.g. "2026-Q1"
    user=Depends(get_current_user),
):
    """LevelTen PPA Index 리포트 업로드 → Claude API로 파싱 → Firebase 저장."""
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 환경변수 미설정")

    raw = await file.read()
    filename = file.filename or "levelten.pdf"
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    # 1) 파일 타입별 텍스트 추출
    source_text = ""
    parse_mode = ""
    parsed = None

    if ext in ("csv", "tsv"):
        try:
            source_text = raw.decode("utf-8", errors="ignore")
        except Exception:
            source_text = raw.decode("latin-1", errors="ignore")
        parse_mode = "csv"

    elif ext in ("xlsx", "xls", "xlsb"):
        parse_mode = "excel"
        tmp_path = None
        try:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}")
            tmp_path = tmp.name
            tmp.write(raw)
            tmp.close()

            if ext == "xlsb":
                lines = []
                with open_workbook(tmp_path) as wb:
                    for sheet_name in wb.sheets[:5]:
                        lines.append(f"\n=== Sheet: {sheet_name} ===")
                        with wb.get_sheet(sheet_name) as sh:
                            for row in sh.rows():
                                vals = [str(c.v) if c.v is not None else "" for c in row]
                                lines.append("\t".join(vals))
                source_text = "\n".join(lines)[:40000]
            else:
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    raise HTTPException(400, "openpyxl이 설치되지 않았습니다. requirements.txt에 추가하세요.")
                wb = load_workbook(tmp_path, data_only=True, read_only=True)
                lines = []
                for sheet_name in wb.sheetnames[:5]:
                    lines.append(f"\n=== Sheet: {sheet_name} ===")
                    for row in wb[sheet_name].iter_rows(values_only=True):
                        vals = [str(v) if v is not None else "" for v in row]
                        lines.append("\t".join(vals))
                wb.close()
                source_text = "\n".join(lines)[:40000]
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Excel 파싱 실패: {str(e)}")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try: os.unlink(tmp_path)
                except Exception: pass

    elif ext == "pdf":
        parse_mode = "pdf"
        import base64
        pdf_b64 = base64.standard_b64encode(raw).decode("utf-8")

        headers = {
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        prompt = (
            "You are parsing a LevelTen Energy PPA Price Index report for a Solar+BESS developer. "
            "Extract structured data into strict JSON (no markdown, no prose).\n\n"
            "IMPORTANT PRINCIPLES:\n"
            "1. We focus on Solar PPA and BESS Storage pricing. Skip Wind data entirely.\n"
            "2. ONLY extract values that are explicitly present in the report — tables, charts, or text.\n"
            "3. DO NOT estimate, guess, or use general market knowledge to fill missing values.\n"
            "4. If a value is not in the report, use null. It is BETTER to return null than to invent data.\n"
            "5. For chart-read values (Storage Price Spreads typically shown as charts only), estimate to nearest $0.5 and mark source as 'chart_read'.\n\n"

            "Required schema:\n"
            "{\n"
            '  "quarter": "YYYY-QN",\n'
            '  "report_date": "YYYY-MM-DD or null",\n'
            '  "solar_iso": [\n'
            '    {"region":"ERCOT|PJM|MISO|CAISO|SPP|NYISO|ISO-NE|AESO", "p25":<number $/MWh>,\n'
            '     "qoq_pct":<number or null>, "yoy_pct":<number or null>}\n'
            "  ],\n"
            '  "solar_continental": {\n'
            '    "p25":<number>, "p50":<number>, "p75":<number>,\n'
            '    "p10":<number or null>, "p90":<number or null>,\n'
            '    "qoq_pct":<number or null>, "yoy_pct":<number or null>\n'
            "  },\n"
            '  "solar_hub": [\n'
            '    {"region":"ERCOT", "hub":"HB_NORTH|HB_WEST|HB_SOUTH|HB_HOUSTON|SP15|Alberta|WESTERN HUB|DOM|N ILLINOIS HUB|AEP-DAYTON HUB|SPPNORTH_HUB|SPPSOUTH_HUB|MINN.HUB|ILLINOIS.HUB|INDIANA.HUB|LOUISIANA.HUB|ARKANSAS.HUB|...", "p25":<number>}\n'
            "  ],\n"
            '  "storage_iso": [\n'
            '    {"region":"AESO|CAISO|ERCOT|MISO|PJM|SPP|...", \n'
            '     "min":<number or null>, "p25":<number or null>, "median":<number or null>, "p75":<number or null>, "max":<number or null>,\n'
            '     "source":"levelten_index|chart_read"}\n'
            "  ],\n"
            '  "storage_duration_mix": [\n'
            '    {"region":"ERCOT", "2h":<pct or null>, "3h":<pct or null>, "4h":<pct or null>, "6h":<pct or null>, "8h":<pct or null>, "10h":<pct or null>}\n'
            "  ],\n"
            '  "solar_psv": [\n'
            '    {"region":"ERCOT", "psv_median":<number $/MWh>, "psv_min":<number>, "psv_max":<number>}\n'
            "  ],\n"
            '  "pipeline_breakdown": [\n'
            '    {"cod_year":"2025|2026|2027|2028|2029|2030+", "solar_mw":<number>, "standalone_storage_mw":<number>, "hybrid_mw":<number>}\n'
            "  ],\n"
            '  "storage_available": <true if BESS data found in report, false otherwise>,\n'
            '  "storage_note": "description of BESS data source",\n'
            '  "key_insights": ["1-line insight 1", "1-line insight 2", ...],\n'
            '  "notes": "2-3 sentence summary of quarter trends (Solar + Storage focus)"\n'
            "}\n\n"

            "CRITICAL — Storage Extraction:\n"
            "- LevelTen's 'Storage Price Spreads by ISO' is a BOX PLOT chart showing MIN, P25, MEDIAN, P75, MAX for each ISO.\n"
            "- Read ALL 5 statistics from the box plot. Round to nearest $0.5. Mark source='levelten_index'.\n"
            "- These are LEVELIZED TOLLING AGREEMENT prices in $/kW-month (confirmed by methodology).\n"
            "- Typical ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP (6 ISOs). ISO-NE and NYISO NOT covered by LevelTen Storage Index.\n"
            "- Also extract 'Storage Duration Distribution by ISO' chart → percent for each duration (2h, 3h, 4h, 6h, 8h, 10h).\n\n"

            "CRITICAL — Hub-level Solar P25 Extraction:\n"
            "- Every ISO has a 'PPA Prices by Hub' section showing maps with Solar P25 values labeled on each hub.\n"
            "- Extract every hub + price combination. Examples:\n"
            "  - ERCOT: HB_NORTH, HB_WEST, HB_SOUTH, HB_HOUSTON (4 hubs)\n"
            "  - CAISO: SP15 (1 hub)\n"
            "  - MISO: MINN.HUB, ILLINOIS.HUB, INDIANA.HUB, LOUISIANA.HUB, ARKANSAS.HUB (5 hubs)\n"
            "  - PJM: WESTERN HUB, DOM, AEP-DAYTON HUB, N ILLINOIS HUB (4 hubs)\n"
            "  - SPP: SPPNORTH_HUB, SPPSOUTH_HUB (2 hubs)\n"
            "  - AESO: Alberta (1 hub)\n\n"

            "CRITICAL — Solar PSV (Projected Settlement Value):\n"
            "- Report has 'Projected Settlement Values by Market: Solar' box plot chart.\n"
            "- Read median, min, max for each ISO shown. Values are in $/MWh (can be NEGATIVE).\n"
            "- Typical ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP.\n\n"

            "CRITICAL — Pipeline Breakdown:\n"
            "- Report has 'Technology Breakdown of Pipelines by COD Year' bar chart (in 'Going Hybrid' section).\n"
            "- Extract MW values for each year × technology. Include Solar, Standalone Storage, Hybrid.\n"
            "- DO NOT include Wind.\n\n"

            "General Rules:\n"
            "- Solar prices: USD/MWh. Storage prices: USD/kW-month.\n"
            "- If data not available for a field, use null. NEVER invent numbers.\n"
            "- DO NOT include any Wind data anywhere in the output.\n"
            "- Return ONLY the JSON object. No explanation. No code fences."
        )
        body = {
            "model": "claude-sonnet-4-5",
            "max_tokens": 4000,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "document", "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_b64,
                    }},
                    {"type": "text", "text": prompt},
                ],
            }],
        }
        try:
            res = requests.post("https://api.anthropic.com/v1/messages",
                                headers=headers, json=body, timeout=150)
            if res.status_code != 200:
                raise HTTPException(502, f"Claude API 오류: {res.text[:300]}")
            ai_text = res.json()["content"][0]["text"].strip()
            if ai_text.startswith("```"):
                ai_text = ai_text.split("```", 2)[1]
                if ai_text.startswith("json"): ai_text = ai_text[4:]
                ai_text = ai_text.rsplit("```", 1)[0]
            parsed = json.loads(ai_text.strip())
        except HTTPException:
            raise
        except requests.Timeout:
            raise HTTPException(504, "Claude API 응답 타임아웃 — 리포트가 너무 크거나 서버 혼잡 (재시도 권장)")
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"AI 응답 JSON 파싱 실패: {str(e)}")
    else:
        raise HTTPException(400, f"지원하지 않는 파일 형식: .{ext} (PDF, CSV, XLSX, XLSB 지원)")

    # CSV/Excel인 경우 Claude에게 텍스트 파싱 요청
    if parse_mode in ("csv", "excel"):
        headers = {
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        prompt = (
            "You are parsing LevelTen Energy PPA Price Index tabular data for a Solar+BESS developer. "
            f"The data below is from a {parse_mode.upper()} export.\n\n"
            f"DATA:\n{source_text[:30000]}\n\n"
            "IMPORTANT PRINCIPLES:\n"
            "1. We focus on Solar PPA and BESS Storage pricing. Skip Wind data.\n"
            "2. ONLY extract values that are EXPLICITLY present in the data. NEVER estimate or invent.\n"
            "3. If a value is missing, use null. Do NOT fill with guesses.\n\n"
            "Extract into strict JSON (no markdown, no prose):\n"
            "{\n"
            '  "quarter": "YYYY-QN",\n'
            '  "report_date": "YYYY-MM-DD or null",\n'
            '  "solar_iso": [{"region":"ERCOT|PJM|MISO|CAISO|SPP|ISO-NE|AESO", "p25":<$/MWh>, "qoq_pct":<null>, "yoy_pct":<null>}],\n'
            '  "solar_continental": {"p25":<number>, "p50":<number>, "p75":<number>, "p10":<number|null>, "p90":<number|null>, "qoq_pct":<null>, "yoy_pct":<null>},\n'
            '  "solar_hub": [{"region":"ERCOT", "hub":"North", "p25":<number>}],\n'
            '  "storage_iso": [{"region":"ERCOT|...", "p25":<$/kW-month|null>, "p50":<number|null>, "p75":<number|null>, "source":"table"}],\n'
            '  "storage_available": <true|false>,\n'
            '  "storage_note": "description or \\"Not included in data\\"",\n'
            '  "key_insights": ["actionable insight 1", "insight 2"],\n'
            '  "notes": "2-3 sentence summary (Solar+Storage focus, no Wind)"\n'
            "}\n"
            "If NO storage data in file: storage_iso=[], storage_available=false.\n"
            "Return ONLY the JSON object. No code fences."
        )
        body = {
            "model": "claude-sonnet-4-5",
            "max_tokens": 4000,
            "messages": [{"role": "user", "content": prompt}],
        }
        try:
            res = requests.post("https://api.anthropic.com/v1/messages",
                                headers=headers, json=body, timeout=150)
            if res.status_code != 200:
                raise HTTPException(502, f"Claude API 오류: {res.text[:300]}")
            ai_text = res.json()["content"][0]["text"].strip()
            if ai_text.startswith("```"):
                ai_text = ai_text.split("```", 2)[1]
                if ai_text.startswith("json"): ai_text = ai_text[4:]
                ai_text = ai_text.rsplit("```", 1)[0]
            parsed = json.loads(ai_text.strip())
        except HTTPException:
            raise
        except requests.Timeout:
            raise HTTPException(504, "Claude API 응답 타임아웃 — 리포트가 너무 크거나 서버 혼잡 (재시도 권장)")
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"AI 응답 JSON 파싱 실패: {str(e)}")

    if not parsed:
        raise HTTPException(500, "파싱 결과가 비어있습니다.")

    # 쿼터 형식 검증 (YYYY-QN)
    import re
    if not re.match(r'^\d{4}-Q[1-4]$', quarter.upper()):
        raise HTTPException(400, "쿼터 형식은 YYYY-Q1 ~ YYYY-Q4 여야 합니다.")
    quarter = quarter.upper()

    # 쿼터 덮어쓰기 (사용자가 명시한 값이 우선)
    parsed["quarter"] = quarter
    parsed["uploaded_at"] = datetime.datetime.utcnow().isoformat()[:19]
    parsed["uploaded_by"] = user["email"]
    parsed["filename"] = filename
    parsed["parse_mode"] = parse_mode

    # Backward compat: 새 파서 스키마를 legacy entries 배열로도 변환
    if "entries" not in parsed:
        legacy = []
        for s in parsed.get("solar_iso", []) or []:
            legacy.append({"tech":"solar", "region": s.get("region",""), "term_yr":10,
                           "p25": s.get("p25"), "p50": None, "p75": None})
        for s in parsed.get("storage_iso", []) or []:
            # 새 스키마: min/p25/median/p75/max → legacy: p25/p50/p75
            legacy.append({"tech":"storage", "region": s.get("region",""), "term_yr":10,
                           "p25": s.get("p25"),
                           "p50": s.get("median") or s.get("p50"),
                           "p75": s.get("p75")})
        parsed["entries"] = legacy

    # Firebase 저장: benchmark/levelten/{quarter}
    fb_put(f"benchmark/levelten/{quarter}", parsed)

    return {
        "ok": True,
        "quarter": quarter,
        "solar_iso_count": len(parsed.get("solar_iso", []) or []),
        "storage_iso_count": len(parsed.get("storage_iso", []) or []),
        "entries_count": len(parsed.get("entries", []) or []),
        "data": parsed,
    }


@app.get("/benchmark/levelten")
def get_levelten_all(user=Depends(get_current_user)):
    """모든 분기별 LevelTen 데이터."""
    return fb_read("benchmark/levelten") or {}


@app.get("/benchmark/levelten/latest")
def get_levelten_latest(user=Depends(get_current_user)):
    """가장 최신 분기의 LevelTen 데이터."""
    all_data = fb_read("benchmark/levelten") or {}
    if not all_data:
        return {}
    # 쿼터 문자열 정렬 (YYYY-QN 포맷이라 사전순 정렬로 충분)
    latest_key = sorted(all_data.keys())[-1]
    return {"quarter": latest_key, **all_data[latest_key]}


@app.delete("/benchmark/levelten/{quarter}")
def delete_levelten(quarter: str, user=Depends(require_admin)):
    """특정 분기 LevelTen 데이터 삭제."""
    import re
    if not re.match(r'^\d{4}-Q[1-4]$', quarter.upper()):
        raise HTTPException(400, "쿼터 형식은 YYYY-Q1 ~ YYYY-Q4 여야 합니다.")
    quarter = quarter.upper()
    try:
        requests.delete(f"{FB_URL}/benchmark/levelten/{quarter}.json",
                        params=fb_auth_param(), timeout=5)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, f"삭제 오류: {str(e)}")


# ── 피어 IRR 벤치마크 (내부 수동 입력값) ────────────
@app.get("/benchmark/peer-irr")
def get_peer_irr(user=Depends(get_current_user)):
    """저장된 피어 IRR 벤치마크 조회."""
    return fb_read("benchmark/peer_irr") or {}


@app.post("/benchmark/peer-irr")
def save_peer_irr(payload: dict, user=Depends(get_current_user)):
    """피어 IRR 벤치마크 저장 (Levered Pre-Tax IRR 레인지)."""
    # 필드 검증
    required_numeric = ["solar_min", "solar_max", "hybrid_min", "hybrid_max", "wind_min", "wind_max"]
    data = {}
    for k in required_numeric:
        v = payload.get(k)
        if v is None:
            raise HTTPException(400, f"누락된 필드: {k}")
        try:
            fv = float(v)
            if fv < 0 or fv > 50:
                raise HTTPException(400, f"{k}: 0~50% 범위여야 합니다.")
            data[k] = round(fv, 2)
        except (ValueError, TypeError):
            raise HTTPException(400, f"{k}: 숫자여야 합니다.")
    # min < max 검증
    for tech in ("solar", "hybrid", "wind"):
        if data[f"{tech}_min"] >= data[f"{tech}_max"]:
            raise HTTPException(400, f"{tech}: min < max 이어야 합니다.")
    # 비고
    note = payload.get("note", "")
    if isinstance(note, str):
        data["note"] = note[:200]
    data["updated_at"] = datetime.datetime.utcnow().isoformat()[:19]
    data["updated_by"] = user["email"]
    fb_put("benchmark/peer_irr", data)
    return {"ok": True, "data": data}


# ══════════════════════════════════════════════════
#  BESS Tolling Market Research (AI Web Search)
# ══════════════════════════════════════════════════
@app.post("/benchmark/bess-tolling/research")
def research_bess_tolling(user=Depends(get_current_user)):
    """
    Claude API + web_search 도구로 ISO별 BESS tolling 가격을 실시간 리서치.
    결과: ISO × Duration 별 P25/P75 + 출처 URL + confidence score.
    캐시: benchmark/bess_tolling/latest (수동 새로고침)
    """
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 미설정")

    today_str = datetime.date.today().isoformat()

    prompt = (
        f"You are an energy market research analyst specializing in US battery energy storage "
        f"system (BESS) tolling agreements AND PPA markets. Today: {today_str}.\n\n"
        "ROLE: This research is COMPLEMENTARY to LevelTen's official PPA Price Index.\n"
        "LevelTen publishes official data for 6 ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP.\n"
        "For these 6 ISOs the dashboard uses LevelTen first — your role is DURATION-level BESS detail only.\n\n"
        "YOUR FOCUS (three objectives):\n"
        "  (A) NON-LEVELTEN ISOs — provide BOTH BESS tolling AND PPA market commentary:\n"
        "      - ISO-NE (New England)\n"
        "      - NYISO (New York)\n"
        "      - WECC_DSW (Desert Southwest: AZ, NM, NV — Arizona/New Mexico/Nevada utilities)\n"
        "      - WECC_RM  (Rocky Mountain: UT, CO, WY, ID — PacifiCorp East/RMP, Xcel Colorado)\n"
        "      - WECC_NW  (Northwest: OR, WA, MT — PacifiCorp West, PGE, Puget Sound Energy)\n"
        "      - SERC (Southeast: TVA, Duke, Southern Company territory — NC, SC, GA, AL, TN, KY)\n"
        "  (B) DURATION BREAKDOWN for LevelTen-covered ISOs (ERCOT, CAISO, PJM, MISO, SPP, AESO):\n"
        "      → duration-level prices (2h / 4h / 6h) — LevelTen only gives ISO-level\n"
        "  (C) For WECC sub-regions: include PPA market commentary since LevelTen has ZERO coverage.\n"
        "      Key utility RFPs to reference: PacifiCorp IRP RFP, URC (Utah Renewable Communities),\n"
        "      APS (Arizona Public Service), NV Energy, Xcel Energy Colorado, Portland General Electric,\n"
        "      Idaho Power, Puget Sound Energy.\n\n"
        "Research methodology — TRIANGULATION:\n"
        "  1. Capacity market clearing prices (PJM, NYISO, ISO-NE) — adjusted for storage\n"
        "  2. Merchant BESS revenue data (ERCOT ~$30-50/kW-yr, CAISO duck curve premium)\n"
        "  3. Utility RFP announcements when prices are disclosed (PacifiCorp, APS, Xcel, etc.)\n"
        "  4. Public company earnings calls (NextEra, Vistra, AES)\n"
        "  5. Duration-adjustment heuristic:\n"
        "     - 2h: 60-75% of 4h price (arbitrage-dominated)\n"
        "     - 4h: reference (capacity-dominated, NERC/ISO standard)\n"
        "     - 6h+: 110-130% of 4h (long-duration premium)\n"
        "  6. Industry rule-of-thumb benchmarks (2025):\n"
        "     - ERCOT 2h: $3-8/kW-mo  | 4h: $5-12/kW-mo\n"
        "     - CAISO 4h: $10-16/kW-mo (duck curve) | 8h: $13-20/kW-mo\n"
        "     - PJM 4h: $8-13/kW-mo (capacity market) | 2h: $5-9/kW-mo\n"
        "     - SPP/MISO 4h: $6-11/kW-mo\n"
        "     - ISO-NE 4h: $12-18/kW-mo (tight capacity, winter peak)\n"
        "     - NYISO 4h: $11-17/kW-mo (DEC mandate, expensive zones J/K)\n"
        "     - WECC_DSW 4h: $7-12/kW-mo (APS/NV Energy RFPs, solar-shifting demand)\n"
        "     - WECC_RM 4h: $6-11/kW-mo (PacifiCorp/Xcel CO — emerging market, thin liquidity)\n"
        "     - WECC_NW 4h: $6-10/kW-mo (hydro-dominant, moderate storage need)\n"
        "     - SERC 4h: $7-12/kW-mo (vertically integrated utilities, bilateral)\n"
        "Use these as STARTING POINTS, then VERIFY/ADJUST via web_search.\n\n"
        "Use web_search to find CURRENT data from:\n"
        "- Wood Mackenzie, BloombergNEF, S&P Global, LCG Consulting\n"
        "- ISO capacity auction results: PJM BRA, ISO-NE FCA, NYISO ICAP\n"
        "- State PUC filings for RFP results (Utah PSC, Colorado PUC, Oregon PUC, Arizona ACC)\n"
        "- Utility IRP documents (PacifiCorp IRP, APS IRP, Xcel Colorado ERP)\n"
        "- Press releases: NextEra, Invenergy, AES, EDP, Engie, Brookfield\n"
        "- News: Utility Dive, Energy Storage News, Reuters, Canary Media\n\n"
        "TARGET REGIONS (10 total):\n"
        "- PRIMARY (no LevelTen coverage, full research required):\n"
        "    ISO-NE, NYISO, WECC_DSW, WECC_RM, WECC_NW, SERC\n"
        "- SECONDARY (LevelTen-covered, provide duration breakdown only):\n"
        "    ERCOT, CAISO, PJM, MISO, SPP, AESO\n\n"
        "TARGET DURATIONS for each region: 2h, 4h, 6h\n\n"
        "Output: ALL text fields (market_note, methodology_note, caveats) MUST BE IN KOREAN.\n"
        "Use formal nominal/concise style ('~확인됨', '~추정됨', '~범위').\n"
        "For WECC_* regions, market_note MUST include PPA market commentary (utility RFP landscape, "
        "recent clearing prices, Neptune-like Utah projects context).\n"
        "Numbers stay numeric ($X/kW-mo). Region names stay English.\n\n"
        "Return ONLY this JSON structure (no markdown, no code fences):\n"
        "{\n"
        '  "research_date": "YYYY-MM-DD",\n'
        '  "iso_data": [\n'
        '    {\n'
        '      "region": "ERCOT|CAISO|PJM|MISO|SPP|AESO|ISO-NE|NYISO|WECC_DSW|WECC_RM|WECC_NW|SERC",\n'
        '      "levelten_covered": true,  // 6 LevelTen ISOs=true, 나머지 4 (ISO-NE/NYISO/WECC_*/SERC)=false\n'
        '      "durations": [\n'
        '        {"hours": 2, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"},\n'
        '        {"hours": 4, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"},\n'
        '        {"hours": 6, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"}\n'
        '      ],\n'
        '      "market_note": "(한국어) 시장 특성 1-2문장. WECC_*는 PPA 시장 commentary 포함 (주요 utility RFP, recent clearing prices, 인접 주 벤치마크)",\n'
        '      "sources": [\n'
        '        {"url": "https://...", "title": "source title", "date": "YYYY-MM", "key_data": "핵심 수치/인용 (한국어 번역 OK)"}\n'
        '      ]\n'
        '    }\n'
        '  ],\n'
        '  "methodology_note": "(한국어) 추정 방법 요약. LevelTen 공식 index와의 관계 명시: LevelTen은 6개 ISO만 커버, 본 리서치는 (1) 미커버 4개 지역(ISO-NE/NYISO/WECC_DSW/WECC_RM/WECC_NW/SERC) 보완, (2) 전체 duration별(2h/4h/6h) 세분화 목표. capacity market + merchant 수익 + utility RFP 삼각 검증",\n'
        '  "confidence_overall": "high|medium|low",\n'
        '  "caveats": "(한국어) 1-2문장. 예: 본 수치는 AI 리서치 기반 추정치. LevelTen 6개 ISO는 공식 데이터 우선. WECC sub-region 및 SERC는 공식 index 부재 — RFP/IRP 참고치"\n'
        "}\n\n"
        "Rules:\n"
        "- All prices in USD/kW-month, levelized over contract term.\n"
        "- ALWAYS include all 12 regions (10 + WECC split into 3 sub-regions):\n"
        "  ERCOT, CAISO, PJM, MISO, SPP, AESO, ISO-NE, NYISO, WECC_DSW, WECC_RM, WECC_NW, SERC.\n"
        "- Each region must have 3 durations (2h, 4h, 6h) — use benchmark if no evidence, mark confidence='low'.\n"
        "- Confidence guide: 'high' if 3+ sources corroborate; 'medium' if 1-2 sources; 'low' if benchmark/inference only.\n"
        "- For WECC_* regions, market_note MUST include PPA context (not just BESS) — target utilities and recent RFP clearing prices.\n"
        "- Dates must be 2024-2026 (recent only).\n"
        "- All text fields in Korean formal nominal style.\n"
        "- Return valid JSON only."
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-5",
                "max_tokens": 8000,
                "tools": [{
                    "type": "web_search_20250305",
                    "name": "web_search",
                    "max_uses": 10,
                }],
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=180,  # 웹서치 여러 번 → 최대 3분
        )
        if resp.status_code != 200:
            raise HTTPException(502, f"Claude API 오류: {resp.text[:400]}")

        data = resp.json()
        # content blocks 중 text 타입만 합쳐서 JSON 파싱
        text_parts = []
        for block in data.get("content", []):
            if block.get("type") == "text":
                text_parts.append(block.get("text", ""))
        full_text = "".join(text_parts).strip()

        # JSON 추출 (code fence 제거 + { } 범위)
        import re as _re
        clean = _re.sub(r"```(?:json)?\s*", "", full_text).strip().strip("`")
        start = clean.find("{")
        end = clean.rfind("}") + 1
        if start < 0 or end <= start:
            raise HTTPException(500, f"AI 응답에서 JSON을 찾을 수 없음: {full_text[:300]}")
        clean = clean[start:end]

        try:
            parsed = json.loads(clean)
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"JSON 파싱 실패: {str(e)}. 응답: {clean[:400]}")

        # 메타데이터 추가
        parsed["generated_at"] = datetime.datetime.utcnow().isoformat()[:19]
        parsed["generated_by"] = user["email"]
        parsed["source"] = "ai_research"
        # 토큰 사용량 (cost 추적용)
        usage = data.get("usage", {})
        parsed["tokens"] = {
            "input": usage.get("input_tokens", 0),
            "output": usage.get("output_tokens", 0),
        }

        # Firebase 저장 (latest로)
        fb_put("benchmark/bess_tolling/latest", parsed)
        # 히스토리도 (월별 캐시)
        month_key = datetime.date.today().strftime("%Y-%m")
        fb_put(f"benchmark/bess_tolling/history/{month_key}", parsed)

        return {"ok": True, "data": parsed}

    except HTTPException:
        raise
    except requests.Timeout:
        raise HTTPException(504, "AI 리서치 타임아웃 (3분 초과)")
    except Exception as e:
        raise HTTPException(500, f"BESS 리서치 실패: {str(e)}")


@app.get("/benchmark/bess-tolling")
def get_bess_tolling(user=Depends(get_current_user)):
    """저장된 BESS tolling 리서치 결과 조회 (latest)."""
    return fb_read("benchmark/bess_tolling/latest") or {}


@app.get("/benchmark/bess-tolling/history")
def get_bess_tolling_history(user=Depends(get_current_user)):
    """월별 히스토리 (stale 확인용)."""
    return fb_read("benchmark/bess_tolling/history") or {}


# ══════════════════════════════════════════════════
#  Valuation Calculate (Stage 1 Engine)
# ══════════════════════════════════════════════════
import numpy as np
import numpy_financial as npf

def _calc_engine(inputs: dict) -> dict:
    # ═══════════════════════════════════════════════════════
    # MODE 분기: Neptune Calibration vs 일반 Prediction
    # ═══════════════════════════════════════════════════════
    # 
    # 'prediction' (기본): 업계 표준 PF 가정
    #   - 99/5 Partnership Flip
    #   - 균등 Debt amortization
    #   - MACRS tax benefit 정상 Sponsor 귀속
    #   - CAPEX 전액 Y0 현금 지출
    #   - Debt/TE 100% Y0 drawdown
    # 
    # 'calibration' (Neptune 실측): Neptune 모델 재현 목적
    #   - Sculpted Debt (DSCR 기반 비율)
    #   - NOL 상쇄 (Y1-Y9 tax benefit 0)
    #   - Construction Cost 별도 (FMV와 다름)
    #   - 실측 draw ratio
    #   - 25.5/7 Partnership Flip
    #
    # 신규 프로젝트 분석 시 → 'prediction' 권장
    # Neptune 검증/재현 시 → 'calibration'
    mode = inputs.get('calibration_mode', 'prediction')
    is_calibration = (mode == 'calibration')

    pv_mwac   = inputs.get('pv_mwac', 199)
    pv_mwdc   = inputs.get('pv_mwdc') or pv_mwac * inputs.get('dc_ac_ratio', 1.34)
    bess_mw   = inputs.get('bess_mw', 199)
    bess_mwh  = inputs.get('bess_mwh', 796)
    life      = int(inputs.get('life', 35))

    # CAPEX 구성 ───────────────────────────────────────────────
    module_cwp   = inputs.get('module_cwp', 31.5)        # c/Wdc
    pv_bos_cwp   = inputs.get('bos_cwp', 42.88)          # c/Wdc (PV BOS+Construction)
    ess_per_kwh  = inputs.get('ess_per_kwh', 234.5)      # $/kWh (BESS Equipment)
    bess_bos_per_kwh = inputs.get('bess_bos_per_kwh', 130.0)  # $/kWh (BESS BOS — NEW)
    epc_cont_pct = inputs.get('epc_cont_pct', 8.0)       # %
    owner_pct    = inputs.get('owner_pct', 3.0)          # %
    softcost_pct = inputs.get('softcost_pct', 5.0)
    intercon_m   = inputs.get('intercon_m', 22.5)        # $M (Sub+Gentie+GSU+Trans avg Neptune)
    dev_cost_m   = inputs.get('dev_cost_m', 20.0)        # $M
    capex_etc    = inputs.get('capex_etc', 0)

    # 하드웨어/BOS 비용
    pv_module    = pv_mwdc*1000*module_cwp/100           # $K
    pv_bos       = pv_mwdc*1000*pv_bos_cwp/100           # $K
    ess_equip    = bess_mwh*ess_per_kwh                  # $K
    bess_bos     = bess_mwh*bess_bos_per_kwh             # $K (NEW)
    epc_base     = pv_module + pv_bos + ess_equip + bess_bos
    epc_total    = epc_base * (1 + epc_cont_pct/100)
    pre_capex    = (epc_total*(1+owner_pct/100+softcost_pct/100)
                    + intercon_m*1000 + dev_cost_m*1000 + capex_etc*1000)
    int_rate     = inputs.get('int_rate', 5.5) / 100
    debt_ratio   = inputs.get('debt_ratio', 47.6) / 100
    base_capex   = pre_capex * (1 + debt_ratio*int_rate*0.75 + 0.012)
    total_capex  = float(inputs['capex_total_override'])*1000 if inputs.get('capex_total_override') else base_capex

    # Dev Margin: c/Wac × (PV + BESS) MW × 10 (Neptune 표준 공식)
    dev_margin_cwac_v = inputs.get('dev_margin_cwac', 10.0)
    dev_margin   = dev_margin_cwac_v * (pv_mwac + bess_mw) * 10  # $K
    epc_margin   = epc_base * inputs.get('epc_margin_pct', 7.95)/100
    total_margin = dev_margin + epc_margin

    loan_term  = int(inputs.get('loan_term', 28 if is_calibration else 18))
    debt       = total_capex * debt_ratio
    ann_ds     = float(npf.pmt(int_rate, loan_term, -debt)) if debt > 0 else 0

    # Credit System ─────────────────────────────────────────────
    # Mode: ITC (capital credit) vs PTC (production credit)
    # ITC는 PV와 BESS를 분리해서 적용 가능 (Neptune: PV 0%, BESS 30%)
    # PTC는 PV generation에만 적용 (BESS는 ITC만 가능)
    credit_mode  = inputs.get('credit_mode', 'ITC').upper()  # 'ITC' or 'PTC'
    itc_elig     = inputs.get('itc_elig', 97) / 100

    # PV/BESS ITC 분리 (inputs 없으면 레거시 단일 credit_val 사용)
    pv_itc_rate  = inputs.get('pv_itc_rate')
    bess_itc_rate = inputs.get('bess_itc_rate')
    if pv_itc_rate is None and bess_itc_rate is None:
        # 레거시: credit_val을 전체에 적용
        legacy = inputs.get('itc_rate') or inputs.get('credit_val', 30)
        pv_itc_rate  = legacy
        bess_itc_rate = legacy
    pv_itc_rate  = (pv_itc_rate or 0) / 100
    bess_itc_rate = (bess_itc_rate or 0) / 100

    # PTC Rate ($/kWh) — production-based
    ptc_rate_per_kwh = inputs.get('ptc_rate_per_kwh') or inputs.get('credit_val')
    if credit_mode == 'PTC':
        # credit_val이 30 같이 크면 잘못 입력된 것 — 0.03으로 자동 보정
        if ptc_rate_per_kwh and ptc_rate_per_kwh > 1:
            ptc_rate_per_kwh = ptc_rate_per_kwh / 100.0
    ptc_rate_per_kwh = ptc_rate_per_kwh or 0.0

    # PV CAPEX과 BESS CAPEX 분리 (ITC basis용)
    pv_capex_share   = (pv_module + pv_bos) / epc_base if epc_base > 0 else 0.5
    bess_capex_share = (ess_equip + bess_bos) / epc_base if epc_base > 0 else 0.5

    if credit_mode == 'ITC':
        # 가중 평균 ITC rate (CAPEX 비중 기준)
        effective_itc_rate = pv_itc_rate * pv_capex_share + bess_itc_rate * bess_capex_share
    else:
        # PTC 모드: ITC 적용 안 함 (BESS도 ITC 받을 수 있지만 단순화 위해 일단 미적용)
        # BESS ITC는 PTC와 병행 가능하므로 bess_itc_rate 살아있으면 그대로
        effective_itc_rate = bess_itc_rate * bess_capex_share  # BESS만 ITC

    # TE Flip
    _fy_raw = inputs.get('flip_yield', 8.75)
    if _fy_raw > 50: _fy_raw = _fy_raw / 100
    flip_yield = _fy_raw / 100
    flip_term  = int(inputs.get('flip_term', 7))
    te_mult    = inputs.get('te_mult', 1.115)
    yield_adj  = 1 / (1 + (flip_yield - 0.0875) * 8)

    # ── TE Invest 산정 + Sponsor Equity 최소선 확보 ──────────────────
    te_theoretical = total_capex * itc_elig * effective_itc_rate * te_mult * yield_adj

    min_sponsor_eq_pct = inputs.get('min_sponsor_eq_pct', 10.0) / 100
    max_te_invest = total_capex - debt - total_capex * min_sponsor_eq_pct

    te_invest = max(0, min(te_theoretical, max_te_invest))

    # Capital Stack Override: te_ratio / sponsor_eq_ratio 명시하면 해당 비율 사용
    # (예: Neptune은 Debt 47.6%, TE 32.5%, Eq 19.8% — ITC 기반 공식으로 못 맞춤)
    te_ratio_override = inputs.get('te_ratio_override')
    if te_ratio_override is not None:
        te_invest = total_capex * te_ratio_override / 100
        te_invest = max(0, min(te_invest, max_te_invest))

    sponsor_eq = total_capex - debt - te_invest
    effective_eq = sponsor_eq * (1 - int_rate * 0.75)

    # ═══ Sponsor Y0 Cash Outflow ═══
    # 
    # Prediction mode (기본, 신규 프로젝트):
    #   Sponsor Y0 = Sponsor Equity (effective_eq) 전액 Y0 지출
    #   Debt/TE는 CAPEX 전액 Y0 drawdown
    #
    # Calibration mode (Neptune 재현):
    #   Sponsor Y0 = Construction Cost + Txn + CapInt - Debt Drawdown - TE Proceeds
    #   FMV와 구분된 실제 Y0 현금 흐름 추적
    
    # ═══ Sponsor Y0 Cash Outflow ═══
    # 
    # 모든 사업 공통 구조:
    #   CAPEX (FMV) = Construction Cost + Dev Margin + EPC Margin
    #   Debt / TE sizing은 FMV 기준 (시장 관행)
    #   Sponsor Y0 현금 = Construction + Txn + CapInt - Debt draw - TE proceeds
    #
    # Dev Margin = dev_margin_cwac(c/Wac) × (PV MWac + BESS MW) × 10
    #   Neptune 표준: 10 c/Wac
    #   모든 HWR 프로젝트에 공통 적용
    
    # Dev Margin & EPC Margin
    dev_margin_cwac = inputs.get('dev_margin_cwac', 10.0)  # c/Wac 기본 10
    total_mw_ac = pv_mwac + bess_mw
    dev_margin_k = dev_margin_cwac * total_mw_ac * 10  # $K
    
    # EPC Margin: CAPEX 대비 % (Neptune 기본 7.95%)
    epc_margin_pct_calc = inputs.get('epc_margin_pct', 7.95) / 100
    epc_margin_k = total_capex * epc_margin_pct_calc
    
    # Construction Cost = FMV CAPEX - Dev Margin - EPC Margin
    # (모든 모드 공통 로직)
    construction_cost_override = inputs.get('construction_cost_m')
    if construction_cost_override:
        construction_cost = construction_cost_override * 1000  # $M → $K (수동 override)
    else:
        construction_cost = total_capex - dev_margin_k - epc_margin_k
    
    # Transaction costs & Capitalized Interest
    # 기본값: CAPEX 대비 작은 비율 (Neptune: Txn 1.27%, CapInt 1.71%)
    txn_costs = inputs.get('txn_costs_m')
    if txn_costs is not None:
        txn_costs = txn_costs * 1000  # $M → $K
    else:
        txn_costs = total_capex * 0.0127  # Neptune 비율
    
    cap_interest = inputs.get('cap_interest_m')
    if cap_interest is not None:
        cap_interest = cap_interest * 1000
    else:
        cap_interest = total_capex * 0.0171  # Neptune 비율
    
    # Debt / TE 현금 drawdown 비율
    # Calibration (Neptune): 77.5% / 93.5%
    # Prediction (표준): 100% / 100% (Y0 전액 drawdown)
    if is_calibration:
        debt_drawdown_ratio = inputs.get('debt_drawdown_ratio', 0.775)
        te_proceeds_ratio = inputs.get('te_proceeds_ratio', 0.935)
    else:
        debt_drawdown_ratio = inputs.get('debt_drawdown_ratio', 1.0)
        te_proceeds_ratio = inputs.get('te_proceeds_ratio', 1.0)
    
    debt_drawdown = debt * debt_drawdown_ratio
    te_proceeds = te_invest * te_proceeds_ratio
    
    # Sponsor Y0 실제 현금 outflow
    sponsor_y0_cash = construction_cost + txn_costs + cap_interest - debt_drawdown - te_proceeds
    
    # Flip year event cash (Neptune Y10 pattern, 일반 프로젝트는 0)
    flip_event_cf = inputs.get('flip_event_cf', 0)

    # 레거시 호환
    itc_rate = effective_itc_rate
    net_sponsor_y0 = sponsor_y0_cash
    
    # Dev Margin 최종 값 저장 (output에 노출)
    computed_dev_margin_k = dev_margin_k

    # MACRS depreciation
    tax_rate    = inputs.get('tax_rate', 21) / 100
    macrs_basis = total_capex * itc_elig * (1 - itc_rate/2)
    depr_sched  = {i+1: macrs_basis*r for i,r in enumerate(MACRS_5YR)}
    
    # Depreciation share — Partnership Flip 구조에 따라 다름
    # Prediction mode 기본값:
    #   Pre-flip: TE가 depr 99% → Sponsor depr_share = 1%
    #   Post-flip: Sponsor 95% → Sponsor depr_share = 95%
    # Calibration mode (Neptune): 0.7721 (실측)
    # 사용자 override 가능
    if is_calibration:
        depr_share_pre = inputs.get('depr_share', 0.7721)
        depr_share_post = inputs.get('depr_share_post', depr_share_pre)
    else:
        depr_share_pre = inputs.get('depr_share_pre', 0.01)  # TE 99%
        depr_share_post = inputs.get('depr_share_post', 0.95)  # Sponsor 95%
    # 레거시 호환: depr_share만 override 주면 pre/post 모두 동일
    if 'depr_share' in inputs and not is_calibration:
        depr_share_pre = depr_share_post = inputs.get('depr_share')
    depr_share = depr_share_pre  # backward compat

    # Cash allocation
    # Prediction mode (기본): 표준 Partnership Flip 99%/5%
    # Calibration mode (Neptune): 실측 25.5%/7%
    if is_calibration:
        default_pre_flip = 25.5
        default_post_flip = 7
    else:
        default_pre_flip = 99
        default_post_flip = 5
    pre_flip_cash_te  = inputs.get('pre_flip_cash_te', default_pre_flip) / 100
    post_flip_cash_te = inputs.get('post_flip_cash_te', default_post_flip) / 100

    # Revenue
    cf_pct       = inputs.get('cf_pct', 21.24)
    net_prod_yr1 = inputs.get('net_prod_yr1', None)
    ann_prod_yr1 = float(net_prod_yr1) if net_prod_yr1 else pv_mwac*cf_pct/100*8760
    ppa_price   = inputs.get('ppa_price', 68.82)
    ppa_term    = int(inputs.get('ppa_term', 25))
    ppa_esc     = inputs.get('ppa_esc', 0) / 100
    # bess_toll: CF_Annual Y1 실제값 우선, 없으면 Summary 파싱값
    bess_toll   = inputs.get('bess_toll_y1_effective') or inputs.get('bess_toll', 14.50)
    bess_toll_t = int(inputs.get('bess_toll_term', 20))
    bess_toll_esc = inputs.get('bess_toll_esc', 0) / 100  # Toll escalation (%)
    # BESS Toll 월 적용: 표준 12개월, Neptune calibration은 12.72 (pro-rated)
    bess_months_per_yr = inputs.get('bess_months_per_yr', 12.72 if is_calibration else 12.0)
    merch_ppa   = inputs.get('merchant_ppa', 45.0)
    merch_esc   = inputs.get('merchant_esc', 3.0) / 100
    degradation = inputs.get('degradation', 0.0064)
    # Neptune의 CF%는 availability 내재 → Calibration 모드에서 1.0
    # 일반 프로젝트는 별도 availability factor 적용
    avail_1     = inputs.get('availability_yr1', 1.0 if is_calibration else 0.977)
    avail_2     = inputs.get('availability_yr2', 1.0 if is_calibration else 0.982)

    # OPEX
    pv_om=inputs.get('pv_om',4.5); pv_om_nc=inputs.get('pv_om_nc',1.0)
    pv_aux=inputs.get('pv_aux',1.56); bess_om=inputs.get('bess_om',8.64)
    bess_om_nc=inputs.get('bess_om_nc',1.0); bess_aux=inputs.get('bess_aux',3.84)
    ins_pv=inputs.get('insurance_pv',10.57); ins_bess=inputs.get('insurance_bess',5.05)
    asset_mgmt=inputs.get('asset_mgmt',210); prop_tax=inputs.get('prop_tax_yr1',3162)
    land_rent=inputs.get('land_rent_yr1',437); opex_etc=inputs.get('opex_etc', 0.56 if is_calibration else 0)
    opex_esc=inputs.get('opex_esc',2.0)/100

    # Augmentation (Neptune: Y4, Y8, Y14 × $22.5M / 표준: Y4, Y8, Y12)
    default_aug_y3 = 14 if is_calibration else 12
    aug_price=inputs.get('aug_price',150); aug_mwh_pct=inputs.get('aug_mwh_pct',18.8)
    aug_mwh_ea=bess_mwh*aug_mwh_pct/100
    aug_years=[int(y) for y in [inputs.get('aug_y1',4),inputs.get('aug_y2',8),inputs.get('aug_y3',default_aug_y3)] if y and int(y)>0]
    aug_cost_ea=aug_mwh_ea*aug_price

    # Full 35-year CF schedule
    # Sponsor Y0 = 실제 현금 outflow (Neptune R25 방식)
    # Unlev Y0 = -CAPEX + ITC tax credit Y0 benefit
    #   Neptune R51 Y0 = -385,453 → CAPEX -639,855에서 +254,402 TE proceeds 반영 + 추가 조정
    #   단순화: Unlev는 project 전체 관점이므로 full CAPEX
    effective_itc_value = total_capex * itc_elig * effective_itc_rate
    # Unlev Y0 (Neptune Row 26 방식): -Construction + TE proceeds
    # Neptune Row 26 Y0 = -385,453 ≈ -639,855 + 254,405 (Construction - TE proceeds)
    # txn + cap_interest는 Partnership 관점에서 Y0 cash flow 이전의 financing 비용이라
    # Unlevered IRR 계산에는 포함 안 함 (엑셀 실측 일치)
    unlev_y0 = -construction_cost + te_proceeds

    cashflows=[-effective_eq]; unlev_cfs=[unlev_y0]
    sponsor_cfs=[-sponsor_y0_cash]; pretax_cfs=[-sponsor_y0_cash]
    debt_bal=debt; detail=[]; ebitda_yr1=None

    for yr in range(1, life+1):
        avail = avail_1 if yr==1 else avail_2
        prod  = ann_prod_yr1 * avail * ((1-degradation)**(yr-1))

        # CF_Annual parsed schedule 우선 사용 (실제 Neptune 모델값)
        pv_sched   = inputs.get('pv_rev_schedule', [])
        bess_sched = inputs.get('bess_rev_schedule', [])
        merch_sched= inputs.get('merch_rev_schedule', [])

        if pv_sched and yr-1 < len(pv_sched):
            pv_rev = pv_sched[yr-1]
        elif yr <= ppa_term:
            pv_rev = prod*ppa_price*((1+ppa_esc)**(yr-1))/1000
        else:
            # Merchant 기간: escalation 기산점은 merchant 시작 연도
            # Neptune: Y26 $61/MWh, Y35 $73/MWh → merchant_esc가 Y26부터 적용됨
            merch_yr = yr - ppa_term  # merchant 경과년수 (Y26 → 1)
            pv_rev = prod*merch_ppa*((1+merch_esc)**(merch_yr-1))/1000

        if bess_sched and yr-1 < len(bess_sched):
            bess_rev = bess_sched[yr-1]
        else:
            bess_rev = bess_mw*1000*bess_toll*((1+bess_toll_esc)**(yr-1))*bess_months_per_yr/1000 if yr<=bess_toll_t else 0

        if merch_sched and yr-1 < len(merch_sched) and merch_sched[yr-1] > 0:
            pv_rev = merch_sched[yr-1]  # merchant 기간은 merch_sched 우선

        total_rev = pv_rev + bess_rev

        esc=(1+opex_esc)**(yr-1); prop_esc=max(0.35,1-0.025*(yr-1))
        opex=(pv_mwdc*1000*pv_om/1000*esc + pv_mwac*1000*pv_om_nc/1000*esc +
              pv_mwac*1000*pv_aux/1000*esc + bess_mw*1000*bess_om/1000*esc +
              bess_mw*1000*bess_om_nc/1000*esc + bess_mw*1000*bess_aux/1000*esc +
              pv_mwac*1000*ins_pv/1000*esc + bess_mw*1000*ins_bess/1000*esc +
              asset_mgmt*esc + prop_tax*prop_esc + land_rent*esc + opex_etc*1000*esc)

        ebitda = total_rev - opex
        if yr==1: ebitda_yr1=ebitda
        aug_c = aug_cost_ea if yr in aug_years else 0

        # Partnership CF = EBITDA - Aug (Neptune R19 방식)
        partnership_cf = ebitda - aug_c

        # Debt Service: Neptune sculpted schedule (DSCR 기반)
        # Calibration mode 기본 True, Prediction mode 기본 False (균등 amortization)
        use_sculpted = inputs.get('use_sculpted_debt', is_calibration)
        if use_sculpted and yr <= 28:
            # Neptune Case 2 Debt/Partnership 비율 (실측)
            neptune_debt_ratios = {
                1: 0.602, 2: 0.604, 3: 0.606, 4: 0.570, 5: 0.800,
                6: 0.663, 7: 0.663, 8: 0.677, 9: 0.663, 10: 0.263,
                11: 0.607, 12: 0.700, 13: 0.701, 14: 0.711, 15: 0.701,
                16: 0.701, 17: 0.701, 18: 0.701, 19: 0.702, 20: 0.700,
                21: 0.905, 22: 0.914, 23: 0.923, 24: 0.933, 25: 0.938,
                26: 0.504, 27: 0.504, 28: 0.503,
            }
            ratio = neptune_debt_ratios.get(yr, 0)
            ds = partnership_cf * ratio
            debt_bal = max(0, debt_bal - max(0, ds - debt_bal * int_rate))
        elif yr <= loan_term and debt_bal > 0:
            # 기존 amortization (non-Neptune)
            int_p = debt_bal * int_rate
            prin = ann_ds - int_p
            ds = ann_ds
            debt_bal = max(0, debt_bal - prin)
        else:
            ds = 0

        # MACRS depreciation tax benefit
        # Neptune NOL 이월 효과로 Y1-Y9 tax benefit이 Partnership tax와 상쇄 (R42 + R33 ≈ 0)
        # 따라서 Sponsor aftertax CF ≈ Sponsor pretax CF (MACRS 따로 더하지 않음)
        # ITC는 이미 Y0 Sponsor cash outflow에 반영됨 (Construction Cost 기반)
        depr = depr_sched.get(yr, 0)
        # NOL 상쇄 로직
        # Calibration mode 기본 True (Neptune처럼 Y1-Y9 tax benefit 상쇄)
        # Prediction mode 기본 False (MACRS tax benefit 정상 반영)
        use_nol_offset = inputs.get('use_nol_offset', is_calibration)
        # flip year 기반 Sponsor depr share (pre/post flip)
        current_depr_share = depr_share_pre if yr <= flip_term else depr_share_post
        if use_nol_offset:
            s_tax = 0  # NOL로 상쇄
        else:
            s_tax = depr * tax_rate * current_depr_share

        # PTC (Production Tax Credit) — PV만, COD 후 10년간
        ptc_benefit = 0
        if credit_mode == 'PTC' and yr <= 10:
            ptc_benefit = prod * ptc_rate_per_kwh  # MWh × $/kWh → $K
            # PTC도 flip 기반 배분
            s_tax += ptc_benefit * current_depr_share

        op_cf = ebitda - ds - aug_c
        # ── Sponsor CF: Neptune Returns Row 25 방식 ──
        # Row 25 = Row 19 (Partnership) - Row 21 (Debt net) - Row 22 (TE dist) + Row 23 (Pay-Go)
        # TE distribution은 Partnership CF에 비례 (Debt와 독립)
        # 이전 엔진은 op_cf × (1-TE%) 로 계산했으나, 이는 Debt 영향을 TE 배분에 섞음 → 교정
        te_cash_pct = pre_flip_cash_te if yr <= flip_term else post_flip_cash_te
        te_dist_cash = partnership_cf * te_cash_pct
        s_cf = partnership_cf - ds - te_dist_cash + s_tax

        # Flip Year event: TE buyout 직후 Sponsor 일시 대금 수령
        if yr == flip_term + 1 and flip_event_cf > 0:
            s_cf += flip_event_cf

        s_cf_pretax = partnership_cf - ds - te_dist_cash

        # Unlevered aftertax CF (Neptune R51 구조):
        # = Partnership CF - TE distribution (Debt 제외한 Sponsor+TE 관점)
        # TE dist 비율은 위 Sponsor CF와 동일한 te_cash_pct 사용 (일관성)
        te_dist_unlev = partnership_cf * te_cash_pct
        unlev_aftertax_cf = partnership_cf - te_dist_unlev + (ptc_benefit if credit_mode == 'PTC' and yr <= 10 else 0)

        cashflows.append(op_cf); unlev_cfs.append(unlev_aftertax_cf); sponsor_cfs.append(s_cf); pretax_cfs.append(s_cf_pretax)
        if yr<=10:
            detail.append({'yr':yr,'rev':round(total_rev,0),'opex':round(opex,0),
                'ebitda':round(ebitda,0),'ds':round(ds,0),'aug':round(aug_c,0),
                'depr':round(depr,0),'s_cf':round(s_cf,0),'ptc':round(ptc_benefit,0)})

    lirr = _irr_robust(pretax_cfs, guess=0.10)   # Sponsor pretax levered (Neptune Row 26 ~10%)
    uirr = _irr_robust(unlev_cfs, guess=0.05)    # Asset-level unlevered (Neptune Row 27 ~8%)
    sirr = _irr_robust(sponsor_cfs, guess=0.10)  # Sponsor after-tax w/ MACRS (Full Life)
    try:
        sirr_c = float(npf.irr(sponsor_cfs[:ppa_term+1]))
        if np.isnan(sirr_c): sirr_c = None
    except Exception:
        sirr_c = None
    ebitda_yield = ebitda_yr1/total_capex*100 if total_capex else 0

    # ── NPV 계산 (Hurdle 기준 할인) ────────────────────────────────
    # Sponsor NPV: Hurdle IRR(예: 10%)로 할인 — 매수자 관점 가치
    # Project NPV: WACC로 할인 — 프로젝트 자체 가치
    hurdle_sponsor = inputs.get('hurdle_sponsor_irr', 9.0) / 100  # Default 9%
    # WACC 계산 (approximation): tax-adjusted weighted cost
    wacc_debt_cost = int_rate * (1 - tax_rate)  # after-tax
    wacc_te_cost = 0.07   # TE 조달 비용 (typical)
    wacc_eq_cost = 0.11   # Sponsor eq 비용 (typical)
    debt_w = debt / total_capex if total_capex else 0
    te_w = te_invest / total_capex if total_capex else 0
    eq_w = sponsor_eq / total_capex if total_capex else 0
    wacc = (debt_w * wacc_debt_cost) + (te_w * wacc_te_cost) + (eq_w * wacc_eq_cost)
    if wacc <= 0 or wacc > 0.5: wacc = 0.072  # fallback

    def _npv(cfs, rate):
        try:
            return float(npf.npv(rate, cfs))
        except Exception:
            return None

    sponsor_npv = _npv(sponsor_cfs, hurdle_sponsor)
    project_npv = _npv(unlev_cfs, wacc)
    # ───────────────────────────────────────────────────────────────

    return {
        'capex_total':   round(total_capex,0),
        'epc_base':      round(epc_base,0),
        'pv_module':     round(pv_module,0),
        'pv_bos':        round(pv_bos,0),
        'bess_equip':    round(ess_equip,0),
        'bess_bos':      round(bess_bos,0),
        'debt':          round(debt,0),
        'equity':        round(sponsor_eq+te_invest,0),
        'te_invest':     round(te_invest,0),
        'sponsor_equity':round(sponsor_eq,0),
        'dev_margin':    round(dev_margin,0),
        'epc_margin':    round(epc_margin,0),
        'total_margin':  round(total_margin,0),
        'levered_irr':   round(lirr,6) if (lirr is not None and not np.isnan(lirr)) else None,
        'unlevered_irr': round(uirr,6) if (uirr is not None and not np.isnan(uirr)) else None,
        'sponsor_irr':   round(sirr,6) if (sirr is not None and not np.isnan(sirr)) else None,
        'sponsor_irr_contract': round(sirr_c,6) if (sirr_c is not None and not np.isnan(sirr_c)) else None,
        'sponsor_npv':   round(sponsor_npv,0) if sponsor_npv is not None else None,
        'project_npv':   round(project_npv,0) if project_npv is not None else None,
        'wacc':          round(wacc,6),
        'hurdle_sponsor_irr_used': round(hurdle_sponsor,6),
        'ebitda_yield':  round(ebitda_yield,2),
        'aug_cost_ea':   round(aug_cost_ea,0),
        'life_yrs':      life,
        'credit_mode':   credit_mode,
        'pv_itc_rate':   round(pv_itc_rate*100, 2),
        'bess_itc_rate': round(bess_itc_rate*100, 2),
        'ptc_rate':      round(ptc_rate_per_kwh, 4),
        'annual_detail': detail,
        'cashflows':     [round(x,0) for x in cashflows[:36]],
    }


# ══════════════════════════════════════════════════════════════
# Calibration Auto-Merge
# ──────────────────────────────────────────────────────────────
# 프런트 사이드바는 Neptune 구조적 파라미터를 전송하지 않으므로,
# calibration_mode='calibration' 일 때 백엔드가 자동 주입한다.
# (값은 get_calc_defaults endpoint와 동일 — single source로 가려면 추후 리팩터)
# ══════════════════════════════════════════════════════════════

# 구조적 Neptune 파라미터 — calibration 모드에서 항상 이 값 사용
# (사이드바의 prediction default를 덮어쓴다; 예: loan_term 18 → 28)
_CALIB_STRUCTURAL = {
    'loan_term': 28,
    'aug_y3': 14,
    'bess_months_per_yr': 12.72,
    'opex_etc': 0.56,
    'construction_cost_m': 639.855,
    'txn_costs_m': 10.6,
    'cap_interest_m': 14.3,
    'debt_drawdown_ratio': 0.775,
    'te_proceeds_ratio': 0.935,
    # Neptune Returns 시트 Row 22 실측: TE dist ≈ 9.2% of Partnership CF (Y1-9)
    # Y10에서 5%로 내려감 (Flip effective Y10) → flip_term = 9
    'pre_flip_cash_te': 9.2,
    'post_flip_cash_te': 5.0,
    'flip_term': 9,
    'depr_share': 0.7721,
    'use_nol_offset': True,
    'use_sculpted_debt': True,
    'flip_event_cf': 0,
}
# 사이드바에서 오기도 하는 파라미터 — 없을 때만 Neptune 값으로 채움
_CALIB_FILL_IF_MISSING = {
    'availability_yr1': 1.0,
    'availability_yr2': 1.0,
    'capex_total_override': 836.7,
    'te_ratio_override': 32.52,
    'flip_yield': 8.75,
}

def _apply_calibration_defaults(inputs: dict) -> dict:
    """calibration_mode='calibration'일 때 Neptune 구조적 파라미터 자동 주입.
    
    주의: _calc_engine에 직접 넣지 않는다 — _decompose_irr_difference가
    step별로 실험적 param 변경 (예: pre_flip_cash_te 25.5→99) 할 때
    auto-merge가 덮어쓰면 decompose가 깨진다.
    따라서 endpoint 레벨에서만 호출한다.
    """
    if inputs.get('calibration_mode') != 'calibration':
        return inputs
    merged = dict(inputs)
    # 구조적 파라미터는 Neptune 값 강제
    for k, v in _CALIB_STRUCTURAL.items():
        merged[k] = v
    # 나머지는 fill-if-missing
    for k, v in _CALIB_FILL_IF_MISSING.items():
        if k not in merged or merged[k] is None:
            merged[k] = v
    return merged


@app.post("/valuation/calculate")
def calculate_valuation(req: ValuationCalcRequest, user=Depends(get_current_user)):
    """Run PF calculation engine with given inputs"""
    try:
        inputs = _apply_calibration_defaults(dict(req.inputs))
        result = _calc_engine(inputs)
        return {"ok": True, "project_id": req.project_id, "result": result}
    except Exception as e:
        raise HTTPException(500, f"Calculation error: {str(e)}")


# ══════════════════════════════════════════════════════════════
# IRR 차이 분해 (Calibration vs Prediction)
# ══════════════════════════════════════════════════════════════
def _decompose_irr_difference(inputs_base: dict) -> dict:
    """Calibration → Prediction 전환 시 각 요인이 IRR에 미치는 기여도 분해
    
    방법: 순차적 ON/OFF
      1) Full Calibration IRR (starting point)
      2) 각 Neptune-specific 파라미터를 하나씩 '해제' (Prediction 값으로)
      3) 각 단계의 IRR 변화량 = 해당 요인의 기여도
      4) 최종 = Prediction IRR
    
    4개 주요 요인:
      - NOL 상쇄 (use_nol_offset)
      - Sculpted Debt (use_sculpted_debt)
      - Partnership Flip 25.5/7 vs 99/5
      - Y0 현금 구조 (Construction < FMV)
    
    Returns:
        {
            'calib_irr': 11.14,
            'predict_irr': 6.17,
            'total_delta': 4.97,
            'factors': [
                {'name': 'NOL 상쇄', 'delta_pp': 3.2, 'from': '...', 'to': '...'},
                ...
            ]
        }
    """
    def _get_irr(inp):
        """Full Life Sponsor IRR (%)"""
        try:
            r = _calc_engine(inp)
            v = r.get('sponsor_irr')
            return (v * 100) if v is not None else 0.0
        except:
            return 0.0
    
    # Start: full calibration
    base = dict(inputs_base)
    base['calibration_mode'] = 'calibration'
    calib_irr = _get_irr(base)
    
    # End point: full prediction
    predict = dict(inputs_base)
    predict['calibration_mode'] = 'prediction'
    # prediction에서는 calibration 전용 파라미터 제거
    for k in ['construction_cost_m', 'txn_costs_m', 'cap_interest_m',
              'debt_drawdown_ratio', 'te_proceeds_ratio',
              'pre_flip_cash_te', 'post_flip_cash_te',
              'depr_share', 'use_nol_offset', 'use_sculpted_debt']:
        predict.pop(k, None)
    predict_irr = _get_irr(predict)
    
    total_delta = predict_irr - calib_irr  # 보통 음수 (Prediction이 낮음)
    
    factors = []
    current = dict(base)  # Calibration 상태에서 시작
    current_irr = calib_irr
    
    # 순서 중요: 영향 큰 구조적 요인부터 해제 (현실적 기여도 계산)
    # 1. Partnership Flip (가장 큰 구조 차이)
    # 2. Y0 현금 구조
    # 3. Sculpted Debt
    # 4. NOL 상쇄 (마지막, 세금 효과)
    
    # ─── Factor 1: Partnership Flip 25.5/7 → 99/5 ───
    step1 = dict(current)
    step1['pre_flip_cash_te'] = 99
    step1['post_flip_cash_te'] = 5
    step1['depr_share_pre'] = 0.01
    step1['depr_share_post'] = 0.95
    step1.pop('depr_share', None)
    irr1 = _get_irr(step1)
    delta1 = irr1 - current_irr
    factors.append({
        'name_ko': 'Partnership Flip 구조 (25/7 → 99/5)',
        'name_en': 'Partnership Flip Structure (25/7 → 99/5)',
        'delta_pp': round(delta1, 2),
        'from_calib': '25.5/7 (Neptune Pay-Go 추정)',
        'to_predict': '99/5 (표준 Yield-Based Flip)',
        'explain_ko': 'Neptune은 pre-flip cash를 TE 25.5% / Sponsor 74.5%로 배분 (Pay-Go 또는 hybrid 구조 추정). Prediction은 표준 99/5 flip으로 pre-flip Sponsor 현금이 1%로 줄어듦.',
        'explain_en': 'Neptune allocates pre-flip cash as TE 25.5% / Sponsor 74.5% (likely Pay-Go or hybrid). Prediction uses standard 99/5 flip, reducing pre-flip Sponsor cash to just 1%.',
        'excel_hint_ko': 'Excel Partnership Flip 탭에서 pre-flip cash split 확인. 표준 99/1 아니면 Pay-Go 구조인지 또는 별도 hybrid 로직인지 문서화 필요.',
        'excel_hint_en': 'Check Partnership Flip tab for pre-flip cash split. If not standard 99/1, document whether Pay-Go or hybrid.',
    })
    current = step1
    current_irr = irr1
    
    # ─── Factor 2: Y0 현금 구조 (Construction ≠ FMV) ───
    step2 = dict(current)
    step2['calibration_mode'] = 'prediction'
    for k in ['construction_cost_m', 'txn_costs_m', 'cap_interest_m',
              'debt_drawdown_ratio', 'te_proceeds_ratio']:
        step2.pop(k, None)
    # 단, NOL과 Debt는 유지 (뒤에서 순차 해제)
    step2['use_nol_offset'] = current.get('use_nol_offset', True)
    step2['use_sculpted_debt'] = current.get('use_sculpted_debt', True)
    irr2 = _get_irr(step2)
    delta2 = irr2 - current_irr
    factors.append({
        'name_ko': 'Y0 현금 구조 (Construction ≠ FMV)',
        'name_en': 'Y0 Cash Structure (Construction ≠ FMV)',
        'delta_pp': round(delta2, 2),
        'from_calib': 'Construction + Txn + CapInt - Debt draw - TE proceeds',
        'to_predict': 'Sponsor Equity 전액 Y0 지출',
        'explain_ko': 'Neptune은 Y0에 Construction Cost $640M (FMV $837M 아님) + Txn Cost $10.6M + Cap Interest $14.3M 지출, Debt 77.5% / TE 93.5%만 drawdown. 나머지는 후속 기간에 drawdown. Prediction은 전액 Y0.',
        'explain_en': 'Neptune Y0 uses Construction Cost $640M (not FMV $837M) + Txn $10.6M + Cap Interest $14.3M, with Debt 77.5% / TE 93.5% drawn. Rest drawn later. Prediction uses full drawdown at Y0.',
        'excel_hint_ko': 'Excel Sources & Uses 탭에서 Y0 Debt/TE drawdown 비율 확인. 전체 Debt 대비 construction 기간 drawdown 비율이 77.5%인지 검증.',
        'excel_hint_en': 'Check Sources & Uses tab for Y0 Debt/TE drawdown ratios. Verify if construction-period drawdown is 77.5% of total Debt.',
    })
    current = step2
    current_irr = irr2
    
    # ─── Factor 3: Sculpted Debt 해제 ───
    step3 = dict(current)
    step3['use_sculpted_debt'] = False
    irr3 = _get_irr(step3)
    delta3 = irr3 - current_irr
    factors.append({
        'name_ko': 'Sculpted Debt (DSCR 기반)',
        'name_en': 'Sculpted Debt (DSCR-based)',
        'delta_pp': round(delta3, 2),
        'from_calib': 'DSCR 1.30 맞춤형 상환',
        'to_predict': '균등 amortization',
        'explain_ko': 'Neptune은 각 연도 DSCR 1.30 맞추기 위해 상환 금액을 동적으로 조정 (Sculpted). Prediction은 균등 상환으로 초기 DSCR이 낮고 후반 높음.',
        'explain_en': 'Neptune dynamically adjusts principal to maintain DSCR 1.30 (Sculpted). Prediction uses level amortization—lower DSCR upfront, higher later.',
        'excel_hint_ko': 'Excel Debt 탭 상환 스케줄이 DSCR 기반 sculpted인지 확인. R161~R180 부근의 IF(DSCR...) 수식.',
        'excel_hint_en': 'Verify Debt tab amortization schedule is DSCR-based sculpted. Check IF(DSCR...) formulas near R161-R180.',
    })
    current = step3
    current_irr = irr3
    
    # ─── Factor 4: NOL 상쇄 해제 ───
    step4 = dict(current)
    step4['use_nol_offset'] = False
    irr4 = _get_irr(step4)
    delta4 = irr4 - current_irr
    factors.append({
        'name_ko': 'NOL 상쇄 (Y1~Y9 Tax 상쇄)',
        'name_en': 'NOL Offset (Y1~Y9 Tax offset)',
        'delta_pp': round(delta4, 2),
        'from_calib': 'Y1-Y9 Sponsor tax = $0',
        'to_predict': 'Sponsor tax = MACRS × share',
        'explain_ko': 'Neptune은 NOL 이월로 Y1~Y9 Partnership tax를 상쇄. Prediction은 MACRS tax benefit이 정상적으로 Sponsor에게 귀속.',
        'explain_en': 'Neptune offsets Y1~Y9 Partnership tax via NOL carryforward. Prediction allocates MACRS tax benefit normally to Sponsor.',
        'excel_hint_ko': 'Excel에서 NOL carryforward 로직이 MACRS benefit을 과도하게 소진하는지 확인. IRS 80% 규칙 적용 여부.',
        'excel_hint_en': 'Verify NOL carryforward in Excel is not over-consuming MACRS benefit. Check IRS 80% limitation.',
    })
    
    return {
        'calib_irr': round(calib_irr, 2),
        'predict_irr': round(predict_irr, 2),
        'total_delta': round(total_delta, 2),
        'factors': factors,
        'note_ko': '각 요인은 Calibration 상태에서 순차적으로 해제한 기여도. 순서에 따라 값이 조금씩 달라질 수 있음 (상호작용 효과).',
        'note_en': 'Each factor is measured by sequentially disabling from Calibration state. Values may vary slightly by order (interaction effects).',
    }


class DecomposeIRRRequest(BaseModel):
    project_id: str = ""
    inputs: dict

@app.post("/valuation/decompose-irr")
def decompose_irr(req: DecomposeIRRRequest, user=Depends(get_current_user)):
    """Calibration vs Prediction IRR 차이를 4개 요인별로 분해"""
    try:
        result = _decompose_irr_difference(req.inputs)
        return {"ok": True, "project_id": req.project_id, "result": result}
    except Exception as e:
        raise HTTPException(500, f"Decomposition error: {str(e)}")


class ExplainDiffRequest(BaseModel):
    project_id: str = ""
    decomposition: dict  # _decompose_irr_difference 결과
    lang: str = 'ko'

@app.post("/valuation/explain-diff")
def explain_diff(req: ExplainDiffRequest, user=Depends(get_current_user)):
    """Claude API로 IRR 차이에 대한 자연어 해설 + 엑셀 수정 제안"""
    try:
        import anthropic
        client = anthropic.Anthropic()
        
        d = req.decomposition
        lang = req.lang
        
        # 프롬프트 구성 (결정적 숫자를 context로 제공 → 환각 방지)
        if lang == 'en':
            system_prompt = """You are a PF (Project Finance) Solar+BESS expert. 
Based on the decomposition data provided, explain why Calibration IRR differs from Prediction IRR 
and suggest specific Excel modifications. Be concise, practical, and cite actual numbers. 
Max 5 paragraphs. Do not invent data not in context."""
            user_prompt = f"""
Calibration IRR: {d['calib_irr']}%
Prediction IRR: {d['predict_irr']}%
Total Difference: {d['total_delta']}%p

Factor Breakdown:
"""
            for f in d['factors']:
                user_prompt += f"\n• {f['name_en']}: {f['delta_pp']:+.2f}%p\n  From (Calib): {f['from_calib']}\n  To (Predict): {f['to_predict']}\n  Excel hint: {f['excel_hint_en']}\n"
            user_prompt += "\nExplain the key drivers of the difference and what the Excel modeler should verify/modify in their spreadsheet. Focus on actionable Excel-level advice."
        else:
            system_prompt = """당신은 PF (Project Finance) Solar+BESS 전문가입니다.
제공된 분해 데이터를 바탕으로 Calibration IRR과 Prediction IRR 차이의 원인을 설명하고, 
구체적인 엑셀 수정 제안을 해주세요. 간결하고 실용적으로, 실제 수치를 인용하세요.
최대 5문단. context에 없는 데이터를 지어내지 마세요."""
            user_prompt = f"""
Calibration IRR: {d['calib_irr']}%
Prediction IRR: {d['predict_irr']}%
총 차이: {d['total_delta']:+.2f}%p

요인별 분해:
"""
            for f in d['factors']:
                user_prompt += f"\n• {f['name_ko']}: {f['delta_pp']:+.2f}%p\n  Calib → : {f['from_calib']}\n  Predict → : {f['to_predict']}\n  Excel 힌트: {f['excel_hint_ko']}\n"
            user_prompt += "\n이 차이의 핵심 원인을 설명하고, 엑셀 모델러가 스프레드시트에서 확인/수정해야 할 사항을 알려주세요. 실행 가능한 엑셀 레벨 조언에 집중."
        
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=1500,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}]
        )
        
        explanation = response.content[0].text if response.content else ""
        return {"ok": True, "explanation": explanation, "lang": lang}
    except Exception as e:
        raise HTTPException(500, f"Explain error: {str(e)}")


# ── Break-Even Analysis (Newton-Raphson) ─────────
class BreakEvenRequest(BaseModel):
    project_id: str = ""
    inputs: dict
    target_irr_pct: float  # e.g., 11.0 for 11%
    target_var: str = "ppa_price"  # 현재는 PPA만 지원 (확장 가능)

@app.post("/valuation/breakeven")
def break_even(req: BreakEvenRequest, user=Depends(get_current_user)):
    """
    Newton-Raphson 기반 정확한 PPA 역산.
    Phase 1: PPA ±25% 11점 민감도 스캔
    Phase 2: Newton-Raphson (tolerance 0.01% IRR)
    """
    try:
        base_inputs = dict(req.inputs)
        base_ppa = float(base_inputs.get("ppa_price", 68.82))
        target_irr = req.target_irr_pct / 100.0  # 0.11
        tol = 0.0001  # 0.01% IRR tolerance
        h = 0.50  # finite difference step $/MWh
        max_iter = 10

        def calc_irr(ppa_val):
            """Calc engine 호출 → Sponsor IRR (Full Life 우선) 반환.
            엔진 발산 시 None 반환 (caller가 처리)."""
            inp = dict(base_inputs)
            inp["ppa_price"] = ppa_val
            try:
                res = _calc_engine(inp)
            except Exception:
                return None
            # sponsor_irr 우선, 없으면 contract, 둘 다 None이면 None
            s_irr = res.get("sponsor_irr")
            if s_irr is None:
                s_irr = res.get("sponsor_irr_contract")
            return s_irr  # None or float

        # ── Phase 1: ±25% 민감도 스캔 ──
        pcts = [-25, -20, -15, -10, -5, 0, 5, 10, 15, 20, 25]
        sensitivity = []
        for pct in pcts:
            ppa_p = base_ppa * (1 + pct / 100.0)
            irr_p = calc_irr(ppa_p)
            sensitivity.append({
                "pct": pct,
                "ppa": round(ppa_p, 2),
                "irr_pct": round(irr_p * 100, 4) if irr_p is not None else None
            })

        # ── Phase 2: Newton-Raphson ──
        iterations = []
        ppa = base_ppa  # 초기값
        status = "not_started"
        solution = None

        # Sensitivity에서 유효한 값만 추출
        valid_sens = [s for s in sensitivity if s["irr_pct"] is not None]
        if not valid_sens:
            # 엔진이 모든 PPA에서 발산 — break-even 불가
            status = "engine_diverged"
            solution = {
                "ppa": base_ppa,
                "irr_pct": 0,
                "error_pct": 0,
                "iterations": 0,
                "converged": False,
            }
            base_res = _calc_engine(base_inputs)
            return {
                "ok": True,
                "base_ppa": base_ppa,
                "target_irr_pct": req.target_irr_pct,
                "sensitivity": sensitivity,
                "iterations": iterations,
                "solution": solution,
                "status": status,
                "tolerance_pct": tol * 100,
                "dev_margin_k": base_res.get("dev_margin", 0),
            }

        min_irr = min(s["irr_pct"] for s in valid_sens)
        max_irr = max(s["irr_pct"] for s in valid_sens)
        target_irr_pct = target_irr * 100

        # Target이 sensitivity 범위 안에 있는지 먼저 확인
        # 범위 안에 있으면 linear interpolation으로 초기값 설정 (Newton-Raphson에 좋은 시작점)
        if min_irr <= target_irr_pct <= max_irr:
            # 두 인접 sens point 사이에서 linear interp
            sorted_sens = sorted(valid_sens, key=lambda s: s["ppa"])
            for i in range(len(sorted_sens) - 1):
                lo = sorted_sens[i]
                hi = sorted_sens[i+1]
                if (lo["irr_pct"] <= target_irr_pct <= hi["irr_pct"]) or \
                   (hi["irr_pct"] <= target_irr_pct <= lo["irr_pct"]):
                    # linear interp on PPA
                    if hi["irr_pct"] != lo["irr_pct"]:
                        frac = (target_irr_pct - lo["irr_pct"]) / (hi["irr_pct"] - lo["irr_pct"])
                        ppa = lo["ppa"] + frac * (hi["ppa"] - lo["ppa"])
                    else:
                        ppa = (lo["ppa"] + hi["ppa"]) / 2
                    break
        elif target_irr_pct < min_irr:
            ppa = base_ppa * 0.75
            status = "target_below_range"
        else:
            ppa = base_ppa * 1.25
            status = "target_above_range"

        for i in range(max_iter):
            irr_cur = calc_irr(ppa)
            if irr_cur is None:
                # 엔진 발산 → 이 PPA에선 IRR 못 구함, loop 종료
                status = "engine_diverged_mid"
                break
            err = irr_cur - target_irr
            iterations.append({
                "iter": i,
                "ppa": round(ppa, 4),
                "irr_pct": round(irr_cur * 100, 4),
                "error_pct": round(err * 100, 4),
                "status": "converged" if abs(err) < tol else "iterating"
            })

            if abs(err) < tol:
                solution = {
                    "ppa": round(ppa, 4),
                    "irr_pct": round(irr_cur * 100, 4),
                    "error_pct": round(err * 100, 4),
                    "iterations": i + 1,
                    "converged": True
                }
                status = "converged"
                break

            # 미분 (central difference) — 양쪽 발산 체크
            irr_plus = calc_irr(ppa + h)
            irr_minus = calc_irr(ppa - h)
            if irr_plus is None or irr_minus is None:
                status = "engine_diverged_deriv"
                break
            derivative = (irr_plus - irr_minus) / (2 * h)

            if abs(derivative) < 1e-8:
                status = "flat_derivative"
                break

            # Newton step + 안전장치 (최대 20% 한 스텝)
            delta = -err / derivative
            max_step = base_ppa * 0.20
            if abs(delta) > max_step:
                delta = max_step if delta > 0 else -max_step
            ppa = ppa + delta

            # 음수/비정상 방지
            if ppa < 1.0:
                ppa = 1.0
            elif ppa > base_ppa * 3:
                ppa = base_ppa * 3

        if not solution:
            # 수렴 실패 - 마지막 iteration 값 사용
            last = iterations[-1] if iterations else {"ppa": base_ppa, "irr_pct": 0}
            solution = {
                "ppa": last["ppa"],
                "irr_pct": last["irr_pct"],
                "error_pct": last.get("error_pct", 0),
                "iterations": len(iterations),
                "converged": False
            }
            if status == "not_started":
                status = "max_iter_reached"

        # 추가 meta: Dev Margin 고정값 (참고용)
        base_res = _calc_engine(base_inputs)
        dev_margin_k = base_res.get("dev_margin", 0)  # $k

        return {
            "ok": True,
            "base_ppa": round(base_ppa, 2),
            "target_irr_pct": round(target_irr_pct, 2),
            "sensitivity": sensitivity,
            "iterations": iterations,
            "solution": solution,
            "status": status,
            "dev_margin_k": round(dev_margin_k, 0),
            "tolerance_pct": tol * 100,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Break-even calculation error: {str(e)}")


@app.get("/valuation/calculate/defaults")
def get_calc_defaults(mode: str = 'prediction', user=Depends(get_current_user)):
    """Return default input values for the calculator
    
    mode='prediction' (기본): 신규 프로젝트 표준 PF 가정
    mode='calibration': Neptune Case 2 재현용 파라미터
    """
    # ═══ 공통 defaults (모든 모드) ═══
    common = {
        "pv_mwac": 199, "dc_ac_ratio": 1.348,
        "pv_mwdc": 268.3,
        "bess_mw": 199, "bess_mwh": 796,
        "cf_pct": 21.24, "life": 35,
        "module_cwp": 31.5, "bos_cwp": 42.88,
        "ess_per_kwh": 234.5, "bess_bos_per_kwh": 130.0,
        "epc_cont_pct": 8.0, "owner_pct": 3.0, "softcost_pct": 5.0,
        "intercon_m": 22.5, "dev_cost_m": 20.0,
        # Dev Margin: c/Wac × (PV + BESS) MW (HWR 표준 공식, 모든 모드 공통)
        "dev_margin_cwac": 10.0, "epc_margin_pct": 7.95,
        "ppa_price": 68.82, "ppa_term": 25, "ppa_esc": 0,
        "bess_toll": 14.5, "bess_toll_term": 20,
        "merchant_ppa": 61.0, "merchant_esc": 2.0,
        "degradation": 0.0064,
        "pv_om": 4.5, "bess_om": 8.64, "insurance_pv": 10.57,
        "insurance_bess": 5.05, "asset_mgmt": 210,
        "prop_tax_yr1": 3162, "land_rent_yr1": 437, "opex_esc": 2.0,
        "aug_price": 150, "aug_mwh_pct": 18.8, "aug_y1": 4, "aug_y2": 8,
        "debt_ratio": 47.6, "int_rate": 5.5,
        "credit_mode": "ITC",
        "pv_itc_rate": 0, "bess_itc_rate": 30,
        "ptc_rate_per_kwh": 0.027,
        "itc_elig": 97,
        "flip_term": 7, "flip_yield": 8.0,
    }
    
    if mode == 'calibration':
        # ═══ Calibration (Neptune Case 2 재현) ═══
        return {**common,
            "calibration_mode": "calibration",
            "aug_y3": 14,  # Neptune: Y14
            "loan_term": 28,
            "availability_yr1": 1.0, "availability_yr2": 1.0,  # CF%에 내재
            "bess_months_per_yr": 12.72,
            "opex_etc": 0.56,
            # Capital Stack (Neptune 실측)
            "capex_total_override": 836.7,
            "te_ratio_override": 32.52,
            # Y0 Cash Flow (Neptune 실측)
            "construction_cost_m": 639.855,
            "txn_costs_m": 10.6,
            "cap_interest_m": 14.3,
            "debt_drawdown_ratio": 0.775,
            "te_proceeds_ratio": 0.935,
            # Partnership Flip Waterfall (Neptune 실측)
            "pre_flip_cash_te": 25.5,
            "post_flip_cash_te": 7,
            "depr_share": 0.7721,
            "use_nol_offset": True,
            "use_sculpted_debt": True,
            "flip_event_cf": 0,
            "flip_yield": 8.75,  # Neptune 실측
        }
    else:
        # ═══ Prediction (신규 프로젝트 표준) ═══
        return {**common,
            "calibration_mode": "prediction",
            "aug_y3": 12,  # 표준 Y12
            "loan_term": 18,
            "availability_yr1": 0.977, "availability_yr2": 0.982,
            "bess_months_per_yr": 12.0,
            "opex_etc": 0,
            # Capital Stack — 사용자 프로젝트 맞게 입력
            # (override 없이 ITC 기반 자동 계산)
            # Y0 Cash Flow — 100% drawdown (표준)
            "debt_drawdown_ratio": 1.0,
            "te_proceeds_ratio": 1.0,
            # Partnership Flip (표준 99/5)
            "pre_flip_cash_te": 99,
            "post_flip_cash_te": 5,
            "depr_share_pre": 0.01,
            "depr_share_post": 0.95,
            "use_nol_offset": False,
            "use_sculpted_debt": False,
            "flip_event_cf": 0,
        }

# ══════════════════════════════════════════════════
#  헬스체크
# ══════════════════════════════════════════════════
@app.get("/")
def root():
    return {"status": "ok", "service": "HWR Dashboard API"}

@app.get("/health")
def health():
    return {"status": "ok", "ts": datetime.datetime.now().isoformat()}

# ══════════════════════════════════════════════════
#  프로젝트 데이터 (인허가/예산/일정/메모)
# ══════════════════════════════════════════════════
@app.get("/project/{project_id}")
def get_project(project_id: str, user=Depends(get_current_user)):
    return fb_read(f"projects/{project_id}")

@app.post("/project/{project_id}")
def save_project(project_id: str, data: dict, user=Depends(get_current_user)):
    data["updatedAt"] = datetime.datetime.now().isoformat()[:16]
    data["updatedBy"] = user["email"]
    fb_write(f"projects/{project_id}", data)
    return {"ok": True}

@app.get("/projects")
def get_all_projects(user=Depends(get_current_user)):
    return fb_read("projects")

# ══════════════════════════════════════════════════
#  Valuation (PF 모델 업로드 / 조회)
# ══════════════════════════════════════════════════

def _integrity_check_pf_model(filepath: str, lang: str = 'ko') -> dict:
    """PF 엑셀 모델 정합성 체크
    
    5개 카테고리 × 심각도별 (HIGH/MEDIUM/LOW) 검사
    
    Args:
        filepath: 엑셀 파일 경로
        lang: 'ko' or 'en' - 리포트 언어
    
    Returns:
        {'checks': [...], 'summary': {...}, 'metadata': {...}}
    """
    # ─── 번역 딕셔너리 ───
    T = {
        'ko': {
            'cat_formula': '수식 오류', 'cat_capital': 'Capital Stack',
            'cat_irr': 'IRR 합리성', 'cat_debt': 'Debt Schedule',
            'cat_revenue': 'Revenue 현실성', 'cat_summary': '종합',
            # Formula errors
            'conv_fail_t': 'xlsb → xlsx 변환 실패',
            'conv_fail_d': '파일 형식 변환 중 오류 발생. 수식 분석 불가.',
            'conv_fail_a': 'LibreOffice 설치 확인 or xlsx 직접 업로드 권장',
            'conv_err_t': 'xlsb → xlsx 변환 에러',
            'conv_err_d': '변환 오류: ',
            'conv_err_a': 'xlsx로 재저장 후 업로드',
            'formula_skip_t': '수식 스트링 분석 스킵 (xlsb 직접 분석)',
            'formula_skip_d': '서버에 LibreOffice가 없어 xlsb → xlsx 변환을 건너뛰었습니다. 수식 텍스트 분석은 스킵하며, 셀 결과값(#REF!/#VALUE! 등)과 나머지 정합성 체크는 xlsb에서 직접 수행합니다.',
            'formula_skip_a': '수식 정밀 분석이 필요하면 xlsx로 저장 후 재업로드',
            'formula_t': '심각한 수식 에러 {n}개',
            'formula_d': '#NAME?/#REF!/#DIV/0 등 에러. 샘플: ',
            'formula_a': '원본 모델러에게 해당 셀 재확인 요청. Named Range 깨짐 or 외부 참조 오류 추정.',
            'na_many_t': '#N/A 값 {n}개',
            'na_many_d': 'VLOOKUP/INDEX 미스매치 or 빈 lookup 범위 가능. 정상일 수 있음.',
            'na_many_a': '핵심 계산 탭(Returns, CF, Summary)에 #N/A 있는지 집중 확인',
            'ext_ref_t': '외부 파일 참조 {n}곳',
            'ext_ref_d': '담당자 local path 참조 가능성. 샘플: ',
            'ext_ref_a': '링크 끊기(Break Links) or 값 복사 처리 필요',
            'parse_err_t': '수식 파싱 실패',
            'parse_err_d': '엑셀 구조 분석 중 오류: ',
            'parse_err_a': '파일 무결성 확인',
            # Capital Stack
            'stack_t': 'Capital Stack 합계 불일치 ({err:.1f}% 오차)',
            'stack_a': 'Summary 탭 Capital Stack 수식 재검증',
            'debt_ratio_t': 'Debt 비율 이상치 ({r:.1f}%)',
            'debt_ratio_d': '일반적 Debt 40-60% 범위를 벗어남',
            'debt_ratio_a': '금융 자문사/Debt Sizing 결과 재확인',
            'te_high_t': 'TE 비율 과다 ({r:.1f}%)',
            'te_high_d': 'TE 40% 초과 시 Sponsor 희석 심각',
            'te_high_a': 'TE 투자 계약 조건 재검토',
            'sponsor_low_t': 'Sponsor Eq 비율 과소 ({r:.1f}%)',
            'sponsor_low_d': 'Sponsor 5% 미만 시 프로젝트 관여도 낮음',
            'sponsor_low_a': 'Sponsor 실질 참여 구조 확인',
            'summary_t': 'Summary 탭 파싱 제한',
            'summary_d': '자동 체크 일부 스킵: ',
            'summary_a': '수동으로 Summary 탭 확인',
            # IRR
            'irr_high_t': '{label} 과도 ({p:.2f}%)',
            'irr_high_d': '25%+ IRR은 일반 Solar+BESS 드문 수치',
            'irr_high_a': 'Revenue/CAPEX 가정 재검토',
            'irr_low_t': '{label} 과소 ({p:.2f}%)',
            'irr_low_d': '5% 미만 IRR은 투자 매력 부족',
            'irr_low_a': 'CAPEX/OPEX/Revenue 전면 재검토',
            'irr_contract_t': 'Contract IRR > Full Life IRR ({sc:.2f}% vs {sf:.2f}%)',
            'irr_contract_d': 'Merchant 기간(Y26+) CF가 음수 또는 IRR 끌어내림. PPA 종료 후 수익성 악화 신호',
            'irr_contract_a': 'Merchant 가정(PV $61, BESS Toll 종료) 재검토',
            # Debt
            'dscr_low_t': 'DSCR 최저값 위험 ({v:.2f})',
            'dscr_low_d': 'DSCR 1.20 미만 기간 있음. 표준 covenant 1.30 위반 가능.',
            'dscr_low_a': 'Debt sizing 재검토 또는 cash sweep 조항 확인',
            'dscr_var_t': 'DSCR 편차 과다 ({lo:.2f}~{hi:.2f})',
            'dscr_var_d': 'Sculpted debt (DSCR 기반 동적 상환) 구조 추정',
            'dscr_var_a': 'Debt 탭 연도별 상환 스케줄 수동 검토 필수',
            # Revenue
            'ppa_t': 'PPA Price 이상치 (${v:.2f}/MWh)',
            'ppa_d': '일반 Solar PPA $40-90/MWh 범위 밖',
            'ppa_a': 'PPA 계약서 원본 확인',
            'merch_t': 'Merchant Price 높음 (${v:.2f}/MWh)',
            'merch_d': 'Merchant 가격 $100+ 는 공격적 가정',
            'merch_a': 'Ventyx/WM 예측 근거 확인',
            'deg_t': 'Degradation 이상치 ({v:.2f}%/yr)',
            'deg_d': '일반 0.4-0.8%/yr 범위',
            'deg_a': 'Module warranty 확인',
            # Recommend
            'rec_calib_t': '권장 계산 모드: Calibration',
            'rec_calib_d': '이 모델은 Neptune과 유사한 구조로 보입니다. 근거: ',
            'rec_calib_none': 'Neptune 패턴 일부 감지',
            'rec_calib_a': 'Dashboard에서 "🎯 Calibration" 모드로 계산 권장',
            'rec_predict_t': '권장 계산 모드: Prediction',
            'rec_predict_d': '일반 PF 구조로 판단됩니다.',
            'rec_predict_a': 'Dashboard "📈 Prediction" 모드로 계산 (기본값)',
            'clear_t': '주요 이상 징후 없음',
            'clear_d': '자동 체크한 5개 카테고리에서 큰 이슈 미발견.',
            'clear_a': '수동 검토는 여전히 권장',
            # Mode reasons
            'reason_sculpted': 'Sculpted debt 구조 감지 (DSCR 편차 큼)',
            'reason_irr_invert': 'Contract > Full IRR 역전 (Neptune 패턴)',
            'reason_debt_ratio': 'Debt 비율 {r:.1f}% (Neptune 유사 47.6%)',
            'reason_te_ratio': 'TE 비율 {r:.1f}% (Neptune 유사 32.5%)',
            # Capital Stack mismatch
            'stack_mismatch_t': 'Capital Stack 합산 불일치 ({pct:.1f}% 오차)',
            'stack_mismatch_d': 'CAPEX ${capex:,.1f}M ≠ Debt ${debt:,.1f}M + TE ${te:,.1f}M + Eq ${eq:,.1f}M (합계 ${sum:,.1f}M, 차이 ${diff:+,.1f}M)',
            'stack_mismatch_a': '모델의 Capital Stack 섹션에서 자금조달 출처별 금액 재검토 필요. 누락된 자금원 or 이중계산 가능성.',
            # DSCR
            'dscr_low_t': 'DSCR 부족 (최소 {v:.2f}x < 1.0)',
            'dscr_low_d': '특정 연도에 EBITDA가 Debt Service를 감당하지 못함. Covenant breach 가능성.',
            'dscr_low_a': 'Debt 스케줄 재검토 또는 상환 조건 완화 필요',
            'dscr_tight_t': 'DSCR 타이트 (최소 {v:.2f}x, 업계 표준 ≥ 1.20)',
            'dscr_tight_d': '기술적 Default 위험은 낮으나 업계 관행 대비 커버리지 여유 작음.',
            'dscr_tight_a': '대주단이 통상 요구하는 DSCR 1.20~1.30x 대비 낮음. 금융조건 확인',
            # Value range sanity
            'range_t': '값 범위 이상 {n}개',
            'range_a': '업계 일반 범위를 벗어난 값. 입력 오타 or 특수 구조 가능성 확인',
            'range_debt': 'Debt 비율 {v:.1f}% (정상 25~75%)',
            'range_te': 'TE 비율 {v:.1f}% (정상 0~50%)',
            'range_ppa': 'PPA ${v:.1f}/MWh (정상 $20~$200)',
            'range_flip': 'Flip Yield {v:.2f}% (정상 5~15%)',
        },
        'en': {
            'cat_formula': 'Formula Errors', 'cat_capital': 'Capital Stack',
            'cat_irr': 'IRR Validity', 'cat_debt': 'Debt Schedule',
            'cat_revenue': 'Revenue Realism', 'cat_summary': 'Summary',
            'conv_fail_t': 'xlsb → xlsx conversion failed',
            'conv_fail_d': 'Error during file format conversion. Formula analysis not possible.',
            'conv_fail_a': 'Check LibreOffice install or upload xlsx directly',
            'conv_err_t': 'xlsb → xlsx conversion error',
            'conv_err_d': 'Conversion error: ',
            'conv_err_a': 'Re-save as xlsx and re-upload',
            'formula_skip_t': 'Formula-text analysis skipped (xlsb analyzed directly)',
            'formula_skip_d': 'Server lacks LibreOffice; xlsb → xlsx conversion skipped. Formula-string analysis is bypassed, but cell-value errors (#REF!/#VALUE! etc.) and all other integrity checks are performed directly on xlsb.',
            'formula_skip_a': 'Re-save as xlsx and re-upload if detailed formula analysis needed',
            'formula_t': '{n} serious formula errors',
            'formula_d': '#NAME?/#REF!/#DIV/0 errors. Samples: ',
            'formula_a': 'Ask the original modeler to verify cells. Likely broken Named Range or external reference errors.',
            'na_many_t': '{n} #N/A values',
            'na_many_d': 'VLOOKUP/INDEX mismatch or empty lookup range possible. May be intentional.',
            'na_many_a': 'Focus on #N/A in core tabs (Returns, CF, Summary)',
            'ext_ref_t': '{n} external file references',
            'ext_ref_d': 'Possible user local path references. Samples: ',
            'ext_ref_a': 'Break links or copy-paste values required',
            'parse_err_t': 'Formula parsing failed',
            'parse_err_d': 'Error during Excel structure analysis: ',
            'parse_err_a': 'Check file integrity',
            'stack_t': 'Capital Stack sum mismatch ({err:.1f}% error)',
            'stack_a': 'Re-verify Capital Stack formula in Summary tab',
            'debt_ratio_t': 'Debt ratio outlier ({r:.1f}%)',
            'debt_ratio_d': 'Outside typical Debt 40-60% range',
            'debt_ratio_a': 'Re-verify with financial advisor/Debt Sizing',
            'te_high_t': 'TE ratio excessive ({r:.1f}%)',
            'te_high_d': 'TE over 40% means severe Sponsor dilution',
            'te_high_a': 'Review TE investment agreement terms',
            'sponsor_low_t': 'Sponsor Eq ratio too low ({r:.1f}%)',
            'sponsor_low_d': 'Sponsor below 5% suggests low project commitment',
            'sponsor_low_a': 'Verify actual Sponsor participation structure',
            'summary_t': 'Summary tab parsing limited',
            'summary_d': 'Some auto-checks skipped: ',
            'summary_a': 'Manually verify Summary tab',
            'irr_high_t': '{label} too high ({p:.2f}%)',
            'irr_high_d': '25%+ IRR is rare for Solar+BESS',
            'irr_high_a': 'Re-verify Revenue/CAPEX assumptions',
            'irr_low_t': '{label} too low ({p:.2f}%)',
            'irr_low_d': 'IRR below 5% lacks investment appeal',
            'irr_low_a': 'Comprehensive review of CAPEX/OPEX/Revenue',
            'irr_contract_t': 'Contract IRR > Full Life IRR ({sc:.2f}% vs {sf:.2f}%)',
            'irr_contract_d': 'Merchant period (Y26+) CF is negative or drags IRR down. Signal of deteriorating economics post-PPA.',
            'irr_contract_a': 'Re-verify Merchant assumptions (PV $61, BESS Toll end)',
            'dscr_low_t': 'DSCR minimum risk ({v:.2f})',
            'dscr_low_d': 'DSCR below 1.20 in some periods. Standard 1.30 covenant at risk.',
            'dscr_low_a': 'Re-verify Debt sizing or cash sweep provisions',
            'dscr_var_t': 'DSCR variance high ({lo:.2f}~{hi:.2f})',
            'dscr_var_d': 'Sculpted debt (DSCR-based dynamic amortization) likely',
            'dscr_var_a': 'Manual review of Debt tab annual schedule required',
            'ppa_t': 'PPA Price outlier (${v:.2f}/MWh)',
            'ppa_d': 'Outside typical Solar PPA $40-90/MWh range',
            'ppa_a': 'Verify original PPA contract',
            'merch_t': 'Merchant Price high (${v:.2f}/MWh)',
            'merch_d': 'Merchant price $100+ is aggressive',
            'merch_a': 'Verify Ventyx/WM forecast source',
            'deg_t': 'Degradation outlier ({v:.2f}%/yr)',
            'deg_d': 'Typical range 0.4-0.8%/yr',
            'deg_a': 'Verify Module warranty',
            'rec_calib_t': 'Recommended Mode: Calibration',
            'rec_calib_d': 'This model appears similar to Neptune structure. Reasons: ',
            'rec_calib_none': 'Some Neptune patterns detected',
            'rec_calib_a': 'Run calculations in "🎯 Calibration" mode',
            'rec_predict_t': 'Recommended Mode: Prediction',
            'rec_predict_d': 'Looks like standard PF structure.',
            'rec_predict_a': 'Run in "📈 Prediction" mode (default)',
            'clear_t': 'No major anomalies detected',
            'clear_d': 'No major issues found in 5 auto-checked categories.',
            'clear_a': 'Manual review still recommended',
            'reason_sculpted': 'Sculpted debt detected (large DSCR variance)',
            'reason_irr_invert': 'Contract > Full IRR inversion (Neptune pattern)',
            'reason_debt_ratio': 'Debt ratio {r:.1f}% (similar to Neptune 47.6%)',
            'reason_te_ratio': 'TE ratio {r:.1f}% (similar to Neptune 32.5%)',
            # Capital Stack mismatch
            'stack_mismatch_t': 'Capital Stack imbalance ({pct:.1f}% off)',
            'stack_mismatch_d': 'CAPEX ${capex:,.1f}M ≠ Debt ${debt:,.1f}M + TE ${te:,.1f}M + Eq ${eq:,.1f}M (sum ${sum:,.1f}M, diff ${diff:+,.1f}M)',
            'stack_mismatch_a': 'Review financing sources in the model. Possible missing source or double-counting.',
            # DSCR
            'dscr_low_t': 'DSCR insufficient (min {v:.2f}x < 1.0)',
            'dscr_low_d': 'EBITDA fails to cover Debt Service in some year(s). Possible covenant breach.',
            'dscr_low_a': 'Review debt schedule or relax repayment terms',
            'dscr_tight_t': 'DSCR tight (min {v:.2f}x; industry standard ≥ 1.20)',
            'dscr_tight_d': 'Technical default risk is low, but coverage cushion is thin vs market practice.',
            'dscr_tight_a': 'Below typical lender requirement of 1.20–1.30x. Confirm financing terms.',
            # Value range sanity
            'range_t': '{n} values out of normal range',
            'range_a': 'Values outside industry typical range. Check for typos or special structure.',
            'range_debt': 'Debt ratio {v:.1f}% (normal 25–75%)',
            'range_te': 'TE ratio {v:.1f}% (normal 0–50%)',
            'range_ppa': 'PPA ${v:.1f}/MWh (normal $20–$200)',
            'range_flip': 'Flip Yield {v:.2f}% (normal 5–15%)',
        },
    }
    t = T.get(lang, T['ko'])
    
    from openpyxl import load_workbook
    from pyxlsb import open_workbook as open_xlsb
    import tempfile, subprocess, os
    
    checks = []
    metadata = {}
    
    # xlsb인 경우 xlsx로 변환 (openpyxl로 수식 읽기 위해)
    # libreoffice 미설치 환경(Railway 등)에서는 변환 실패 시
    # 수식 체크만 스킵하고 나머지 (Capital Stack / IRR / Debt / Revenue)는 pyxlsb로 진행
    xlsx_path = filepath
    if filepath.endswith('.xlsb'):
        try:
            subprocess.run(['libreoffice', '--headless', '--convert-to', 'xlsx',
                          '--outdir', os.path.dirname(filepath), filepath],
                         check=True, capture_output=True, timeout=60)
            candidate = filepath.replace('.xlsb', '.xlsx')
            if os.path.exists(candidate):
                xlsx_path = candidate
            else:
                xlsx_path = None
                metadata['formula_text_analysis'] = 'skipped_no_libreoffice'
        except Exception:
            xlsx_path = None
            metadata['formula_text_analysis'] = 'skipped_no_libreoffice'
    
    # ═══ 1. 수식 오류 체크 (xlsx_path 있을 때만) ═══
    if xlsx_path and xlsx_path != filepath:
        try:
            wb = load_workbook(xlsx_path, data_only=False)
            sheet_names = wb.sheetnames
            metadata['sheets'] = sheet_names
            metadata['sheet_count'] = len(sheet_names)
            
            formula_errors = []    # 심각한 수식 에러
            na_errors = []
            external_refs = []
            
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        val_str = str(cell.value)
                        for err in ['#NAME?', '#REF!', '#DIV/0!', '#VALUE!', '#NULL!']:
                            if err in val_str:
                                formula_errors.append({
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'error': err,
                                    'formula': val_str[:100],
                                })
                                break
                        if '#N/A' in val_str:
                            na_errors.append({'sheet': sheet_name, 'cell': cell.coordinate})
                        if '[' in val_str and ('\\' in val_str or '.xls' in val_str):
                            if not any(e in val_str for e in ['#NAME?', '#REF!']):
                                external_refs.append({
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'ref': val_str[:150],
                                })
            
            if formula_errors:
                sample = formula_errors[:5]
                sample_text = '; '.join([f"{e['sheet']}!{e['cell']}({e['error']})" for e in sample])
                from collections import Counter
                sheet_counts = Counter(e['sheet'] for e in formula_errors)
                sheet_samples = {}
                for e in formula_errors:
                    sh = e['sheet']
                    if sh not in sheet_samples:
                        sheet_samples[sh] = []
                    if len(sheet_samples[sh]) < 3:
                        sheet_samples[sh].append(e['cell'])
                detail_by_sheet = [
                    {'sheet': sh, 'count': cnt, 'cells': sheet_samples.get(sh, [])}
                    for sh, cnt in sheet_counts.most_common()
                ]
                checks.append({
                    'category': t['cat_formula'], 'severity': 'HIGH',
                    'code': 'FORMULA_ERR',
                    'title': t['formula_t'].format(n=len(formula_errors)),
                    'description': t['formula_d'] + sample_text,
                    'action': t['formula_a'],
                    'detail_by_sheet': detail_by_sheet,
                })
            
            if len(na_errors) > 50:
                checks.append({
                    'category': t['cat_formula'], 'severity': 'LOW',
                    'code': 'NA_MANY',
                    'title': t['na_many_t'].format(n=len(na_errors)),
                    'description': t['na_many_d'],
                    'action': t['na_many_a'],
                })
            
            if external_refs:
                sample = external_refs[:3]
                sample_text = '; '.join([f"{e['sheet']}!{e['cell']}" for e in sample])
                checks.append({
                    'category': t['cat_formula'], 'severity': 'MEDIUM',
                    'code': 'EXT_REF',
                    'title': t['ext_ref_t'].format(n=len(external_refs)),
                    'description': t['ext_ref_d'] + sample_text,
                    'action': t['ext_ref_a'],
                    'detail': external_refs[:10],
                })
        except Exception as e:
            checks.append({
                'category': t['cat_formula'], 'severity': 'HIGH',
                'code': 'PARSE_ERR',
                'title': t['parse_err_t'],
                'description': t['parse_err_d'] + str(e)[:100],
                'action': t['parse_err_a'],
            })
    else:
        # xlsx 변환 불가 → 수식 체크 스킵, 메타데이터는 pyxlsb로 수집
        try:
            xlsb_path = filepath if filepath.endswith('.xlsb') else filepath.replace('.xlsx', '.xlsb')
            with open_xlsb(xlsb_path) as wb_v:
                sheet_names = list(wb_v.sheets)
                metadata['sheets'] = sheet_names
                metadata['sheet_count'] = len(sheet_names)
                # xlsb에서 #REF! / #VALUE! 등 에러 값 전수 조사
                formula_errors = []
                na_count = 0
                # Excel 컬럼 인덱스 → 문자 변환 (0→A, 25→Z, 26→AA)
                def col_letter(idx):
                    s = ''
                    idx += 1  # 1-based
                    while idx > 0:
                        idx, r = divmod(idx - 1, 26)
                        s = chr(65 + r) + s
                    return s
                for sheet_name in sheet_names:
                    try:
                        with wb_v.get_sheet(sheet_name) as ws:
                            for r_idx, row in enumerate(ws.rows()):
                                for cell in row:
                                    if cell.v is None: continue
                                    val_str = str(cell.v)
                                    # 치명 에러
                                    for err in ['#NAME?', '#REF!', '#DIV/0!', '#VALUE!', '#NULL!']:
                                        if err in val_str:
                                            try:
                                                col_idx = cell.c  # pyxlsb cell has .c (column index, 0-based)
                                            except AttributeError:
                                                col_idx = None
                                            coord = f"{col_letter(col_idx)}{r_idx+1}" if col_idx is not None else f"R{r_idx+1}"
                                            formula_errors.append({
                                                'sheet': sheet_name,
                                                'cell': coord,
                                                'error': err,
                                            })
                                            break
                                    # #N/A 별도 카운트 (정상일 수도 있어서 50개 넘을 때만 경고)
                                    if '#N/A' in val_str:
                                        na_count += 1
                    except Exception:
                        continue
                
                # HIGH: 치명 에러
                if formula_errors:
                    sample = formula_errors[:8]
                    sample_text = '; '.join([f"{e['sheet']}!{e['cell']}({e['error']})" for e in sample])
                    # 시트별 집계 (전체 기반, 상위 샘플 아닌 전체)
                    from collections import Counter
                    sheet_counts = Counter(e['sheet'] for e in formula_errors)
                    # 각 시트에서 처음 발견된 에러 3개씩만 샘플 유지
                    sheet_samples = {}
                    for e in formula_errors:
                        sh = e['sheet']
                        if sh not in sheet_samples:
                            sheet_samples[sh] = []
                        if len(sheet_samples[sh]) < 3:
                            sheet_samples[sh].append(e['cell'])
                    # detail_by_sheet: 프론트에서 바로 렌더 가능한 구조
                    detail_by_sheet = [
                        {'sheet': sh, 'count': cnt, 'cells': sheet_samples.get(sh, [])}
                        for sh, cnt in sheet_counts.most_common()
                    ]
                    checks.append({
                        'category': t['cat_formula'], 'severity': 'HIGH',
                        'code': 'FORMULA_ERR_VAL',
                        'title': t['formula_t'].format(n=len(formula_errors)),
                        'description': t['formula_d'] + sample_text,
                        'action': t['formula_a'],
                        'detail_by_sheet': detail_by_sheet,
                    })
                # LOW: #N/A 과다
                if na_count > 50:
                    checks.append({
                        'category': t['cat_formula'], 'severity': 'LOW',
                        'code': 'NA_MANY_VAL',
                        'title': t['na_many_t'].format(n=na_count),
                        'description': t['na_many_d'],
                        'action': t['na_many_a'],
                    })
                metadata['formula_error_count'] = len(formula_errors)
                metadata['na_count'] = na_count
        except Exception:
            pass
    
    # ═══ 2. Capital Stack 정합성 (값 기반 체크) ═══
    try:
        with open_xlsb(filepath if filepath.endswith('.xlsb') else filepath.replace('.xlsx', '.xlsb')) as wb_v:
            try:
                with wb_v.get_sheet('Summary') as ws:
                    rows = list(ws.rows())
                    # 각 행에서 label + 숫자 값들 추출
                    # Neptune은 Case 1 / Case 2 두 컬럼 — 마지막 숫자(Case 2) 우선
                    summary_rows = []
                    for row in rows:
                        vals = [(c.c, c.v) for c in row if c.v is not None]
                        if len(vals) < 2: continue
                        # 첫 문자열 = label
                        label = None
                        for c, v in vals:
                            if isinstance(v, str):
                                label = v.strip()
                                break
                        # 숫자 값들 (label 이후)
                        nums = [v for c, v in vals if isinstance(v, (int, float))]
                        if label and nums:
                            # Case 2 (마지막 유효값) 우선
                            summary_rows.append((label, nums[-1], nums))
                    
                    metadata['summary_rows_found'] = len(summary_rows)
                    
                    # 정확한 라벨 매칭 (Neptune 기준)
                    label_map = {
                        'total project cost': 'capex_m',
                        'debt': 'debt_m',
                        'tax equity investment': 'te_m',
                        'sponsor equity investment': 'sponsor_m',
                        'hqc dev margin': 'dev_margin_m',
                        'levered project irr (full life)': 'lev_irr',
                        'unlevered project irr (full life)': 'unlev_irr',
                        'sponsor levered irr (full life)': 'sponsor_full_irr',
                        'sponsor levered irr (contract)': 'sponsor_contract_irr',
                    }
                    
                    extracted = {}
                    for label, val, all_nums in summary_rows:
                        key = label.lower().strip()
                        # 콜론, 괄호 등 정리
                        key_clean = key.replace(':', '').replace('  ', ' ').strip()
                        for pattern, field in label_map.items():
                            if pattern in key_clean:
                                extracted[field] = val
                                break
                    
                    metadata.update(extracted)
                    
                    capex_m = extracted.get('capex_m')
                    debt_m = extracted.get('debt_m')
                    te_m = extracted.get('te_m')
                    sponsor_m = extracted.get('sponsor_m')
                    
                    # 체크: 합계 일치
                    if all(v is not None for v in [capex_m, debt_m, te_m, sponsor_m]):
                        total = debt_m + te_m + sponsor_m
                        err_pct = abs(total - capex_m) / capex_m * 100 if capex_m else 0
                        if err_pct > 1:
                            checks.append({
                                'category': t['cat_capital'], 'severity': 'HIGH',
                                'code': 'STACK_MISMATCH',
                                'title': t['stack_t'].format(err=err_pct),
                                'description': f'Debt ${debt_m:,.0f}K + TE ${te_m:,.0f}K + Sponsor ${sponsor_m:,.0f}K = ${total:,.0f}K, CAPEX ${capex_m:,.0f}K',
                                'action': t['stack_a'],
                            })
                        
                        debt_ratio = debt_m / capex_m * 100 if capex_m else 0
                        te_ratio = te_m / capex_m * 100 if capex_m else 0
                        sponsor_ratio = sponsor_m / capex_m * 100 if capex_m else 0
                        
                        metadata['debt_ratio'] = round(debt_ratio, 2)
                        metadata['te_ratio'] = round(te_ratio, 2)
                        metadata['sponsor_ratio'] = round(sponsor_ratio, 2)
                        
                        if debt_ratio < 30 or debt_ratio > 70:
                            checks.append({
                                'category': t['cat_capital'], 'severity': 'MEDIUM',
                                'code': 'DEBT_RATIO',
                                'title': t['debt_ratio_t'].format(r=debt_ratio),
                                'description': t['debt_ratio_d'],
                                'action': t['debt_ratio_a'],
                            })
                        if te_ratio > 45:
                            checks.append({
                                'category': t['cat_capital'], 'severity': 'MEDIUM',
                                'code': 'TE_RATIO_HIGH',
                                'title': t['te_high_t'].format(r=te_ratio),
                                'description': t['te_high_d'],
                                'action': t['te_high_a'],
                            })
                        if sponsor_ratio < 5:
                            checks.append({
                                'category': t['cat_capital'], 'severity': 'MEDIUM',
                                'code': 'SPONSOR_LOW',
                                'title': t['sponsor_low_t'].format(r=sponsor_ratio),
                                'description': t['sponsor_low_d'],
                                'action': t['sponsor_low_a'],
                            })
                    
                    # IRR 체크 (Summary 탭에서 직접)
                    sponsor_full = extracted.get('sponsor_full_irr')
                    sponsor_contract = extracted.get('sponsor_contract_irr')
                    lev_irr = extracted.get('lev_irr')
                    unlev_irr = extracted.get('unlev_irr')
                    
                    for label, val in [
                        ('Sponsor IRR Full Life', sponsor_full),
                        ('Sponsor IRR Contract', sponsor_contract),
                        ('Project Levered IRR', lev_irr),
                        ('Project Unlevered IRR', unlev_irr),
                    ]:
                        if val is not None:
                            # 0~1 → percent
                            pct = val * 100 if 0 < val < 1 else val
                            if pct > 25:
                                checks.append({
                                    'category': t['cat_irr'], 'severity': 'MEDIUM',
                                    'code': 'IRR_HIGH',
                                    'title': t['irr_high_t'].format(label=label, p=pct),
                                    'description': t['irr_high_d'],
                                    'action': t['irr_high_a'],
                                })
                            elif pct < 5:
                                checks.append({
                                    'category': t['cat_irr'], 'severity': 'HIGH',
                                    'code': 'IRR_LOW',
                                    'title': t['irr_low_t'].format(label=label, p=pct),
                                    'description': t['irr_low_d'],
                                    'action': t['irr_low_a'],
                                })
                    
                    # Contract > Full 역전 체크
                    if sponsor_full is not None and sponsor_contract is not None:
                        sf = sponsor_full * 100 if 0 < sponsor_full < 1 else sponsor_full
                        sc = sponsor_contract * 100 if 0 < sponsor_contract < 1 else sponsor_contract
                        if sc > sf + 2:  # Contract가 Full보다 2%p 이상 높으면
                            checks.append({
                                'category': t['cat_irr'], 'severity': 'MEDIUM',
                                'code': 'IRR_CONTRACT_HIGH',
                                'title': t['irr_contract_t'].format(sc=sc, sf=sf),
                                'description': t['irr_contract_d'],
                                'action': t['irr_contract_a'],
                            })
            except Exception as e:
                checks.append({
                    'category': t['cat_capital'], 'severity': 'LOW',
                    'code': 'SUMMARY_PARSE',
                    'title': t['summary_t'],
                    'description': t['summary_d'] + str(e)[:80],
                    'action': t['summary_a'],
                })
    except Exception:
        pass
    
    # ═══ 4. Debt Schedule 체크 ═══
    try:
        with open_xlsb(filepath if filepath.endswith('.xlsb') else None) as wb_v:
            with wb_v.get_sheet('Debt') as ws:
                rows = list(ws.rows())
                # DSCR 라인 탐색 (보통 'DSCR' 라벨이 A열에)
                dscr_values = []
                for r_idx, row in enumerate(rows):
                    label = None
                    for c in row:
                        if c.v is not None:
                            label = str(c.v) if isinstance(c.v, str) else None
                            break
                    if label and 'DSCR' in label.upper():
                        # 이 row의 모든 숫자 값
                        vals = [c.v for c in row if c.v is not None and isinstance(c.v, (int, float)) and 0.5 < c.v < 5]
                        if vals:
                            dscr_values = vals[:40]  # 첫 40개
                            break
                
                if dscr_values:
                    metadata['dscr_sample'] = dscr_values[:10]
                    min_dscr = min(dscr_values)
                    max_dscr = max(dscr_values)
                    
                    if min_dscr < 1.20:
                        checks.append({
                            'category': t['cat_debt'], 'severity': 'HIGH',
                            'code': 'DSCR_LOW',
                            'title': t['dscr_low_t'].format(v=min_dscr),
                            'description': t['dscr_low_d'],
                            'action': t['dscr_low_a'],
                        })
                    if max_dscr - min_dscr > 1.0:
                        checks.append({
                            'category': t['cat_debt'], 'severity': 'MEDIUM',
                            'code': 'DSCR_VARIANCE',
                            'title': t['dscr_var_t'].format(lo=min_dscr, hi=max_dscr),
                            'description': t['dscr_var_d'],
                            'action': t['dscr_var_a'],
                        })
    except Exception:
        pass
    
    # ═══ 5. Revenue / OPEX 현실성 ═══
    try:
        with open_xlsb(filepath if filepath.endswith('.xlsb') else None) as wb_v:
            # Dash 탭에서 주요 가정 확인
            try:
                with wb_v.get_sheet('Dash') as ws:
                    dash_data = {}
                    for row in ws.rows():
                        vals = [c.v for c in row]
                        labels = [str(v) for v in vals if isinstance(v, str)]
                        nums = [v for v in vals if isinstance(v, (int, float))]
                        if labels and nums:
                            label = labels[0].strip()
                            dash_data[label] = nums[0]
                    
                    # PPA Price 체크
                    for k, v in dash_data.items():
                        k_lower = k.lower()
                        if 'ppa' in k_lower and 'price' in k_lower:
                            if v < 20 or v > 150:
                                checks.append({
                                    'category': t['cat_revenue'], 'severity': 'MEDIUM',
                                    'code': 'PPA_RANGE',
                                    'title': t['ppa_t'].format(v=v),
                                    'description': t['ppa_d'],
                                    'action': t['ppa_a'],
                                })
                            metadata['ppa_price'] = v
                        elif 'merchant' in k_lower and 'price' in k_lower:
                            if v > 100:
                                checks.append({
                                    'category': t['cat_revenue'], 'severity': 'LOW',
                                    'code': 'MERCHANT_HIGH',
                                    'title': t['merch_t'].format(v=v),
                                    'description': t['merch_d'],
                                    'action': t['merch_a'],
                                })
                            metadata['merchant_price'] = v
                        elif 'degradation' in k_lower:
                            deg_pct = v * 100 if v < 0.1 else v
                            if deg_pct == 0 or deg_pct > 1.5:
                                checks.append({
                                    'category': t['cat_revenue'], 'severity': 'MEDIUM',
                                    'code': 'DEGRADATION',
                                    'title': t['deg_t'].format(v=deg_pct),
                                    'description': t['deg_d'],
                                    'action': t['deg_a'],
                                })
                            metadata['degradation_pct'] = deg_pct
            except Exception:
                pass
    except Exception:
        pass
    # ═══ 추가 체크: DSCR 합리성 ═══
    try:
        dscr_samples = metadata.get('dscr_sample') or []
        if dscr_samples:
            min_dscr = min(dscr_samples)
            if min_dscr < 1.0:
                checks.append({
                    'category': t['cat_debt'], 'severity': 'HIGH',
                    'code': 'DSCR_INSUFFICIENT',
                    'title': t['dscr_low_t'].format(v=min_dscr),
                    'description': t['dscr_low_d'],
                    'action': t['dscr_low_a'],
                })
            elif min_dscr < 1.20:
                checks.append({
                    'category': t['cat_debt'], 'severity': 'MEDIUM',
                    'code': 'DSCR_TIGHT',
                    'title': t['dscr_tight_t'].format(v=min_dscr),
                    'description': t['dscr_tight_d'],
                    'action': t['dscr_tight_a'],
                })
    except Exception:
        pass

    # ═══ 추가 체크: 값 범위 sanity ═══
    try:
        checks_range = []
        debt_r = metadata.get('debt_ratio', 0)
        te_r = metadata.get('te_ratio', 0)
        ppa = metadata.get('ppa_price', 0)
        
        if debt_r > 0 and not (25 <= debt_r <= 75):
            checks_range.append(t['range_debt'].format(v=debt_r))
        if te_r > 0 and not (0 <= te_r <= 50):
            checks_range.append(t['range_te'].format(v=te_r))
        if ppa > 0 and not (20 <= ppa <= 200):
            checks_range.append(t['range_ppa'].format(v=ppa))
        
        if checks_range:
            checks.append({
                'category': t['cat_summary'], 'severity': 'MEDIUM',
                'code': 'RANGE_ANOMALY',
                'title': t['range_t'].format(n=len(checks_range)),
                'description': '; '.join(checks_range),
                'action': t['range_a'],
            })
    except Exception:
        pass

    # ═══ 체크 없음 (No Issues) — 긍정 신호 ═══
    if not checks:
        checks.append({
            'category': t['cat_summary'], 'severity': 'OK',
            'code': 'ALL_CLEAR',
            'title': t['clear_t'],
            'description': t['clear_d'],
            'action': t['clear_a'],
        })
    
    # 심각도별 요약
    summary = {
        'total': len(checks),
        'high': sum(1 for c in checks if c.get('severity') == 'HIGH'),
        'medium': sum(1 for c in checks if c.get('severity') == 'MEDIUM'),
        'low': sum(1 for c in checks if c.get('severity') == 'LOW'),
        'ok': sum(1 for c in checks if c.get('severity') == 'OK'),
    }
    
    return {
        'checks': checks,
        'summary': summary,
        'metadata': metadata,
    }


@app.post("/valuation/integrity-check")
async def integrity_check_upload(
    file: UploadFile = File(...),
    lang: str = 'ko',
    user=Depends(get_current_user)
):
    """PF 엑셀 모델 정합성 체크 (5개 카테고리 × HIGH/MEDIUM/LOW). lang=ko|en"""
    if not (file.filename.endswith(".xlsb") or file.filename.endswith(".xlsx")):
        raise HTTPException(400, "xlsb 또는 xlsx 파일만 가능합니다.")
    
    suffix = ".xlsb" if file.filename.endswith(".xlsb") else ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    
    try:
        result = _integrity_check_pf_model(tmp_path, lang=lang)
        result['filename'] = file.filename
        return {"ok": True, **result}
    except Exception as e:
        raise HTTPException(500, f"정합성 체크 오류: {str(e)}")
    finally:
        try:
            import os
            os.unlink(tmp_path)
            # xlsb 변환본도 삭제
            if tmp_path.endswith('.xlsb') and os.path.exists(tmp_path.replace('.xlsb', '.xlsx')):
                os.unlink(tmp_path.replace('.xlsb', '.xlsx'))
        except Exception:
            pass


def parse_pf_model(filepath: str) -> dict:
    """xlsb에서 핵심 가정값과 아웃풋 추출"""
    assumptions = {}
    outputs = {}

    with open_workbook(filepath) as wb:

        # ── PF Intake → assumptions ───────────────
        try:
            with wb.get_sheet("PF Intake") as ws:
                for row in ws.rows():
                    vals = [c.v for c in row if c.v is not None]
                    if len(vals) < 2:
                        continue
                    label = str(vals[0]).strip()
                    val   = vals[1] if len(vals) > 1 else None

                    intake_map = {
                        "Project Name":             ("project_name", str),
                        "PJ Characteristic":        ("technology",   str),
                        "State":                    ("state",        str),
                        "PV : Project Size (MWac)": ("pv_mwac",      float),
                        "NTP Date":                 ("ntp",          str),
                        "COD":                      ("cod",          str),
                        "DC/AC Ratio":              ("dc_ac_ratio",  float),
                        "Total Site Area":          ("site_area_ac", float),
                    }
                    if label in intake_map:
                        key, typ = intake_map[label]
                        try:
                            assumptions[key] = typ(val) if val is not None else None
                        except Exception:
                            assumptions[key] = str(val) if val is not None else None

                    # BESS: ESS size 행은 "4hr" 같은 duration값 → bess_duration으로 저장
                    if label == "ESS size (MW)":
                        assumptions["bess_duration"] = str(val) if val is not None else None

                    # BESS: ESS Duration 행에서 실제 MW 추출
                    if label == "ESS Duration (Hours)":
                        try:
                            assumptions["bess_mw"] = float(val)
                        except Exception:
                            pass

                    # BESS: ESS storage MWh — hex/int 모두 처리
                    if label == "ESS storage size (MWh)":
                        try:
                            v = val
                            if isinstance(v, str) and v.startswith("0x"):
                                assumptions["bess_mwh"] = float(int(v, 16))
                            else:
                                assumptions["bess_mwh"] = float(v)
                        except Exception:
                            assumptions["bess_mwh"] = None
        except Exception:
            pass

        # ── Quarterly Assumptions → 운영 가정 ─────
        try:
            with wb.get_sheet("Quarterly Assumptions") as ws:
                for row in ws.rows():
                    vals = [c.v for c in row if c.v is not None]
                    if len(vals) < 3:
                        continue
                    label = str(vals[0]).strip()
                    val   = vals[2] if len(vals) > 2 else None

                    qa_map = {
                        "Degradation":              ("degradation",       float),
                        "Availability (yr 1)":      ("availability_yr1",  float),
                        "Availability (yr 2+)":     ("availability_yr2",  float),
                        "PV Covered O&M":           ("pv_om_covered",     float),
                        "PV Non-covered O&M":       ("pv_om_noncovered",  float),
                        "Asset management < 200MW": ("asset_mgmt_sm",     float),
                        "Asset management > 200MW": ("asset_mgmt_lg",     float),
                        "PV Merchant Haircut":      ("merchant_haircut",  float),
                    }
                    if label in qa_map:
                        key, typ = qa_map[label]
                        try:
                            assumptions[key] = typ(val) if val is not None else None
                        except Exception:
                            pass
        except Exception:
            pass

        # ── Summary → outputs (Case 2 = PV+BESS 컬럼 기준, index 3) ──
        try:
            with wb.get_sheet("Summary") as ws:
                for row in ws.rows():
                    vals = [c.v for c in row if c.v is not None]
                    if len(vals) < 3:
                        continue
                    label = str(vals[0]).strip()

                    summary_map = {
                        "levered project IRR (full life)":   "levered_irr",
                        "Unlevered project IRR (full life)": "unlevered_irr",
                        "Sponsor levered IRR (full life)":   "sponsor_irr",
                        "Sponsor levered IRR (contract)":    "sponsor_irr_contract",
                        # 신규: After-Tax IRR (Class B 관점) + WACC
                        "Sponsor levered after-tax IRR (before NOL)":  "sponsor_irr_aftertax_before_nol",
                        "Sponsor levered after-tax IRR (after NOL)":   "sponsor_irr_aftertax_after_nol",
                        "Weighted average cost of capital":  "wacc",
                        "WACC":                              "wacc",
                        "Total Project Cost":                "capex_total",
                        "Debt":                              "debt",
                        "Tax Equity Investment":             "tax_equity",
                        "Sponsor Equity Investment":         "sponsor_equity",
                        "PV : PPA Price":                    "ppa_price",
                        "PV : PPA term":                     "ppa_term",
                        "BESS : Toll rate":                  "bess_toll",
                        "BESS : Toll term":                  "bess_toll_term",
                        "HQC DEV Margin (000$)":             "dev_margin",
                        "Total Margin (000$)":               "total_margin",
                    }
                    if label in summary_map:
                        key = summary_map[label]
                        # Case 2 (PV+BESS) = vals[3], fallback to vals[2]
                        val = vals[3] if len(vals) > 3 else (vals[2] if len(vals) > 2 else None)
                        try:
                            v = float(val)
                            if key in ("levered_irr", "unlevered_irr",
                                       "sponsor_irr", "sponsor_irr_contract",
                                       "sponsor_irr_aftertax_before_nol",
                                       "sponsor_irr_aftertax_after_nol",
                                       "wacc"):
                                outputs[key] = round(v, 6)
                            else:
                                outputs[key] = round(v, 2)
                        except Exception:
                            pass
        except Exception:
            pass

        # ── Returns 시트 → After-Tax IRR (Before/After NOL) 및 기타 세분화 IRR ──
        # Returns 시트의 'Sponsor net aftertax cashflow' 줄에 IRR이 있음
        # 각 IRR 값은 보통 4~5번째 컬럼 위치에 있고, 레이블은 맨 앞
        try:
            with wb.get_sheet("Returns") as ws:
                rows_list = list(ws.rows())
                # 라인 순서대로 처리 (NOL 이전 aftertax는 첫 번째 매칭, 이후는 두 번째)
                aftertax_matches = []
                unlevered_aftertax_matches = []
                for row in rows_list:
                    vals = [c.v for c in row]
                    label = ""
                    # 첫 번째 문자열 셀을 레이블로
                    for v in vals[:3]:
                        if isinstance(v, str) and v.strip():
                            label = v.strip()
                            break
                    if not label:
                        continue
                    # IRR 숫자 찾기 (0 < v < 1 범위의 float)
                    irr_val = None
                    for v in vals:
                        if isinstance(v, float) and 0.001 < v < 0.5 and v != 1.0:
                            # 첫 번째 그럴듯한 IRR 값 (label 이후)
                            irr_val = round(v, 6)
                            break
                    if irr_val is None:
                        continue

                    # Sponsor net pretax cashflow — Levered Pre-Tax
                    #   "(without ITC or PTC)" = baseline (Line 25, ~10.02%)
                    #   "(with PTC)" 는 PTC 모델이므로 제외
                    if (label.startswith("Sponsor net pretax cashflow")
                        and "unlevered" not in label.lower()
                        and "with ptc" not in label.lower()
                        and "with itc" not in label.lower()):
                        if "sponsor_irr_levered_pretax" not in outputs:
                            outputs["sponsor_irr_levered_pretax"] = irr_val
                    # Sponsor net unlevered pretax cashflow
                    elif (label.startswith("Sponsor net unlevered pretax")
                          and "with ptc" not in label.lower()
                          and "with itc" not in label.lower()):
                        if "sponsor_irr_unlevered_pretax" not in outputs:
                            outputs["sponsor_irr_unlevered_pretax"] = irr_val
                    # Sponsor net aftertax cashflow (level IRR, NOL 전/후 두 줄)
                    #   - 첫 등장 = Before NOL (~13.62%)
                    #   - 두 번째 등장 (NOL effect 처리 후) = After NOL (~10.51%)
                    # "(including Residual Value)" 및 State Tax 버전은 제외
                    elif (label == "Sponsor net aftertax cashflow"
                          and "residual" not in label.lower()
                          and "state" not in label.lower()):
                        aftertax_matches.append(irr_val)
                    # Sponsor net unlevered aftertax cashflow
                    elif (label == "Sponsor net unlevered aftertax cashflow"
                          or label == "Sponsor net unlevered aftertax cashflow with NOL"):
                        unlevered_aftertax_matches.append(irr_val)

                # 매칭 순서 기반: first = before NOL, second = after NOL
                if len(aftertax_matches) >= 1:
                    outputs["sponsor_irr_aftertax_before_nol"] = aftertax_matches[0]
                if len(aftertax_matches) >= 2:
                    # 두 번째 매칭이 After NOL (세 번째 이상은 State Tax 변형)
                    outputs["sponsor_irr_aftertax_after_nol"] = aftertax_matches[1]
                if len(unlevered_aftertax_matches) >= 1:
                    outputs["sponsor_irr_unlevered_aftertax_before_nol"] = unlevered_aftertax_matches[0]
                if len(unlevered_aftertax_matches) >= 2:
                    outputs["sponsor_irr_unlevered_aftertax_after_nol"] = unlevered_aftertax_matches[1]
        except Exception:
            pass

        # ── Sensitivities 시트 → WACC, Cost of Debt ──
        # "Weighted average cost of capital" 레이블이 있는 행에서 2번째 컬럼 값
        try:
            with wb.get_sheet("Sensitivities") as ws:
                for row in ws.rows():
                    vals = [c.v for c in row]
                    label = ""
                    for v in vals[:4]:
                        if isinstance(v, str) and v.strip():
                            label = v.strip()
                            break
                    if not label:
                        continue
                    low = label.lower()
                    # WACC - "Weighted average cost of capital"
                    if "weighted average cost of capital" in low and "wacc" not in outputs:
                        for v in vals:
                            if isinstance(v, float) and 0.01 < v < 0.3:
                                outputs["wacc"] = round(v, 6)
                                break
                    # Cost of debt
                    elif label.lower().strip() == "cost of debt" and "cost_of_debt" not in outputs:
                        for v in vals:
                            if isinstance(v, float) and 0.01 < v < 0.3:
                                outputs["cost_of_debt"] = round(v, 6)
                                break
        except Exception:
            pass

    # ── CF_Annual → 연도별 실제 수익 추출
    try:
        with wb.get_sheet("CF_Annual") as ws:
            for row in ws.rows():
                vals = [c.v for c in row if c.v is not None]
                if len(vals) < 5: continue
                label = str(vals[0]).strip()
                # Y1 시작 인덱스: 앞 4개(total, pre-COD, 0, 0) 제거 후 운영연도
                op_vals = [v for v in vals[1:] if isinstance(v, (int, float))]

                if "PPA #2 BESS" in label and "Revenue" in label:
                    try:
                        bess_rev_y1 = float(op_vals[4]) if len(op_vals) > 4 else 0
                        bess_mw = assumptions.get("bess_mw") or 199
                        if bess_rev_y1 > 0 and bess_mw > 0:
                            outputs["bess_toll_y1_effective"] = round(bess_rev_y1/(bess_mw*1000*12)*1000, 4)
                            outputs["bess_rev_y1"] = round(bess_rev_y1, 0)
                        # 연도별 BESS 수익 (인덱스 4~38 = Y1~Y35)
                        outputs["bess_rev_schedule"] = [round(float(v),0) for v in op_vals[4:39] if isinstance(v,(int,float))]
                    except: pass

                if "PPA #1 PV" in label and "Revenue" in label:
                    try:
                        pv_rev_y1 = float(op_vals[4]) if len(op_vals) > 4 else 0
                        if pv_rev_y1 > 0:
                            outputs["pv_rev_y1"] = round(pv_rev_y1, 0)
                        outputs["pv_rev_schedule"] = [round(float(v),0) for v in op_vals[4:39] if isinstance(v,(int,float))]
                    except: pass

                if "Merchant PV Power Revenue" in label:
                    try:
                        outputs["merch_rev_schedule"] = [round(float(v),0) for v in op_vals[4:39] if isinstance(v,(int,float))]
                    except: pass
    except Exception:
        pass

    # bess_mwh 보정: xlsb hex 파싱 한계 → pv_mwac × duration(숫자)으로 계산
    try:
        duration_str = assumptions.get("bess_duration", "")
        duration_h = float("".join(x for x in str(duration_str) if x.isdigit() or x=="."))
        pv_mwac = assumptions.get("pv_mwac") or assumptions.get("bess_mw")
        if pv_mwac and duration_h:
            assumptions["bess_mwh"] = round(float(pv_mwac) * duration_h, 1)
    except Exception:
        pass

    return {"assumptions": assumptions, "outputs": outputs}


@app.post("/valuation/upload")
async def upload_valuation(
    project_id: str = Form(...),
    scenario:   str = Form(default=""),
    reason:     str = Form(default=""),
    approver:   str = Form(default=""),
    file: UploadFile = File(...),
    user=Depends(get_current_user)
):
    """PF 재무모델(xlsb/xlsx) 업로드 → AI 파싱 → Firebase 저장"""
    if not (file.filename.endswith(".xlsb") or file.filename.endswith(".xlsx")):
        raise HTTPException(400, "xlsb 또는 xlsx 파일만 업로드 가능합니다.")

    suffix = ".xlsb" if file.filename.endswith(".xlsb") else ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        parsed = parse_pf_model(tmp_path)
    except Exception as e:
        raise HTTPException(500, f"모델 파싱 실패: {str(e)}")
    finally:
        os.unlink(tmp_path)

    ts      = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    safe_id = project_id.replace("/", "_").replace(".", "_")

    payload = {
        "uploaded_at": datetime.datetime.now().isoformat(),
        "uploaded_by": user["email"],
        "filename":    file.filename,
        "scenario":    scenario,
        "reason":      reason,
        "approver":    approver,
        "assumptions": parsed["assumptions"],
        "outputs":     parsed["outputs"],
    }

    fb_put(f"valuation/{safe_id}/versions/{ts}", payload)
    fb_put(f"valuation/{safe_id}/latest", payload)

    return {
        "ok":         True,
        "project_id": safe_id,
        "timestamp":  ts,
        "parsed":     parsed,
    }


@app.get("/valuation")
def get_all_valuations(user=Depends(get_current_user)):
    """전체 프로젝트 latest 비교 (Valuation 탭용)"""
    all_data = fb_read("valuation") or {}
    result = {}
    for pid, pdata in all_data.items():
        if isinstance(pdata, dict) and "latest" in pdata:
            result[pid] = pdata["latest"]
    return result


@app.get("/valuation/{project_id}")
def get_valuation(project_id: str, user=Depends(get_current_user)):
    safe_id = project_id.replace("/", "_").replace(".", "_")
    return fb_read(f"valuation/{safe_id}")


@app.get("/valuation/{project_id}/latest")
def get_valuation_latest(project_id: str, user=Depends(get_current_user)):
    safe_id = project_id.replace("/", "_").replace(".", "_")
    return fb_read(f"valuation/{safe_id}/latest")


@app.get("/valuation/{project_id}/versions")
def get_valuation_versions(project_id: str, user=Depends(get_current_user)):
    safe_id = project_id.replace("/", "_").replace(".", "_")
    return fb_read(f"valuation/{safe_id}/versions")


@app.post("/valuation/generate-ic-summary")
async def generate_ic_summary(payload: dict, user=Depends(get_current_user)):
    """Claude API로 IC Summary 전문 보고서 생성"""
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 환경변수 미설정")

    proj    = payload.get("project_name", "Project")
    metrics = payload.get("metrics", {})
    scenarios = payload.get("scenarios", [])
    assumptions = payload.get("assumptions", {})
    history = payload.get("history", [])
    today   = payload.get("date", "")

    scen_text = ""
    if scenarios:
        scen_text = "\n\nScenario Analysis:\n"
        for s in scenarios:
            scen_text += f"  {s.get('name','')}: IRR {s.get('irr','—')}, Dev Margin {s.get('margin','—')}\n"

    hist_text = ""
    if history:
        hist_text = "\n\nVersion History (recent):\n"
        for h in history[:3]:
            hist_text += f"  {h.get('date','')} — {h.get('reason','')}\n"

    prompt = (
        "You are a senior investment analyst at a US renewable energy developer. "
        "Write a concise, professional Investment Committee (IC) Summary in Korean (with key financial metrics in English). "
        "Use formal Korean business writing style. Structure it with clear sections.\n\n"
        f"Project: {proj}\n"
        f"Date: {today}\n\n"
        "Financial Metrics:\n"
        f"  Sponsor IRR: {metrics.get('sirr','—')}\n"
        f"  Dev Margin: {metrics.get('dev_margin','—')}\n"
        f"  Levered IRR: {metrics.get('lirr','—')}\n"
        f"  Unlevered IRR: {metrics.get('uirr','—')}\n"
        f"  EBITDA Yield: {metrics.get('ebitda_yield','—')}\n"
        f"  Total CAPEX: {metrics.get('capex','—')}\n"
        f"  Debt: {metrics.get('debt','—')} ({metrics.get('debt_pct','—')})\n"
        f"  Tax Equity: {metrics.get('te','—')}\n"
        f"  Sponsor Equity: {metrics.get('eq','—')}\n"
        f"  PPA: {metrics.get('ppa','—')}\n"
        f"  BESS Toll: {metrics.get('toll','—')}\n"
        f"  ITC/PTC: {metrics.get('credit','—')}\n"
        f"  Flip Yield: {metrics.get('flip','—')}\n"
        f"{scen_text}{hist_text}\n\n"
        "Write the IC Summary with these sections:\n"
        "1. 프로젝트 개요 (2-3 sentences)\n"
        "2. 핵심 재무 지표 (bullet points with brief commentary)\n"
        "3. Deal Structure 특징 (TE flip structure, debt terms 등)\n"
        "4. 리스크 요인 (2-3 key risks)\n"
        "5. 투자 의견 (1 paragraph recommendation)\n\n"
        "Keep it concise — suitable for a 1-page print. "
        "Return ONLY valid JSON: "
        '{{"sections":[{{"title":"섹션제목","content":"내용"}}]}}'
    )

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 2000,
            "messages": [{"role": "user", "content": prompt}]
        },
        timeout=40
    )
    if resp.status_code != 200:
        raise HTTPException(500, f"Claude API 오류: {resp.text[:200]}")

    data = resp.json()
    text = "".join(b.get("text","") for b in data.get("content",[]))
    clean = text.replace("```json","").replace("```","").strip()
    return {"ok": True, "result": clean}


# ══════════════════════════════════════════════════
#  IC Summary PDF Export (WeasyPrint — world-class formatting)
# ══════════════════════════════════════════════════
import base64 as _base64
from fastapi.responses import Response as _Response

def _esc_html(s):
    """HTML escape helper."""
    if s is None:
        return ""
    return (str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))

def _fmt_pct(v, decimals=2):
    """숫자 → 퍼센트 문자열. 이미 문자열이면 그대로."""
    if v is None or v == "—":
        return "—"
    if isinstance(v, str):
        return v
    try:
        return f"{float(v)*100:.{decimals}f}%"
    except Exception:
        return "—"

def _fmt_usd_m(v):
    """$M 포맷 (input in thousands)."""
    if v is None or v == "—":
        return "—"
    if isinstance(v, str):
        return v
    try:
        return f"${float(v)/1000:.1f}M"
    except Exception:
        return "—"

def _build_ic_pdf_html(data: dict) -> str:
    """IC Summary HTML (WeasyPrint용). World-class IB/PE 수준 포맷."""
    proj_name = data.get("project_name", "Project")
    today = data.get("date", datetime.date.today().isoformat())
    verdict = (data.get("verdict") or "").upper() or "—"
    verdict_color = data.get("verdict_color", "amber")

    # 색상 매핑
    color_map = {
        "green": "#059669",  # emerald-600
        "amber": "#D97706",  # amber-600
        "red":   "#DC2626",  # red-600
    }
    v_color = color_map.get(verdict_color, "#6B7280")

    # 지표
    outputs = data.get("outputs", {}) or {}
    assumptions = data.get("assumptions", {}) or {}
    pv_mwac = assumptions.get("pv_mwac") or outputs.get("pv_mwac") or "—"
    bess_mw = assumptions.get("bess_mw") or "—"
    cod = assumptions.get("cod") or "—"
    ntp = assumptions.get("ntp") or "—"
    state = data.get("state") or assumptions.get("state") or "—"
    iso = data.get("iso") or assumptions.get("iso") or "—"

    # 5 IRR 지표
    irr_lev_pre  = _fmt_pct(outputs.get("sponsor_irr_levered_pretax") or outputs.get("sponsor_irr"))
    irr_at_before = _fmt_pct(outputs.get("sponsor_irr_aftertax_before_nol"))
    irr_at_after  = _fmt_pct(outputs.get("sponsor_irr_aftertax_after_nol"))
    irr_unlev    = _fmt_pct(outputs.get("sponsor_irr_unlevered_pretax") or outputs.get("unlevered_irr"))
    wacc_val     = _fmt_pct(outputs.get("wacc"))

    # 재무 요약
    capex = _fmt_usd_m(outputs.get("capex_total"))
    debt  = _fmt_usd_m(outputs.get("debt"))
    te    = _fmt_usd_m(outputs.get("tax_equity"))
    eq    = _fmt_usd_m(outputs.get("sponsor_equity"))
    dev_margin = _fmt_usd_m(outputs.get("dev_margin"))
    margin_cwp = outputs.get("margin_cwp")
    margin_cwp_str = f"{margin_cwp:.2f} c/Wp" if isinstance(margin_cwp, (int, float)) else "—"
    ppa_price = outputs.get("ppa_price") or assumptions.get("ppa_price") or "—"
    ppa_term  = outputs.get("ppa_term") or assumptions.get("ppa_term") or "—"
    bess_toll = outputs.get("bess_toll") or assumptions.get("bess_toll") or "—"
    ebitda_y  = outputs.get("ebitda_yield")
    ebitda_y_str = f"{ebitda_y:.2f}%" if isinstance(ebitda_y, (int, float)) else "—"

    # AI 분석 결과 (IC Opinion 에서 생성된 것)
    ic_analysis = data.get("ic_analysis", {}) or {}
    thesis = ic_analysis.get("thesis", "")
    rec    = ic_analysis.get("rec", "")
    risks  = ic_analysis.get("risks", []) or []
    threshold_status = ic_analysis.get("threshold_status", {}) or {}
    dev_ic = ic_analysis.get("dev_ic", {}) or {}

    # Sensitivity (프론트에서 계산된 값)
    scenarios = data.get("scenarios", []) or []

    # Threshold 메타
    thresholds = data.get("thresholds", {}) or {}
    thr_irr = thresholds.get("sponsor_irr_pct", 9.0)
    thr_margin = thresholds.get("dev_margin_cwp", 10.0)

    # 리스크 분리: compliance_count만큼 앞쪽은 고정 체크리스트, 뒤는 AI 생성
    compliance_count = int(ic_analysis.get("compliance_count", 0) or 0)
    compliance_items = risks[:compliance_count] if compliance_count else []
    ai_risks = risks[compliance_count:] if compliance_count else risks

    # 1) 컴플라이언스 체크리스트 HTML
    compliance_html = ""
    for r in compliance_items:
        title = _esc_html(r.get("title", ""))
        detail = _esc_html(r.get("detail", ""))
        sev = r.get("severity", "Watch")
        compliance_html += f"""
        <div class="compliance-item">
          <div class="compliance-box"></div>
          <div class="compliance-body">
            <div class="compliance-head">
              <span class="compliance-title">{title}</span>
              <span class="compliance-sev">{sev}</span>
            </div>
            <div class="compliance-detail">{detail}</div>
          </div>
        </div>
        """

    # 2) AI 프로젝트별 리스크 HTML
    risks_html = ""
    sev_color = {"Critical": "#DC2626", "Watch": "#D97706", "OK": "#059669"}
    for i, r in enumerate(ai_risks[:8]):
        sev = r.get("severity", "OK")
        c = sev_color.get(sev, "#6B7280")
        title = _esc_html(r.get("title", ""))
        detail = _esc_html(r.get("detail", ""))
        risks_html += f"""
        <div class="risk-item">
          <div class="risk-header">
            <span class="risk-num">{i+1:02d}</span>
            <span class="risk-title">{title}</span>
            <span class="risk-sev" style="background:{c}">{sev}</span>
          </div>
          <div class="risk-detail">{detail}</div>
        </div>
        """

    # Scenario 테이블
    scen_rows = ""
    for s in scenarios:
        scen_rows += f"""
        <tr>
          <td class="scen-name">{_esc_html(s.get('name','—'))}</td>
          <td class="scen-val">{_esc_html(s.get('irr','—'))}</td>
          <td class="scen-val">{_esc_html(s.get('margin','—'))}</td>
        </tr>
        """

    # Threshold 체크
    def _chk(ok):
        return ('<span style="color:#059669;font-weight:700">✓ PASS</span>' if ok
                else '<span style="color:#DC2626;font-weight:700">✗ FAIL</span>')
    thr_irr_ok = threshold_status.get("irr_ok", False)
    thr_margin_ok = threshold_status.get("margin_ok", False)
    thr_irr_gap = _esc_html(threshold_status.get("irr_gap", ""))
    thr_margin_gap = _esc_html(threshold_status.get("margin_gap", ""))

    # HTML 조립
    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>IC Summary - {_esc_html(proj_name)}</title>
<style>
@page {{
  size: A4 portrait;
  margin: 18mm 16mm 18mm 16mm;
  @bottom-center {{
    content: counter(page) " / " counter(pages);
    font-family: 'Noto Sans KR', 'Helvetica', sans-serif;
    font-size: 8pt;
    color: #6B7280;
  }}
  @bottom-left {{
    content: "Hanwha Energy USA Holdings · Internal IC Memo";
    font-family: 'Noto Sans KR', 'Helvetica', sans-serif;
    font-size: 7pt;
    color: #9CA3AF;
  }}
  @bottom-right {{
    content: "{_esc_html(today)}";
    font-family: 'Noto Sans KR', 'Helvetica', sans-serif;
    font-size: 7pt;
    color: #9CA3AF;
  }}
}}
@page :first {{
  @bottom-center {{ content: none; }}
  @bottom-left {{ content: none; }}
  @bottom-right {{ content: none; }}
}}
* {{ box-sizing: border-box; }}
body {{
  font-family: 'Noto Sans KR', 'Helvetica Neue', Helvetica, sans-serif;
  font-size: 10pt;
  line-height: 1.55;
  color: #111827;
  margin: 0;
  padding: 0;
  -webkit-font-smoothing: antialiased;
}}

/* ── Cover ────────────────────────────────────── */
.cover {{
  height: 260mm;
  display: flex;
  flex-direction: column;
  justify-content: space-between;
  padding: 20mm 8mm 8mm 8mm;
}}
.cover-header {{
  font-size: 8pt;
  font-weight: 600;
  color: #6B7280;
  letter-spacing: 0.2em;
  text-transform: uppercase;
  border-bottom: 1px solid #E5E7EB;
  padding-bottom: 12pt;
}}
.cover-main {{
  margin-top: 40mm;
}}
.cover-tag {{
  font-size: 9pt;
  font-weight: 600;
  color: {v_color};
  letter-spacing: 0.16em;
  text-transform: uppercase;
  margin-bottom: 10pt;
}}
.cover-title {{
  font-size: 36pt;
  font-weight: 800;
  color: #111827;
  letter-spacing: -1.2pt;
  line-height: 1.05;
  margin-bottom: 14pt;
}}
.cover-sub {{
  font-size: 12pt;
  color: #4B5563;
  font-weight: 400;
  margin-bottom: 40pt;
}}
.cover-verdict {{
  display: inline-block;
  padding: 10pt 22pt;
  border: 2pt solid {v_color};
  border-radius: 2pt;
  font-size: 22pt;
  font-weight: 800;
  color: {v_color};
  letter-spacing: 4pt;
}}
.cover-stats {{
  display: flex;
  gap: 20pt;
  margin-top: 24pt;
}}
.cover-stat {{
  flex: 1;
  border-left: 2pt solid #E5E7EB;
  padding-left: 10pt;
}}
.cover-stat-label {{
  font-size: 7pt;
  font-weight: 700;
  color: #6B7280;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  margin-bottom: 3pt;
}}
.cover-stat-value {{
  font-size: 16pt;
  font-weight: 700;
  color: #111827;
  font-variant-numeric: tabular-nums;
}}
.cover-footer {{
  margin-top: auto;
  padding-top: 20pt;
  border-top: 1px solid #E5E7EB;
  display: flex;
  justify-content: space-between;
  font-size: 8pt;
  color: #6B7280;
}}

/* ── Content Pages ────────────────────────────── */
.page-break {{ page-break-before: always; }}

h1 {{
  font-size: 14pt;
  font-weight: 800;
  color: #111827;
  margin: 0 0 3pt 0;
  letter-spacing: -0.3pt;
}}
.section-sub {{
  font-size: 8pt;
  color: #6B7280;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  margin-bottom: 12pt;
  border-bottom: 1pt solid #E5E7EB;
  padding-bottom: 6pt;
}}
h2 {{
  font-size: 10pt;
  font-weight: 700;
  color: #111827;
  margin: 14pt 0 6pt 0;
  letter-spacing: 0;
}}
p {{ margin: 4pt 0; color: #1F2937; }}

/* Metrics Grid */
.metrics-grid {{
  display: grid;
  grid-template-columns: repeat(5, 1fr);
  gap: 4pt;
  margin-bottom: 14pt;
}}
.metric-card {{
  padding: 7pt 8pt;
  border: 0.5pt solid #D1D5DB;
  border-radius: 2pt;
  overflow: hidden;
}}
.metric-card-primary {{
  border-left: 2.5pt solid #059669;
}}
.metric-card-secondary {{
  border-left: 2.5pt solid #D97706;
}}
.metric-card-wacc {{
  border-left: 2.5pt solid #2563EB;
}}
.metric-label {{
  font-size: 6.5pt;
  font-weight: 700;
  color: #6B7280;
  letter-spacing: 0.04em;
  text-transform: uppercase;
  margin-bottom: 2pt;
  white-space: nowrap;
}}
.metric-value {{
  font-size: 14pt;
  font-weight: 700;
  color: #111827;
  font-variant-numeric: tabular-nums;
  line-height: 1.1;
  white-space: nowrap;
}}
.metric-sub {{
  font-size: 6.5pt;
  color: #6B7280;
  margin-top: 2pt;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}}

/* Financial table */
.fin-table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 9pt;
  margin: 8pt 0 14pt 0;
}}
.fin-table th {{
  text-align: left;
  padding: 6pt 8pt;
  border-bottom: 1pt solid #111827;
  font-size: 7pt;
  font-weight: 700;
  letter-spacing: 0.08em;
  text-transform: uppercase;
  color: #374151;
}}
.fin-table td {{
  padding: 5pt 8pt;
  border-bottom: 0.3pt solid #E5E7EB;
  font-variant-numeric: tabular-nums;
}}
.fin-table td.val {{ text-align: right; font-weight: 600; }}
.fin-table tr.subtotal td {{
  background: #F9FAFB;
  font-weight: 700;
  border-top: 0.5pt solid #6B7280;
}}

/* Threshold check */
.thr-box {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 10pt;
  margin: 10pt 0;
}}
.thr-item {{
  padding: 10pt 12pt;
  border: 1pt solid #E5E7EB;
  border-radius: 2pt;
  background: #F9FAFB;
}}
.thr-label {{
  font-size: 8pt;
  font-weight: 600;
  color: #6B7280;
  margin-bottom: 4pt;
}}
.thr-status {{ font-size: 11pt; margin-bottom: 3pt; }}
.thr-gap {{ font-size: 9pt; color: #374151; }}

/* Thesis / Recommendation boxes */
.thesis-box {{
  padding: 12pt 14pt;
  background: #F9FAFB;
  border-left: 3pt solid #2563EB;
  border-radius: 0 2pt 2pt 0;
  margin: 10pt 0;
  font-size: 10pt;
  line-height: 1.7;
  color: #1F2937;
}}
.rec-box {{
  padding: 14pt 16pt;
  background: #FEF3C7;
  border-left: 3pt solid #D97706;
  border-radius: 0 2pt 2pt 0;
  margin: 10pt 0;
  font-size: 10pt;
  line-height: 1.7;
  color: #78350F;
  font-weight: 500;
}}

/* Risk items */
.risk-item {{
  padding: 10pt 0;
  border-bottom: 0.5pt solid #E5E7EB;
}}
.risk-item:last-child {{ border-bottom: none; }}
.risk-header {{
  display: flex;
  align-items: center;
  gap: 8pt;
  margin-bottom: 4pt;
}}
.risk-num {{
  font-size: 8pt;
  font-weight: 700;
  color: #9CA3AF;
  font-variant-numeric: tabular-nums;
  min-width: 18pt;
}}
.risk-title {{
  font-size: 10pt;
  font-weight: 700;
  color: #111827;
  flex: 1;
}}
.risk-sev {{
  color: #fff;
  font-size: 7pt;
  font-weight: 700;
  padding: 2pt 7pt;
  border-radius: 2pt;
  letter-spacing: 0.05em;
}}
.risk-detail {{
  font-size: 9pt;
  color: #4B5563;
  line-height: 1.6;
  margin-left: 26pt;
}}

/* Compliance Checklist items */
.compliance-note {{
  font-size: 8pt;
  color: #6B7280;
  font-style: italic;
  margin-bottom: 8pt;
}}
.compliance-item {{
  display: flex;
  gap: 10pt;
  padding: 9pt 12pt;
  background: #FFFBEB;
  border: 0.5pt solid #FCD34D;
  border-left: 3pt solid #D97706;
  border-radius: 2pt;
  margin-bottom: 6pt;
}}
.compliance-box {{
  flex-shrink: 0;
  width: 12pt;
  height: 12pt;
  border: 1pt solid #9CA3AF;
  border-radius: 2pt;
  margin-top: 2pt;
}}
.compliance-body {{ flex: 1; }}
.compliance-head {{
  display: flex;
  align-items: center;
  gap: 8pt;
  margin-bottom: 3pt;
}}
.compliance-title {{
  font-size: 10pt;
  font-weight: 700;
  color: #111827;
  flex: 1;
}}
.compliance-sev {{
  font-size: 7pt;
  font-weight: 700;
  color: #D97706;
  border: 0.5pt solid #D97706;
  padding: 1pt 6pt;
  border-radius: 8pt;
  letter-spacing: 0.05em;
}}
.compliance-detail {{
  font-size: 9pt;
  color: #78350F;
  line-height: 1.6;
}}

/* Scenario table */
.scen-table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 9pt;
  margin: 8pt 0;
}}
.scen-table th {{
  text-align: left;
  padding: 7pt 10pt;
  background: #111827;
  color: #fff;
  font-size: 7.5pt;
  font-weight: 700;
  letter-spacing: 0.1em;
  text-transform: uppercase;
}}
.scen-table td {{
  padding: 7pt 10pt;
  border-bottom: 0.5pt solid #E5E7EB;
}}
.scen-name {{ font-weight: 700; color: #111827; }}
.scen-val {{ font-variant-numeric: tabular-nums; text-align: right; }}

/* Dev IC Grid */
.devic-grid {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 10pt;
  margin: 10pt 0;
}}
.devic-item {{
  padding: 9pt 12pt;
  border: 0.5pt solid #E5E7EB;
  border-radius: 2pt;
  background: #FEFEFE;
}}
.devic-label {{
  font-size: 7pt;
  font-weight: 700;
  color: #6B7280;
  letter-spacing: 0.08em;
  text-transform: uppercase;
  margin-bottom: 3pt;
}}
.devic-value {{
  font-size: 9pt;
  color: #1F2937;
  line-height: 1.5;
}}

/* Footer note */
.confidential-note {{
  margin-top: 20pt;
  padding-top: 10pt;
  border-top: 0.3pt solid #E5E7EB;
  font-size: 7pt;
  color: #9CA3AF;
  font-style: italic;
  text-align: center;
}}
</style>
</head>
<body>

<!-- ══ COVER PAGE ══ -->
<div class="cover">
  <div>
    <div class="cover-header">Hanwha Energy USA Holdings · Investment Committee Memo</div>
  </div>

  <div class="cover-main">
    <div class="cover-tag">Confidential · Internal Use Only</div>
    <div class="cover-title">{_esc_html(proj_name)}</div>
    <div class="cover-sub">{_esc_html(pv_mwac)} MWac Solar + BESS · {_esc_html(state)} ({_esc_html(iso)}) · COD {_esc_html(cod)}</div>

    <div class="cover-verdict">{_esc_html(verdict)}</div>

    <div class="cover-stats">
      <div class="cover-stat">
        <div class="cover-stat-label">Sponsor IRR</div>
        <div class="cover-stat-value">{irr_lev_pre}</div>
      </div>
      <div class="cover-stat">
        <div class="cover-stat-label">Dev Margin</div>
        <div class="cover-stat-value">{dev_margin}</div>
      </div>
      <div class="cover-stat">
        <div class="cover-stat-label">Total CAPEX</div>
        <div class="cover-stat-value">{capex}</div>
      </div>
      <div class="cover-stat">
        <div class="cover-stat-label">WACC</div>
        <div class="cover-stat-value">{wacc_val}</div>
      </div>
    </div>
  </div>

  <div class="cover-footer">
    <span>Prepared: {_esc_html(today)}</span>
    <span>{_esc_html(data.get("prepared_by",""))}</span>
  </div>
</div>

<!-- ══ PAGE 2 — EXECUTIVE SUMMARY ══ -->
<div class="page-break">
  <h1>Executive Summary</h1>
  <div class="section-sub">투자 의견 · 핵심 논거</div>

  <h2>투자 근거 (Investment Rationale)</h2>
  <div class="thesis-box">{_esc_html(thesis) if thesis else "(AI 분석 미완료 — IC Opinion 탭에서 Run AI Analysis 실행 후 재생성)"}</div>

  <h2>Recommendation</h2>
  <div class="rec-box">{_esc_html(rec) if rec else "(AI 분석 미완료)"}</div>
</div>

<!-- ══ PAGE 3 — FINANCIAL SUMMARY ══ -->
<div class="page-break">
  <h1>Financial Summary</h1>
  <div class="section-sub">재무 지표 · 자본 구조 · 계약 조건</div>

  <h2>Returns Detail</h2>
  <div class="metrics-grid">
    <div class="metric-card metric-card-primary">
      <div class="metric-label">Sponsor IRR</div>
      <div class="metric-value">{irr_lev_pre}</div>
      <div class="metric-sub">Lev · Pre-Tax</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Sponsor IRR</div>
      <div class="metric-value">{irr_at_before}</div>
      <div class="metric-sub">A-Tax · Pre-NOL</div>
    </div>
    <div class="metric-card metric-card-secondary">
      <div class="metric-label">Sponsor IRR</div>
      <div class="metric-value">{irr_at_after}</div>
      <div class="metric-sub">A-Tax · Post-NOL</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Project IRR</div>
      <div class="metric-value">{irr_unlev}</div>
      <div class="metric-sub">Unlev · Pre-Tax</div>
    </div>
    <div class="metric-card metric-card-wacc">
      <div class="metric-label">WACC</div>
      <div class="metric-value">{wacc_val}</div>
      <div class="metric-sub">Capital Cost</div>
    </div>
  </div>

  <h2>Investment Thresholds (기준 달성)</h2>
  <div class="thr-box">
    <div class="thr-item">
      <div class="thr-label">Sponsor IRR (After-TE-Flip, Full Life · min {thr_irr}%)</div>
      <div class="thr-status">{_chk(thr_irr_ok)}</div>
      <div class="thr-gap">{thr_irr_gap}</div>
    </div>
    <div class="thr-item">
      <div class="thr-label">Dev Margin (min {thr_margin} c/Wp)</div>
      <div class="thr-status">{_chk(thr_margin_ok)}</div>
      <div class="thr-gap">{thr_margin_gap}</div>
    </div>
  </div>

  <h2>Capital Structure & Deal Terms</h2>
  <table class="fin-table">
    <thead><tr><th>Item</th><th style="text-align:right">Value</th></tr></thead>
    <tbody>
      <tr><td>Total CAPEX</td><td class="val">{capex}</td></tr>
      <tr><td>Senior Debt</td><td class="val">{debt}</td></tr>
      <tr><td>Tax Equity</td><td class="val">{te}</td></tr>
      <tr><td>Sponsor Equity</td><td class="val">{eq}</td></tr>
      <tr class="subtotal"><td>Dev Margin</td><td class="val">{dev_margin} ({margin_cwp_str})</td></tr>
      <tr><td>EBITDA Yield (Y1)</td><td class="val">{ebitda_y_str}</td></tr>
      <tr><td>PPA Price × Term</td><td class="val">${_esc_html(ppa_price)}/MWh × {_esc_html(ppa_term)}yr</td></tr>
      <tr><td>BESS Toll</td><td class="val">${_esc_html(bess_toll)}/kW-mo</td></tr>
    </tbody>
  </table>

  <h2>Scenario Analysis</h2>
  <table class="scen-table">
    <thead><tr><th>Scenario</th><th style="text-align:right">Sponsor IRR</th><th style="text-align:right">Dev Margin</th></tr></thead>
    <tbody>{scen_rows if scen_rows else '<tr><td colspan="3" style="color:#9CA3AF;text-align:center">시나리오 미실행</td></tr>'}</tbody>
  </table>
</div>

<!-- ══ PAGE 4 — RISK ASSESSMENT ══ -->
<div class="page-break">
  <h1>Risk Assessment</h1>
  <div class="section-sub">규정 준수 체크 · 프로젝트별 리스크 (의사결정에 반영되지 않음)</div>

  <h2 style="margin-top:10pt">📋 규정 준수 체크리스트 (IC 승인 전 확인 필수)</h2>
  <div class="compliance-note">고정 체크리스트 · 모든 프로젝트 공통 적용</div>
  {compliance_html if compliance_html else '<p style="color:#9CA3AF;font-size:9pt">체크리스트 없음</p>'}

  <h2 style="margin-top:18pt">🔍 프로젝트별 리스크 (AI 모니터링)</h2>
  <div class="compliance-note">정보 제공 · 경제성 판정에 영향 없음</div>
  {risks_html if risks_html else '<p style="color:#9CA3AF;font-size:9pt">AI 분석 미완료 — IC Opinion 탭에서 Run AI Analysis 실행 후 재생성</p>'}

  <div class="confidential-note">
    본 문서는 Hanwha Energy USA Holdings 내부 투자심의 목적으로만 작성되었으며, 외부 유출을 금합니다.<br>
    수치 및 가정은 {_esc_html(today)} 기준 엑셀 재무모델 및 시장 데이터를 근거로 하며, 시장 변동에 따라 달라질 수 있습니다.<br>
    경제성 판정(PROCEED/RECUT/STOP)은 Dev Margin · Sponsor IRR · Unlev IRR vs WACC 기준의 순수 경제 분석 결과이며, 규정 준수 체크리스트와 개별 리스크는 별도 관리 대상입니다.
  </div>
</div>

</body>
</html>"""
    return html


@app.post("/valuation/export-pdf")
async def export_ic_pdf(payload: dict, user=Depends(get_current_user)):
    """IC Summary PDF 생성 (WeasyPrint, world-class formatting)."""
    import traceback
    import sys

    # Step 1: WeasyPrint import
    try:
        from weasyprint import HTML
        print(f"[export-pdf] WeasyPrint import OK", flush=True)
    except Exception as e:
        print(f"[export-pdf] WeasyPrint import FAILED: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        raise HTTPException(500, f"WeasyPrint import 실패: {str(e)[:300]}")

    # Step 2: HTML 문자열 생성
    try:
        html_str = _build_ic_pdf_html(payload)
        print(f"[export-pdf] HTML built, length={len(html_str)}", flush=True)
    except Exception as e:
        print(f"[export-pdf] HTML build FAILED: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        raise HTTPException(500, f"HTML 생성 오류: {str(e)[:300]}")

    # Step 3: PDF 렌더링
    try:
        pdf_bytes = HTML(string=html_str).write_pdf()
        print(f"[export-pdf] PDF rendered, size={len(pdf_bytes)} bytes", flush=True)
    except Exception as e:
        print(f"[export-pdf] PDF render FAILED: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        raise HTTPException(500, f"PDF 렌더링 오류: {str(e)[:300]}")

    proj_name = payload.get("project_name", "IC_Summary").replace(" ", "_")
    date_str = payload.get("date", datetime.date.today().isoformat()).replace("-", "")
    filename = f"IC_Summary_{proj_name}_{date_str}.pdf"

    return _Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ── WeasyPrint 진단용 미니멀 테스트 엔드포인트 ───────
@app.get("/valuation/export-pdf-test")
async def export_pdf_test(user=Depends(get_current_user)):
    """WeasyPrint가 살아있는지 간단히 테스트."""
    import traceback, sys
    try:
        from weasyprint import HTML
        simple_html = "<html><body><h1>Test</h1><p>안녕하세요, WeasyPrint 테스트</p></body></html>"
        pdf = HTML(string=simple_html).write_pdf()
        return _Response(
            content=pdf,
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="weasyprint_test.pdf"'}
        )
    except Exception as e:
        traceback.print_exc(file=sys.stdout)
        raise HTTPException(500, f"테스트 실패: {str(e)[:500]}")


@app.post("/valuation/analyze-cf")
async def analyze_cf(payload: dict, user=Depends(get_current_user)):
    """CF 데이터를 Claude API로 분석하여 인사이트 반환"""
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 환경변수 미설정")

    cf_text   = payload.get("cf_text", "")
    proj_name = payload.get("project_name", "프로젝트")

    context        = payload.get("context", "")
    proj_context   = payload.get("project_context", "")  # PPV 탭 프로젝트 메타데이터
    lang           = payload.get("lang", "en")
    mode           = payload.get("mode", "full")
    proj_meta      = payload.get("project_meta", {})
    stage    = proj_meta.get("stage", "")
    iso      = proj_meta.get("iso", "")
    proj_type= proj_meta.get("type", "")
    ntp_date = proj_meta.get("ntp", "")
    cod_date = proj_meta.get("cod", "")
    risk_pct = proj_meta.get("risk_factor", "")
    if risk_pct != "": risk_pct = f"{float(risk_pct)*100:.0f}%"
    itc_risk = proj_meta.get("itc_expiry_risk", "")
    proj_ctx = proj_meta.get("proj_ctx", "")
    thresholds     = payload.get("thresholds", {})
    current_metrics= payload.get("current_metrics", {})

    irr_thr    = thresholds.get("sponsor_irr_pct", 9.0)
    margin_thr = thresholds.get("dev_margin_cwp", 10.0)
    itc_thr    = thresholds.get("itc_min_pct", 30.0)

    curr_irr    = current_metrics.get("sponsor_irr_pct", "?")
    curr_irr_basis = current_metrics.get("sponsor_irr_basis", "After-TE-Flip, Full Life")
    curr_margin = current_metrics.get("dev_margin_cwp", "?")
    curr_itc    = current_metrics.get("itc_rate_pct", "?")
    sponsor_npv_m = current_metrics.get("sponsor_npv_m")  # $M (optional)
    project_npv_m = current_metrics.get("project_npv_m")  # $M (optional)
    ppa_term    = current_metrics.get("ppa_term", "?")
    toll_term   = current_metrics.get("toll_term", "?")
    pv_mwac     = current_metrics.get("pv_mwac", "?")

    if mode == "interp":
        prompt = (
            "미국 태양광+BESS PF 전문가로서 아래 연도별 Sponsor CF 패턴을 분석해줘.\n"
            f"프로젝트: {proj_name}\n"
            f"CF: {cf_text}\n\n"
            "3~4개 핵심 인사이트를 JSON으로 반환 (다른 텍스트 없이):\n"
            '{"insights":[{"title":"제목","detail":"설명(80자이내)"}]}'
        )
    else:
        # 동적 날짜 계산
        _today = datetime.date.today()
        _current_year = _today.year
        _current_quarter = f"{_current_year}-Q{(_today.month - 1) // 3 + 1}"
        _prev_q_month = _today.month - 3
        _prev_q_year = _current_year
        if _prev_q_month <= 0:
            _prev_q_month += 12
            _prev_q_year -= 1
        _prev_quarter = f"{_prev_q_year}-Q{(_prev_q_month - 1) // 3 + 1}"

        # 시장 데이터 컨텍스트 (payload에서 주입)
        market_context = payload.get("market_context", {}) or {}
        rates_txt = market_context.get("rates_summary", "")
        levelten_txt = market_context.get("levelten_summary", "")
        # BESS 소스 priority: LevelTen Storage (official) 1순위, AI Research (fallback) 2순위
        levelten_storage_txt = market_context.get("levelten_storage_summary", "")  # 공식 ISO-level
        bess_tolling_txt = market_context.get("bess_tolling_summary", "")          # AI Research duration별
        our_bess_duration = market_context.get("our_bess_duration", 4)
        # LevelTen 커버 여부 + 지역 해석 (WECC sub-region, SERC 등)
        lt_covered = market_context.get("levelten_covered", True)  # 기본 True (기존 프로젝트 호환)
        region_display = market_context.get("region_display", "")  # "WECC Rocky Mountain (UT)" 등
        sub_region = market_context.get("sub_region", "")          # WECC_RM, WECC_DSW, etc.
        continental_avg_txt = market_context.get("continental_avg_summary", "")  # Market-Averaged Continental (대용 비교용)

        market_block = ""
        if rates_txt or levelten_txt or levelten_storage_txt or bess_tolling_txt or continental_avg_txt:
            market_block = "=== CURRENT MARKET DATA (most recent; use this INSTEAD of training knowledge) ===\n"
            if rates_txt:
                market_block += f"  Interest Rates: {rates_txt}\n"
            # 지역 해석 명시
            if region_display:
                market_block += f"  Project Region: {region_display}"
                if lt_covered:
                    market_block += " [LevelTen 직접 커버 ISO]\n"
                else:
                    market_block += f" [LevelTen 미커버 — 대용 비교 필요]\n"
            if levelten_txt:
                market_block += f"  LevelTen PPA Benchmark (Solar): {levelten_txt}\n"
                if lt_covered:
                    market_block += "  → Our ISO가 LevelTen에 있음. USE IT to compare against project PPA directly.\n"
                    market_block += "    Reference: 'PPA $X.XX vs LevelTen P25 $Y.YY in {ISO}'.\n"
                else:
                    market_block += "  → Our region is NOT in LevelTen. Use as market context reference only.\n"
            # 대용 비교 (WECC/SERC 등 LevelTen 미커버 지역)
            if not lt_covered and continental_avg_txt:
                market_block += f"  Market-Averaged Continental Index (대용 비교용): {continental_avg_txt}\n"
                market_block += f"  → Use this as PRIMARY benchmark since {sub_region or 'project region'} has NO direct LevelTen coverage.\n"
                market_block += "  → Cite as 'LevelTen Market-Averaged Continental (전 대륙 ISO 평균, 대용치)'.\n"
                market_block += "  → Explicitly note '해당 지역 공식 P25 데이터 없음 → 대륙 평균 대비 비교' in risk commentary.\n"
            # Priority 1: LevelTen Storage (official, ISO-level)
            if levelten_storage_txt:
                market_block += f"  LevelTen Storage Index (OFFICIAL tolling offers, Q4 2025): {levelten_storage_txt}\n"
                market_block += f"  → Project BESS duration: {our_bess_duration}h (ISO-level price applies broadly, consider duration fit)\n"
                market_block += "  → USE THIS OFFICIAL DATA as primary BESS benchmark. Cite as 'LevelTen 공식 Storage Index'.\n"
                market_block += "  → If project toll EXCEEDS ISO P75 → risk 'BESS Toll 시장 상단 초과' (severity: Critical if >20% over, Watch if slight).\n"
                market_block += "  → If project toll is BELOW ISO P25 → positive flag '보수적 산정'.\n"
            # Priority 2: AI Research (fallback for non-LevelTen ISOs: ISO-NE/NYISO/WECC_*/SERC, or duration-level detail)
            if bess_tolling_txt:
                if levelten_storage_txt:
                    market_block += f"  AI Research Duration Detail (supplementary — LevelTen only provides ISO-level): {bess_tolling_txt}\n"
                    market_block += f"  → Use ONLY to add duration-specific nuance ({our_bess_duration}h). LevelTen ISO-level is primary.\n"
                    market_block += "  → Caveat: 'duration 세부는 AI 추정치'.\n"
                else:
                    # LevelTen 없는 지역 (WECC_*, ISO-NE, NYISO, SERC)
                    market_block += f"  BESS Tolling Estimate (AI Research — {sub_region or 'non-LevelTen region'}): {bess_tolling_txt}\n"
                    market_block += f"  → Project BESS duration: {our_bess_duration}h\n"
                    market_block += "  → CAVEAT: '시장 추정치, 공식 index 아님' when citing.\n"
                    if sub_region and sub_region.startswith("WECC"):
                        market_block += f"  → For {sub_region}: reference relevant utility RFPs (PacifiCorp IRP, URC, APS, Xcel Colorado, etc.) if AI Research provided commentary.\n"
            market_block += "\n"

        # 경제성 지표 추출 (Unlevered vs WACC 비교용)
        unlev_irr = current_metrics.get("unlevered_irr_pct")
        wacc_val  = current_metrics.get("wacc_pct")
        wacc_block = ""
        if unlev_irr is not None and wacc_val is not None:
            wacc_block = (
                f"  Unlevered Pre-Tax IRR : {unlev_irr}% (project-level)\n"
                f"  WACC                  : {wacc_val}% (hurdle)\n"
                f"  Value Creation        : Unlev - WACC = "
                f"{'POSITIVE' if float(unlev_irr)>float(wacc_val) else 'NEGATIVE'}\n"
            )

        prompt = (
        "You are the head of Investment Committee at Hanwha Energy USA (HEUH), "
        "a renewable energy developer whose sole business model is: develop → sell at NTP (before COD). "
        "The IC decision: should we continue spending development capital on this project?\n\n"

        f"TODAY'S DATE: {_today.isoformat()} (current quarter: {_current_quarter}, prior: {_prev_quarter}).\n\n"

        "═══ KNOWN REGULATORY & OPERATIONAL FACTS (treat as GIVEN; do not second-guess) ═══\n"
        "1. ITC Section 48E — Solar PV:\n"
        "   - 'Beginning of Construction' (BOC) is a LEGAL construct, not physical construction start.\n"
        "   - BOC deadline: July 4, 2026 — established via Physical Work Test (on-site or off-site binding work).\n"
        "   - Continuity Safe Harbor preserved: if BOC is established, project has until Dec 31, 2030 (4 years) to reach PIS.\n"
        "   - Projects missing BOC by July 4, 2026 must be Placed-in-Service by Dec 31, 2027.\n"
        "2. ITC Section 48E — BESS (SEPARATE TRACK from PV):\n"
        "   - Begin Construction by Dec 31, 2033 → 100% ITC\n"
        "   - 2034 → 75%, 2035 → 50%, 2036 → expires\n"
        "   - BESS is NOT subject to the 2026 solar cliff. Do NOT flag BESS ITC as imminent risk.\n"
        "3. HEUH Business Model & BOC Status:\n"
        "   - HEUH develops → sells at NTP (pre-COD). Post-COD execution risk does NOT affect IC decision.\n"
        "   - HEUH has established BOC for its project pool via Physical Work Test, managed by its compliance team.\n"
        "   - Individual project matching to BOC pool is operational matter — do NOT flag as financial risk.\n"
        "   - Post-BOC physical construction schedule is flexible within 4-year Continuity Safe Harbor.\n"
        "4. FEOC (Foreign Entity of Concern): compliance checklist item — do NOT use as verdict driver.\n\n"

        f"PROJECT: {proj_name} | Size: {pv_mwac} MWac\n"
        f"FINANCIAL SUMMARY: {context}\n"
        f"PROJECT METADATA: {proj_ctx}\n"
        f"ANNUAL SPONSOR CF (Y1-Y10): {cf_text}\n\n"

        + market_block +

        "=== INVESTMENT THRESHOLDS (firm hurdles) ===\n"
        f"  Primary   · Sponsor IRR ≥ {irr_thr}% (After-TE-Flip, Full Life) — 매수자 요구 수익률\n"
        f"  Secondary · Dev Margin  ≥ {margin_thr} c/Wp — HEUH 내부 마진 기준\n"
        "  Both must PASS for IC approval.\n\n"

        "=== CURRENT PROJECT METRICS ===\n"
        f"  Sponsor IRR : {curr_irr}% ({curr_irr_basis})\n"
        f"  Dev Margin  : {curr_margin} c/Wp\n"
        + (f"  Sponsor NPV : ${sponsor_npv_m}M (discounted at {irr_thr}% hurdle)\n" if sponsor_npv_m is not None else "")
        + (f"  Project NPV : ${project_npv_m}M (discounted at WACC)\n" if project_npv_m is not None else "")
        + wacc_block +
        f"  ITC Rate    : {curr_itc}%\n"
        f"  PPA Term    : {ppa_term} yrs | Toll Term: {toll_term} yrs\n\n"

        "═══ VERDICT FRAMEWORK (PURE ECONOMICS ONLY) ═══\n"
        "The verdict is determined ONLY by economic criteria. Development risks are monitoring items and do NOT affect verdict.\n\n"
        "Economic criteria:\n"
        "  1. Dev Margin vs threshold (primary: HEUH's exit value)\n"
        "  2. Sponsor IRR (After-TE-Flip, Full Life) vs threshold (market-clearing for buyer)\n"
        "  3. Unlevered IRR vs WACC (true value creation — leverage-independent)\n\n"
        "VERDICT RULES:\n"
        "  PROCEED:\n"
        "    - Dev Margin ≥ threshold AND Sponsor IRR ≥ threshold AND Unlev IRR > WACC\n"
        "    - Express threshold headroom explicitly if positive (e.g., '+1.5%p 여유')\n"
        "  RECUT:\n"
        "    - 1~2 criteria near miss (gap < 1.5%p from threshold) AND recoverable via negotiation\n"
        "    - Typical levers: PPA price revision, CAPEX reduction, TE/debt terms\n"
        "  STOP:\n"
        "    - Multiple criteria missed OR Unlev IRR < WACC (value destruction)\n"
        "    - Unrecoverable: gap too wide to close via normal levers\n\n"

        "═══ RISK ANALYSIS (monitoring only — NOT verdict driver) ═══\n"
        "Identify project-specific risks AI can assess:\n"
        "  - EPC price adequacy: $/Wdc vs current market (use supplied MARKET DATA if provided)\n"
        "  - ISO / interconnection queue risk based on ISO and state\n"
        "  - PPA market competitiveness vs supplied LevelTen P25 data (if given)\n"
        "  - Construction timeline vs PIS deadlines (Solar PV: 4-year continuity → Dec 31, 2030 PIS if BOC established; BESS: flexible to 2033)\n"
        "  - BESS replacement CAPEX / augmentation assumption sanity\n"
        "DO NOT generate risks for:\n"
        "  - Safe Harbor matching or BOC status (handled separately as fixed checklist item)\n"
        "  - FEOC compliance (handled separately as fixed checklist item)\n"
        "  - BESS ITC expiry (not imminent — 2033+ horizon)\n"
        "  - 'Must begin physical construction by 2026' — INCORRECT; BOC is a legal construct already managed via HEUH's Physical Work Test completion\n"
        "  - Generic 'market uncertainty' or 'policy risk' without specifics\n\n"

        "═══ LANGUAGE ═══\n"
        + ("ALL text fields in KOREAN (한국어).\n"
           "\n"
           "CRITICAL — Korean ENDING STYLE (IC memo convention, formal & concise):\n"
           "Use short nominal/verbal endings, NOT 존대체 (하다/한다) nor 합쇼체 (합니다).\n"
           "Required endings:\n"
           "  - 명사형 종결: '~ 충족', '~ 확인', '~ 권고', '~ 필요', '~ 가능', '~ 부족'\n"
           "  - 축약 서술: '~됨', '~함', '~임', '~없음', '~확보됨'\n"
           "Examples (GOOD):\n"
           "  ✓ '개발 마진 20.0 c/Wp로 기준 대비 +10.0%p 여유 확보'\n"
           "  ✓ '가중평균자본비용 대비 +0.88%p 상회로 가치 창출 확인'\n"
           "  ✓ 'PPA 재협상 또는 CAPEX 3% 절감 필요'\n"
           "  ✓ '경제성 기준 3개 모두 충족, 개발자본 투입 계속 권고'\n"
           "Examples (BAD — do NOT use):\n"
           "  ✗ '~제공한다' (→ '~제공')\n"
           "  ✗ '~확인된다' (→ '~확인됨')\n"
           "  ✗ '~충족한다' (→ '~충족')\n"
           "  ✗ '~하도록 한다' (→ '~권고')\n"
           "  ✗ '~해야 합니다' / '~할 수 있습니다' (too formal/verbose)\n"
           "Maintain consistency — ALL sentences end in the nominal/concise style.\n"
           "\n"
           "CRITICAL — INDUSTRY TERMINOLOGY:\n"
           "  ✗ '증설' (WRONG — means 'capacity expansion')\n"
           "  ✓ 'Augmentation' (English preferred, industry standard)\n"
           "  ✓ '배터리 교체 (용량 유지)' or '배터리 보강 (성능 유지)' (if Korean needed)\n"
           "  Augmentation = replacing/adding cells to MAINTAIN capacity over degradation,\n"
           "  NOT adding new capacity. Never call it '증설'.\n"
           "\n"
           "Only 'verdict' (PROCEED/RECUT/STOP) and 'verdict_color' (green/amber/red) stay English.\n"
           "Financial numbers with units can stay English-style (e.g., '10.38%', '$68.82/MWh').\n"
           "DO NOT mix languages within a single field.\n"
           if payload.get("lang","en")=="kr" else
           "ALL text fields in ENGLISH only. Formal institutional investor tone.\n")
        + "\n"
        "Be direct. Cite specific numbers. No hedging.\n\n"

        "Respond ONLY with valid JSON (no markdown, no code blocks).\n"
        "Required keys:\n"
        "  verdict: \"PROCEED\" | \"RECUT\" | \"STOP\"\n"
        "  verdict_color: \"green\" | \"amber\" | \"red\"\n"
        "  threshold_status: {\n"
        "    margin_ok: bool, margin_gap: str,\n"
        "    irr_ok: bool, irr_gap: str,\n"
        "    wacc_spread_ok: bool, wacc_spread: str  (e.g., '+0.88%p' or '-1.20%p')\n"
        "  }\n"
        "  metrics: ONE compact line, under 120 chars, pipe-delimited.\n"
        "    Example: '199 MWac | 10.4% IRR | $39.8M Margin | $68.8 PPA | $836M CAPEX | 30% ITC'\n"
        "  sensitivity_en: dev margin upside/downside in English with c/Wp numbers\n"
        "  sensitivity_kr: same in Korean (nominal ending style)\n"
        "  thesis: 3-4 sentence economic rationale (경제성 수치 기반 근거)\n"
        "  risks: array of {title, severity: Critical|Watch|OK, detail}\n"
        "    (project-specific only; Safe Harbor/FEOC/BESS ITC are handled separately)\n"
        "  rec: 2-3 sentence actionable recommendation (경제성 관점)\n"
        "All strings double-quoted. No trailing commas. No extra text outside JSON.\n"
        "NOTE: Do NOT include 'dev_ic' field — it has been removed from the schema."
    )

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 4000,
            "messages": [{"role": "user", "content": prompt}]
        },
        timeout=45
    )
    if resp.status_code != 200:
        raise HTTPException(500, f"Claude API 오류: {resp.text[:200]}")

    data = resp.json()
    text = "".join(b.get("text","") for b in data.get("content",[]))

    # JSON 정제 — 코드블록, 줄바꿈, 특수문자 처리
    import re as _re
    clean = text.strip()
    clean = _re.sub(r"```(?:json)?\s*", "", clean).strip()
    clean = clean.strip("`")
    # { ... } 범위만 추출
    start = clean.find("{")
    end   = clean.rfind("}") + 1
    if start >= 0 and end > start:
        clean = clean[start:end]

    # ── 고정 규정 준수 체크리스트 2개 항목을 응답에 주입 ───────
    # AI가 판단하는 risks와 완전히 분리된, 모든 프로젝트 공통 체크리스트
    is_kr = payload.get("lang", "en") == "kr"

    if is_kr:
        compliance_checklist = [
            {
                "title": "ITC BOC(Beginning of Construction) 매칭 확인",
                "severity": "Watch",
                "detail": (
                    "HEUH는 Physical Work Test 방식으로 BOC 요건을 확보하여 관리 중. "
                    "본 프로젝트가 기확보된 BOC pool과 매칭되는지 NTP 전 확인 권고. "
                    "매칭 확보 시 Continuity Safe Harbor에 따라 2030년 말까지 PIS 여유."
                )
            },
            {
                "title": "FEOC 공급망 적격성 검토",
                "severity": "Watch",
                "detail": (
                    "OBBBA에 따라 2026년 착공 프로젝트는 비PFE(중국/러시아/이란/북한 외) "
                    "부품 비중 요건 적용: PV ≥40%, BESS ≥55% (매년 5%p 상향). "
                    "EPC 계약 체결 전 배터리 셀·PV 모듈 원산지 증빙 확보 필요."
                )
            }
        ]
    else:
        compliance_checklist = [
            {
                "title": "ITC BOC Matching Verification",
                "severity": "Watch",
                "detail": (
                    "HEUH has established BOC (Beginning of Construction) for its project pool "
                    "via Physical Work Test, managed by its compliance team. Verify this project "
                    "is matched to the secured BOC pool before NTP. Once matched, Continuity Safe "
                    "Harbor extends PIS to Dec 31, 2030."
                )
            },
            {
                "title": "FEOC Supply Chain Compliance Review",
                "severity": "Watch",
                "detail": (
                    "Under OBBBA, 2026-start projects face non-PFE (China/Russia/Iran/DPRK excluded) "
                    "content thresholds: PV ≥40%, BESS ≥55% (ramping +5%p annually). Verify battery "
                    "cell and PV module country-of-origin documentation before EPC contract."
                )
            }
        ]

    # JSON 파싱해서 risks 배열 앞에 삽입
    try:
        import json as _json
        parsed = _json.loads(clean)
        ai_risks = parsed.get("risks", []) or []
        # 컴플라이언스 체크리스트를 최상단, AI 리스크를 그 뒤에
        parsed["risks"] = compliance_checklist + ai_risks
        # 구분 위해 플래그 추가 (프론트엔드에서 활용 가능)
        parsed["compliance_count"] = len(compliance_checklist)
        clean = _json.dumps(parsed, ensure_ascii=False)
    except Exception as _e:
        # 파싱 실패 시 원본 그대로 반환 (프론트가 처리)
        print(f"[analyze-cf] JSON merge failed: {_e}", flush=True)

    return {"ok": True, "result": clean}


@app.post("/valuation/{project_id}/save")
async def save_valuation_version(
    project_id: str,
    payload: dict,
    user=Depends(get_current_user)
):
    """버전 저장 → 즉시 저장 (승인 flow 제거). 100개 한도."""
    safe_id = project_id.replace("/", "_").replace(".", "_")
    ts = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    payload["uploaded_by"] = user["email"]
    payload["uploaded_at"] = datetime.datetime.now().isoformat()
    # 승인 flow 제거 → 즉시 "saved" 상태
    payload["status"] = "saved"
    payload["requested_by"] = user["email"]
    # "approver" 필드 레거시 호환: 존재하면 "shared_with"로 마이그레이트
    if "approver" in payload and payload["approver"] and "shared_with" not in payload:
        payload["shared_with"] = payload["approver"]

    fb_put(f"valuation/{safe_id}/versions/{ts}", payload)
    fb_put(f"valuation/{safe_id}/latest", payload)

    # 100개 한도 — 초과 시 가장 오래된 것 삭제
    versions = fb_read(f"valuation/{safe_id}/versions") or {}
    keys = sorted(versions.keys())
    if len(keys) > 100:
        for old_key in keys[:len(keys)-100]:
            try:
                requests.delete(
                    f"{FB_URL}/valuation/{safe_id}/versions/{old_key}.json",
                    params=fb_auth_param(),
                    timeout=5
                )
            except Exception:
                pass

    return {"ok": True, "timestamp": ts}


# 레거시 승인/반려 엔드포인트 — 하위호환 유지하되 no-op화 (존재하는 pending 버전 정리용)
@app.post("/valuation/{project_id}/versions/{ts}/approve")
def approve_version(project_id: str, ts: str, user=Depends(require_admin)):
    """[Deprecated] 승인 flow 제거됨. 하위호환용: pending을 saved로 마이그레이트."""
    safe_id = project_id.replace("/", "_").replace(".", "_")
    fb_patch(f"valuation/{safe_id}/versions/{ts}", {
        "status": "saved",
        "approved_by": user["email"],
        "approved_at": datetime.datetime.now().isoformat()
    })
    return {"ok": True}


@app.post("/valuation/{project_id}/versions/{ts}/reject")
def reject_version(project_id: str, ts: str, body: dict = {}, user=Depends(require_admin)):
    """[Deprecated] 승인 flow 제거됨. 하위호환용: 버전 삭제."""
    safe_id = project_id.replace("/", "_").replace(".", "_")
    fb_patch(f"valuation/{safe_id}/versions/{ts}", {
        "status": "rejected",
        "rejected_by": user["email"],
        "rejected_at": datetime.datetime.now().isoformat(),
        "reject_reason": body.get("reason", "")
    })
    return {"ok": True}
